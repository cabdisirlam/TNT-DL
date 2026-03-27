"""
Imprest Surrender AP Loader — Engine
Converts IFMIS AP invoice rows from Excel into IFMIS keystrokes and loads them.

Excel format (Data_Entry sheet):
  Row 1: Title
  Row 2: Column headers
  Row 3: Format hints
  Row 4: Sample row (do not edit)
  Row 5+: Data rows (one row = one AP invoice)

Columns (A–K):
  Supplier_Num, Invoice_Date, Invoice_Num, Invoice_Amount, Description,
  Payment_Method, Terms_Date, GL_Date, Auth_Ref_No, Administrative_Code,
  Distribution_Account

Keystroke template (per row) — 82-cell DL grid (C1–C82):
  Requires "Use Alternate Method for processing Macros" in DL Load Settings.
  \\%2\\{ESC} = Alt+2, then Esc = jump to the Lines block.
  \\%d        = Alt+D = jump to the Distributions block.
  \\{ENTER} = close modal / confirm / dismiss prepayment alert.
  \\{BACKSPACE} = clear pre-filled field.
  Macro flow:
    General fields → \\{ENTER} (dismiss prepayment alert if any) → "Provisional"
    → \\{ENTER} (close modal) → \\%2\\{ESC} (→Lines block)
    → Tab to Amount → enter amount → \\%d (→Distributions block)
    → Tab to Amount → enter amount → GL Date → Dist Account → Save → Down
"""

import time

from PySide6.QtCore import QThread, Signal


# ── Column definitions ────────────────────────────────────────────────────────

COLUMNS = [
    "Supplier_Num",
    "Invoice_Date",
    "Invoice_Num",
    "Invoice_Amount",
    "Description",
    "Payment_Method",
    "Terms_Date",
    "GL_Date",
    "Auth_Ref_No",
    "Administrative_Code",
    "Distribution_Account",
    "Old_Imprest_No",
]

COLUMN_HINTS = {
    "Supplier_Num":        "e.g. 117711",
    "Invoice_Date":        "DD-MMM-YYYY",
    "Invoice_Num":         "e.g. SURR0001",
    "Invoice_Amount":      "e.g. 42,000.00",
    "Description":         "e.g. SURRENDER OF IMPREST",
    "Payment_Method":      "CHECK or ELECTRONIC",
    "Terms_Date":          "DD-MMM-YYYY",
    "GL_Date":             "DD-MMM-YYYY",
    "Auth_Ref_No":         "e.g. CFO",
    "Administrative_Code": "e.g. 5322000201",
    "Distribution_Account":"Full SCOA e.g. 0-5322-0000000000-00001001-...",
    "Old_Imprest_No":      "Imported old imprest no e.g. IMP5711075",
}

COLUMN_SAMPLE = {
    "Supplier_Num":        "117711",
    "Invoice_Date":        "30-JUN-2020",
    "Invoice_Num":         "SURR0001",
    "Invoice_Amount":      "42,000.00",
    "Description":         "SURRENDER OF IMPREST",
    "Payment_Method":      "CHECK",
    "Terms_Date":          "30-JUN-2020",
    "GL_Date":             "30-JUN-2020",
    "Auth_Ref_No":         "CFO",
    "Administrative_Code": "5322000201",
    "Distribution_Account":"0-5322-0000000000-00001001-0000000000-6580101-53100001-000",
    "Old_Imprest_No":      "IMP5711075",
}

# ── Keystroke grid row builder ────────────────────────────────────────────────

_T       = "{Tab}"
_BS      = "\\{BACKSPACE}"
_ENTER   = "\\{ENTER}"
_ALT2ESC = "\\%2\\{ESC}"   # Alt+2 + Esc  → Lines block
_ALTD    = "\\%d"           # Alt+D        → Distributions block
_CTRLS   = "\\^s"           # Ctrl+S       → save
_CTRLF4  = "\\^{F4}"        # Ctrl+F4      → clear record
_ALTKEY  = "\\%"            # Alt alone    → activate menu
_DOWN    = "\\{DOWN}"       # Down arrow
_SHTAB   = "\\+{TAB}"       # Shift+Tab

# ── Imprest: 82-cell DataLoad grid row ────────────────────────────────────────
# Navigation: \%2\{ESC} (Alt+2+Esc) → Lines,  \%d (Alt+D) → Distributions.
# Post-save: \^{F4} + Alt + 4×Down + Enter + 3×Shift+Tab → ready for next row.
# REQUIRES: "Use Alternate Method for processing Macros" ticked in DL settings.
def build_keystroke_row(row: dict) -> list:
    """
    Convert one invoice row-dict into an 82-cell DataLoad grid row.
    Matches the full working macro exactly:
      General fields → Enter (dismiss prepayment alert) → "Provisional"
      → Alt+2+Esc (Lines) → Tab×2 → Amount
      → Alt+D (Distributions) → Tab×2 → Amount → GL_Date → Account
      → Ctrl+S → Ctrl+F4 → Alt → Down×4 → Enter → Shift+Tab×3
    """
    row = _normalize_invoice_row(row)
    sup   = row.get("Supplier_Num", "")
    idate = row.get("Invoice_Date", "")
    inum  = row.get("Invoice_Num", "")
    amt   = row.get("Invoice_Amount", "")
    desc  = row.get("Description", "")
    pmeth = row.get("Payment_Method", "") or "CHECK"
    gldt  = row.get("GL_Date", "")
    auth  = row.get("Auth_Ref_No", "")
    admc  = row.get("Administrative_Code", "")
    dist  = row.get("Distribution_Account", "")

    return [
        _T,           # C1
        _T,           # C2
        _BS,          # C3   clear pre-filled field
        _T,           # C4
        "Standard",   # C5   Invoice Type (fixed)
        _T,           # C6
        _BS,          # C7   clear pre-filled field
        _T,           # C8
        _BS,          # C9   clear pre-filled field
        _T,           # C10
        sup,          # C11  Supplier_Num
        _T,           # C12
        _ENTER,       # C13  dismiss prepayment alert if present (harmless otherwise)
        "Provisional",# C14  Supplier Site / batch type (fixed)
        _T,           # C15
        idate,        # C16  Invoice_Date
        _T,           # C17
        inum,         # C18  Invoice_Num
        _T,           # C19
        _T,           # C20
        amt,          # C21  Invoice_Amount
        _T,           # C22
        _T,           # C23
        _T,           # C24
        _T,           # C25
        _T,           # C26
        _T,           # C27
        _T,           # C28
        desc,         # C29  Description
        _T,           # C30
        _T,           # C31
        _T,           # C32
        "IMMEDIATE",  # C33  Pay Terms (fixed)
        _T,           # C34
        pmeth,        # C35  Payment_Method
        _T,           # C36
        _T,           # C37
        _T,           # C38
        _T,           # C39
        _T,           # C40
        _T,           # C41
        _T,           # C42
        _T,           # C43
        _T,           # C44
        _T,           # C45
        _T,           # C46
        _T,           # C47
        _T,           # C48
        _T,           # C49
        _T,           # C50
        _T,           # C51
        _T,           # C52
        auth,         # C53  Auth_Ref_No
        _T,           # C54
        admc,         # C55  Administrative_Code
        _T,           # C56
        _ENTER,       # C57  \{ENTER} — close modal
        _ALT2ESC,     # C58  \%2\{ESC} — Alt+2+Esc → Lines block
        _T,           # C59
        _T,           # C60
        amt,          # C61  Line_Amount (= Invoice_Amount)
        _T,           # C62
        _ALTD,        # C63  \%d — Alt+D → Distributions block
        _T,           # C64
        _T,           # C65
        amt,          # C66  Dist_Amount (= Invoice_Amount)
        _T,           # C67
        gldt,         # C68  GL_Date
        _T,           # C69
        dist,         # C70  Distribution_Account
        _T,           # C71
        _CTRLS,       # C72  \^s — Ctrl+S save
        _CTRLF4,      # C73  \^{F4} — clear record
        _ALTKEY,      # C74  \% — Alt → activate menu
        _DOWN,        # C75  \{DOWN}
        _DOWN,        # C76  \{DOWN}
        _DOWN,        # C77  \{DOWN}
        _DOWN,        # C78  \{DOWN} — navigate to menu item
        _ENTER,       # C79  \{ENTER} — select menu item
        _SHTAB,       # C80  \+{TAB}
        _SHTAB,       # C81  \+{TAB}
        _SHTAB,       # C82  \+{TAB} — ready for next invoice
    ]


# ── Template action sequence ──────────────────────────────────────────────────
# Action tuples:
#   ("tab",    n)            press Tab n times via SendInput
#   ("key",    name)         press a named key (must be in _SI_VK_MAP)
#   ("hotkey", mods, key)    modifier+key  e.g. (["shift"], "pagedown")
#   ("delay",  ms)           sleep ms milliseconds (interruptible)
#   ("field",  col_name)     inject field value via SendInput unicode
#   ("text",   value)        type a fixed literal string via SendInput unicode

# Single imprest template — mirrors build_keystroke_row exactly.
# Navigation: Alt+2+Esc → Lines,  Alt+D → Distributions.
# Post-save: Ctrl+F4 + Alt + Down×4 + Enter + Shift+Tab×3 → ready for next.
# 500 ms delay after Alt+2+Esc for Lines block to open.
TEMPLATE_ACTIONS = (
    # ── General block ──────────────────────────────────────────────────────
    ("tab",    2),                           # C1–C2
    ("key",    "backspace"),                 # C3   clear pre-filled
    ("tab",    1),                           # C4
    ("text",   "Standard"),                  # C5   Invoice Type
    ("tab",    1),                           # C6
    ("key",    "backspace"),                 # C7   clear pre-filled
    ("tab",    1),                           # C8
    ("key",    "backspace"),                 # C9   clear pre-filled
    ("tab",    1),                           # C10
    ("field",  "Supplier_Num"),              # C11
    ("tab",    1),                           # C12
    ("key",    "enter"),                     # C13  dismiss prepayment alert (if any)
    ("text",   "Provisional"),               # C14  Supplier Site / batch type
    ("tab",    1),                           # C15
    ("field",  "Invoice_Date"),              # C16
    ("tab",    1),                           # C17
    ("field",  "Invoice_Num"),               # C18
    ("tab",    2),                           # C19–C20
    ("field",  "Invoice_Amount"),            # C21
    ("tab",    7),                           # C22–C28
    ("field",  "Description"),               # C29
    ("tab",    3),                           # C30–C32
    ("text",   "IMMEDIATE"),                 # C33  Pay Terms (fixed)
    ("tab",    1),                           # C34
    ("text",   "CHECK"),                     # C35  Payment Method (fixed)
    ("tab",    17),                          # C36–C52
    ("field",  "Auth_Ref_No"),               # C53
    ("tab",    1),                           # C54
    ("field",  "Administrative_Code"),       # C55
    ("tab",    1),                           # C56
    ("key",    "enter"),                     # C57  close modal
    # ── Lines block (Alt+2+Esc) ────────────────────────────────────────────
    ("hotkey", ["alt"], "2"),                # C58a Alt+2 → Lines
    ("key",    "escape"),                    # C58b Esc
    ("delay",  500),                         # wait for Lines block to open
    ("tab",    2),                           # C59–C60
    ("field",  "Invoice_Amount"),            # C61  Line Amount
    ("tab",    1),                           # C62
    # ── Distributions block (Alt+D) ────────────────────────────────────────
    ("hotkey", ["alt"], "d"),                # C63  Alt+D → Distributions
    ("tab",    2),                           # C64–C65
    ("field",  "Invoice_Amount"),            # C66  Dist Amount
    ("tab",    1),                           # C67
    ("field",  "GL_Date"),                   # C68
    ("tab",    1),                           # C69
    ("field",  "Distribution_Account"),      # C70
    ("tab",    1),                           # C71
    # ── Save and advance to next invoice ──────────────────────────────────
    ("hotkey", ["ctrl"], "s"),               # C72  Ctrl+S save
    ("hotkey", ["ctrl"], "f4"),              # C73  Ctrl+F4 clear record
    ("key",    "alt"),                       # C74  Alt → activate menu
    ("key",    "down"),                      # C75  ↓
    ("key",    "down"),                      # C76  ↓
    ("key",    "down"),                      # C77  ↓
    ("key",    "down"),                      # C78  ↓ (4th menu item)
    ("key",    "enter"),                     # C79  select menu item
    ("hotkey", ["shift"], "tab"),            # C80  ⇧Tab
    ("hotkey", ["shift"], "tab"),            # C81  ⇧Tab
    ("hotkey", ["shift"], "tab"),            # C82  ⇧Tab → ready for next invoice
)

# Legacy alternate navigation template kept for reference only.
# Uses Shift+PageDown instead of Alt+2+Esc to jump to the Lines block.
TEMPLATE_ACTIONS_PGDN = (
    # ── General block ──────────────────────────────────────────────────────
    ("tab",    2),
    ("key",    "backspace"),
    ("tab",    1),
    ("text",   "Standard"),
    ("tab",    1),
    ("key",    "backspace"),
    ("tab",    1),
    ("key",    "backspace"),
    ("tab",    1),
    ("field",  "Supplier_Num"),
    ("tab",    1),
    ("key",    "enter"),                     # dismiss prepayment alert (if any)
    ("text",   "Provisional"),
    ("tab",    1),
    ("field",  "Invoice_Date"),
    ("tab",    1),
    ("field",  "Invoice_Num"),
    ("tab",    2),
    ("field",  "Invoice_Amount"),
    ("tab",    7),
    ("field",  "Description"),
    ("tab",    3),
    ("text",   "IMMEDIATE"),
    ("tab",    1),
    ("text",   "CHECK"),
    ("tab",    17),
    ("field",  "Auth_Ref_No"),
    ("tab",    1),
    ("field",  "Administrative_Code"),
    ("tab",    1),
    ("key",    "enter"),                     # close modal
    # ── Lines block (Shift+PageDown) ───────────────────────────────────────
    ("hotkey", ["shift"], "pagedown"),       # Shift+PgDn → Lines block
    ("delay",  500),                         # wait for Lines block to open
    ("tab",    2),
    ("field",  "Invoice_Amount"),            # Line Amount
    ("tab",    1),
    # ── Distributions block (Alt+D) ────────────────────────────────────────
    ("hotkey", ["alt"], "d"),
    ("tab",    2),
    ("field",  "Invoice_Amount"),            # Dist Amount
    ("tab",    1),
    ("field",  "GL_Date"),
    ("tab",    1),
    ("field",  "Distribution_Account"),
    ("tab",    1),
    # ── Save and advance to next invoice ──────────────────────────────────
    ("hotkey", ["ctrl"], "s"),
    ("hotkey", ["ctrl"], "f4"),
    ("key",    "alt"),
    ("key",    "down"),
    ("key",    "down"),
    ("key",    "down"),
    ("key",    "down"),
    ("key",    "enter"),
    ("hotkey", ["shift"], "tab"),
    ("hotkey", ["shift"], "tab"),
    ("hotkey", ["shift"], "tab"),
)


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def build_keystroke_row(row: dict) -> list:
    """Updated 109-cell DataLoad grid row for Imprest surrender application."""
    row = _normalize_invoice_row(row)
    sup = row.get("Supplier_Num", "")
    idate = row.get("Invoice_Date", "")
    inum = row.get("Invoice_Num", "")
    amt = row.get("Invoice_Amount", "")
    apply_amt = (amt or "").replace(",", "")
    desc = row.get("Description", "")
    pmeth = row.get("Payment_Method", "") or "CHECK"
    gldt = row.get("GL_Date", "")
    auth = row.get("Auth_Ref_No", "")
    admc = row.get("Administrative_Code", "")
    dist = row.get("Distribution_Account", "")
    old_imp = row.get("Old_Imprest_No", "")

    return [
        _T, _T, _BS, _T, "Standard", _T, _BS, _T, _BS, _T,
        sup, _T, _ENTER, "Provisional", _T, idate, _T, inum, _T, _T,
        amt, _T, _T, _T, _T, _T, _T, _T, desc, _T,
        _T, _T, "IMMEDIATE", _T, pmeth, _T, _T, _T, _T, _T,
        _T, _T, _T, _T, _T, _T, _T, _T, _T, _T,
        _T, _T, auth, _T, admc, _T, _ENTER, _ALT2ESC, _T, _T,
        apply_amt, _T, _ALTD, _T, _T, apply_amt, _T, gldt, _T, dist,
        _T, _CTRLS, _CTRLF4, "\\%c", "\\%u", "\\%k", "\\%v", "\\{DOWN}", "\\{DOWN}", "\\{ENTER}",
        old_imp, "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", apply_amt, "\\{TAB}",
        "\\{TAB}", "\\{ENTER}", "\\{SPACE}", "\\{TAB}", apply_amt, "\\{TAB}", gldt, "\\{TAB}", _CTRLS, _CTRLF4,
        "\\%", "\\{DOWN}", "\\{DOWN}", "\\{DOWN}", "\\{DOWN}", "\\{ENTER}", "\\+{TAB}", "\\+{TAB}", "\\+{TAB}",
    ]


TEMPLATE_ACTIONS = (
    ("tab", 2),
    ("key", "backspace"),
    ("tab", 1),
    ("text", "Standard"),
    ("tab", 1),
    ("key", "backspace"),
    ("tab", 1),
    ("key", "backspace"),
    ("tab", 1),
    ("field", "Supplier_Num"),
    ("tab", 1),
    ("key", "enter"),
    ("text", "Provisional"),
    ("tab", 1),
    ("field", "Invoice_Date"),
    ("tab", 1),
    ("field", "Invoice_Num"),
    ("tab", 2),
    ("field", "Invoice_Amount"),
    ("tab", 7),
    ("field", "Description"),
    ("tab", 3),
    ("text", "IMMEDIATE"),
    ("tab", 1),
    ("text", "CHECK"),
    ("tab", 17),
    ("field", "Auth_Ref_No"),
    ("tab", 1),
    ("field", "Administrative_Code"),
    ("tab", 1),
    ("key", "enter"),
    ("hotkey", ["alt"], "2"),
    ("key", "escape"),
    ("delay", 500),
    ("tab", 2),
    ("field", "Application_Amount"),
    ("tab", 1),
    ("hotkey", ["alt"], "d"),
    ("tab", 2),
    ("field", "Application_Amount"),
    ("tab", 1),
    ("field", "GL_Date"),
    ("tab", 1),
    ("field", "Distribution_Account"),
    ("tab", 1),
    ("hotkey", ["ctrl"], "s"),
    ("hotkey", ["ctrl"], "f4"),
    ("delay", 500),
    ("hotkey", ["alt"], "c"),
    ("hotkey", ["alt"], "u"),
    ("hotkey", ["alt"], "k"),
    ("hotkey", ["alt"], "v"),
    ("key", "down"),
    ("key", "down"),
    ("key", "enter"),
    ("delay", 500),
    ("field", "Old_Imprest_No"),
    ("tab", 7),
    ("field", "Application_Amount"),
    ("tab", 2),
    ("key", "enter"),
    ("key", "space"),
    ("tab", 1),
    ("field", "Application_Amount"),
    ("tab", 1),
    ("field", "GL_Date"),
    ("tab", 1),
    ("hotkey", ["ctrl"], "s"),
    ("delay", 700),
    ("hotkey", ["ctrl"], "f4"),
    ("delay", 700),
    ("key", "alt"),
    ("delay", 250),
    ("key", "down"),
    ("key", "down"),
    ("key", "down"),
    ("key", "down"),
    ("key", "enter"),
    ("delay", 350),
    ("hotkey", ["shift"], "tab"),
    ("hotkey", ["shift"], "tab"),
    ("hotkey", ["shift"], "tab"),
)


def read_invoice_rows(filepath: str) -> tuple:
    """
    Read invoice rows from the Data_Entry sheet (or first sheet) of filepath.
    Supports .xlsx, .xls, and .csv files.
    Data rows start at row 4 (rows 1–3 are title / headers / hints) for Excel,
    or row 2 for CSV (row 1 = headers).
    Returns (rows: list[dict], error: str).
    """
    import os
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        return _read_invoice_rows_csv(filepath)

    try:
        import openpyxl
    except ImportError:
        return [], "openpyxl is not installed."

    try:
        wb = openpyxl.load_workbook(
            filepath,
            data_only=True,
            read_only=True,
            keep_links=False,
        )
    except Exception as exc:
        return [], f"Cannot open file: {exc}"

    try:
        # Prefer a sheet named Data_Entry (or containing "data"/"entry"), else first sheet
        sheet_name = wb.sheetnames[0]
        for name in wb.sheetnames:
            if any(k in name.lower() for k in ("data", "entry", "invoice")):
                sheet_name = name
                break

        ws = wb[sheet_name]
        rows = []
        for row_values in ws.iter_rows(min_row=5, values_only=True):
            if not row_values or all(v is None or str(v).strip() == "" for v in row_values):
                continue
            row_dict = _normalize_invoice_row({
                col: row_values[i] if i < len(row_values) else ""
                for i, col in enumerate(COLUMNS)
            })
            if not row_dict["Supplier_Num"]:
                continue
            rows.append(row_dict)

        return rows, ""
    finally:
        close_wb = getattr(wb, "close", None)
        if callable(close_wb):
            close_wb()


def _read_invoice_rows_csv(filepath: str) -> tuple:
    """Read invoice rows from a CSV file. First row is headers, data from row 2."""
    import csv
    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            all_rows = list(reader)
    except Exception as exc:
        return [], f"Cannot open CSV file: {exc}"

    if len(all_rows) < 2:
        return [], "CSV file has no data rows."

    rows = []
    for row_values in all_rows[1:]:
        if all(not v.strip() for v in row_values):
            continue
        row_dict = _normalize_invoice_row({
            col: row_values[i] if i < len(row_values) else ""
            for i, col in enumerate(COLUMNS)
        })
        if not row_dict["Supplier_Num"]:
            continue
        rows.append(row_dict)

    return rows, ""


def export_template(filepath: str) -> str:
    """
    Write the Data_Entry Excel template to filepath.
    Returns an error string, or "" on success.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl is not installed."

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data_Entry"

        ncols = len(COLUMNS)

        # ── Shared style helpers ──────────────────────────────────────────────
        def _side(style="thin", color="BFBFBF"):
            return Side(style=style, color=color)

        def _border():
            s = _side()
            return Border(left=s, right=s, top=s, bottom=s)

        def _fill(color):
            return PatternFill(fill_type="solid", fgColor=color)

        BLUE      = "0070C0"   # header bg
        WHITE_TXT = "FFFFFF"
        HINT_BG   = "D6E4F0"   # light blue — hint row
        SAMPLE_BG = "EAF4FB"   # very light blue — sample row
        ROW_ODD   = "F2F2F2"   # alternating grey
        ROW_EVEN  = "FFFFFF"   # alternating white

        # ── Row 1: title bar ─────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=ncols)
        title = ws.cell(row=1, column=1,
                        value="IFMIS AP Invoice Loader  —  NT_DL  |  "
                              "Fill from row 5 only. One row = one invoice.")
        title.font      = Font(bold=True, size=11, color=WHITE_TXT)
        title.fill      = _fill(BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

        # ── Row 2: column headers ─────────────────────────────────────────────
        hdr_border_s = _side("medium", "FFFFFF")
        hdr_border   = Border(left=hdr_border_s, right=hdr_border_s,
                              top=hdr_border_s, bottom=hdr_border_s)
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=ci, value=col.replace("_", " "))
            cell.fill      = _fill(BLUE)
            cell.font      = Font(bold=True, color=WHITE_TXT, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = hdr_border
        ws.row_dimensions[2].height = 28

        # ── Row 3: format hints ───────────────────────────────────────────────
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=3, column=ci, value=COLUMN_HINTS.get(col, ""))
            cell.fill      = _fill(HINT_BG)
            cell.font      = Font(italic=True, color="1F5C8B", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[3].height = 15

        # ── Row 4: sample data row ────────────────────────────────────────────
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=4, column=ci, value=COLUMN_SAMPLE.get(col, ""))
            cell.fill      = _fill(SAMPLE_BG)
            cell.font      = Font(italic=True, color="1F5C8B", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[4].height = 15

        # ── Rows 5–104: data entry rows (alternating, no colour fill) ─────────
        for ri in range(5, 105):
            bg = ROW_ODD if ri % 2 == 1 else ROW_EVEN
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill      = _fill(bg)
                cell.font      = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = _border()
            ws.row_dimensions[ri].height = 15

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [14, 14, 14, 14, 30, 16, 14, 14, 12, 18, 68, 18]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # ── Freeze panes below header rows ────────────────────────────────────
        ws.freeze_panes = "A5"

        wb.save(filepath)
        return ""
    except Exception as exc:
        return f"Export failed: {exc}"


# ── IFMIS export importer ─────────────────────────────────────────────────────

# Columns that cannot be auto-mapped from the IFMIS export — user must fill them.
IFMIS_BLANK_COLS = {"Auth_Ref_No", "Administrative_Code", "Distribution_Account"}

# IFMIS header name fragment → our field name (case-insensitive contains match)
_IFMIS_COL_MAP = {
    "supplier num":    "Supplier_Num",
    "invoice date":    "Invoice_Date",
    "invoice num":     "Invoice_Num",
    "invoice amount":  "Invoice_Amount",
    "description":     "Description",
    "payment method":  "Payment_Method",
    "gl date":         "GL_Date",
    "terms date":      "Terms_Date",
}


def _fmt_ifmis_date(v) -> str:
    """Convert an IFMIS date value (datetime or 'YYYY-MM-DD …') to DD-MMM-YYYY."""
    if v is None:
        return ""
    from datetime import datetime as _dt
    if isinstance(v, _dt):
        return v.strftime("%d-%b-%Y").upper()
    s = str(v).strip()
    if not s or s.lower() == "none":
        return ""
    # "2022-06-08 00:00:00" or "2022-06-08"
    try:
        return _dt.strptime(s[:10], "%Y-%m-%d").strftime("%d-%b-%Y").upper()
    except ValueError:
        return s


def _normalize_invoice_row(row: dict) -> dict:
    """Normalize invoice values before exporting or sending them."""
    normalized = {}
    for col in COLUMNS:
        value = row.get(col, "") if isinstance(row, dict) else ""
        if col in {"Invoice_Date", "Terms_Date", "GL_Date"}:
            normalized[col] = _fmt_ifmis_date(value)
        else:
            normalized[col] = "" if value is None else str(value).strip()
    if not normalized.get("Payment_Method"):
        normalized["Payment_Method"] = "CHECK"
    return normalized


def import_ifmis_export(filepath: str) -> tuple:
    """
    Read an IFMIS-exported AP invoice file and map to our 11-column format.
    Supports .xlsx, .xls, and .csv files.
    Only rows where the 'Type' column equals 'Prepayment' are imported.

    Returns (rows: list[dict], skipped: int, error: str).
      rows    — list of dicts keyed by COLUMNS; blank fields = "".
      skipped — number of non-Prepayment rows ignored.
      error   — non-empty string if the file could not be read.
    """
    import os as _os
    ext = _os.path.splitext(filepath)[1].lower()

    try:
        import openpyxl
    except ImportError:
        return [], 0, "openpyxl is not installed."

    if ext == ".csv":
        import csv
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                for row_vals in csv.reader(f):
                    ws.append(row_vals)
        except Exception as exc:
            return [], 0, f"Cannot open CSV file: {exc}"
    else:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as exc:
            return [], 0, f"Cannot open file: {exc}"
        ws = wb.active

    # ── Build header index (0-based) ──────────────────────────────────────────
    headers = [str(ws.cell(1, c).value or "").strip().lower()
               for c in range(1, ws.max_column + 1)]

    def _find(fragment):
        for i, h in enumerate(headers):
            if fragment in h:
                return i
        return -1

    col_type = _find("type")
    field_cols = {field: _find(frag) for frag, field in _IFMIS_COL_MAP.items()}

    def _val(row_num, col_idx):
        if col_idx < 0:
            return ""
        v = ws.cell(row_num, col_idx + 1).value
        return "" if v is None else str(v).strip()

    def _date_val(row_num, col_idx):
        if col_idx < 0:
            return ""
        return _fmt_ifmis_date(ws.cell(row_num, col_idx + 1).value)

    def _amount_val(row_num, col_idx):
        if col_idx < 0:
            return ""
        v = ws.cell(row_num, col_idx + 1).value
        if v is None:
            return ""
        try:
            f = float(v)
            return f"{f:,.2f}" if f != int(f) else f"{int(f):,}"
        except (TypeError, ValueError):
            return str(v).strip()

    rows = []
    skipped = 0

    for r in range(2, ws.max_row + 1):
        # Filter: Prepayment only
        if col_type >= 0:
            t = _val(r, col_type).lower()
            if t and t != "prepayment":
                skipped += 1
                continue

        supplier = _val(r, field_cols.get("Supplier_Num", -1))
        if not supplier:
            continue   # skip blank rows

        amt_col = field_cols.get("Invoice_Amount", -1)

        row_dict = {
            "Supplier_Num":         supplier,
            "Invoice_Date":         "",   # left blank — user fills the dates
            "Invoice_Num":          _val(r,  field_cols.get("Invoice_Num", -1)),
            "Invoice_Amount":       _amount_val(r, amt_col),
            "Description":          _val(r,  field_cols.get("Description", -1)),
            "Payment_Method":       "CHECK",   # always CHECK for imprest
            "Terms_Date":           "",   # left blank — user fills the dates
            "GL_Date":              "",   # left blank — user fills the dates
            "Auth_Ref_No":          "",   # must be filled by user
            "Administrative_Code":  "",   # must be filled by user
            "Distribution_Account": "",   # must be filled by user
        }
        rows.append(row_dict)

    return rows, skipped, ""


def import_ifmis_export(filepath: str) -> tuple:
    """
    Optimized IFMIS import for large workbooks.

    Returns (rows: list[dict], skipped: int, error: str).
    """
    import os as _os

    ext = _os.path.splitext(filepath)[1].lower()

    try:
        import openpyxl
    except ImportError:
        return [], 0, "openpyxl is not installed."

    if ext == ".csv":
        import csv
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(filepath, newline="", encoding="utf-8-sig") as f:
                for row_vals in csv.reader(f):
                    ws.append(row_vals)
        except Exception as exc:
            return [], 0, f"Cannot open CSV file: {exc}"
    else:
        try:
            wb = openpyxl.load_workbook(
                filepath,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        except Exception as exc:
            return [], 0, f"Cannot open file: {exc}"

    try:
        ws = wb.active
        headers = [
            str(value or "").strip().lower()
            for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ]

        def _find(fragment):
            for i, header in enumerate(headers):
                if fragment in header:
                    return i
            return -1

        col_type = _find("type")
        field_cols = {field: _find(frag) for frag, field in _IFMIS_COL_MAP.items()}

        def _raw_val(row_values, col_idx):
            if col_idx < 0 or col_idx >= len(row_values):
                return None
            return row_values[col_idx]

        def _text_val(row_values, col_idx):
            value = _raw_val(row_values, col_idx)
            return "" if value is None else str(value).strip()

        def _amount_val(row_values, col_idx):
            value = _raw_val(row_values, col_idx)
            if value is None:
                return ""
            try:
                amount = float(value)
                return f"{amount:,.2f}" if amount != int(amount) else f"{int(amount):,}"
            except (TypeError, ValueError):
                return str(value).strip()

        rows = []
        skipped = 0

        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if not row_values or all(v is None or str(v).strip() == "" for v in row_values):
                continue

            if col_type >= 0:
                row_type = _text_val(row_values, col_type).lower()
                if row_type and row_type != "prepayment":
                    skipped += 1
                    continue

            supplier = _text_val(row_values, field_cols.get("Supplier_Num", -1))
            if not supplier:
                continue

            rows.append(
                {
                    "Supplier_Num": supplier,
                    "Invoice_Date": "",
                    "Invoice_Num": _text_val(row_values, field_cols.get("Invoice_Num", -1)),
                    "Invoice_Amount": _amount_val(
                        row_values,
                        field_cols.get("Invoice_Amount", -1),
                    ),
                    "Description": _text_val(row_values, field_cols.get("Description", -1)),
                    "Payment_Method": "CHECK",
                    "Terms_Date": "",
                    "GL_Date": "",
                    "Auth_Ref_No": "",
                    "Administrative_Code": "",
                    "Distribution_Account": "",
                    "Old_Imprest_No": _text_val(row_values, field_cols.get("Invoice_Num", -1)),
                }
            )

        return rows, skipped, ""
    finally:
        close_wb = getattr(wb, "close", None)
        if callable(close_wb):
            close_wb()


def export_prefilled_template(filepath: str, rows: list) -> str:
    """
    Export the 11-column template pre-filled with `rows` data.
    Blank columns (Auth_Ref_No, Administrative_Code, Distribution_Account)
    are highlighted orange to remind the user to fill them in.
    Returns an error string, or "" on success.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl is not installed."

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data_Entry"

        ncols = len(COLUMNS)

        def _side(style="thin", color="BFBFBF"):
            return Side(style=style, color=color)

        def _border():
            s = _side()
            return Border(left=s, right=s, top=s, bottom=s)

        def _fill(color):
            return PatternFill(fill_type="solid", fgColor=color)

        BLUE      = "0070C0"
        WHITE_TXT = "FFFFFF"
        HINT_BG   = "D6E4F0"
        SAMPLE_BG = "EAF4FB"
        ROW_ODD   = "F2F2F2"
        ROW_EVEN  = "FFFFFF"
        NEED_FILL = "FFD966"   # amber — user must fill

        # Indices (0-based) of the blank/must-fill columns
        blank_ci = {ci for ci, col in enumerate(COLUMNS, start=1)
                    if col in IFMIS_BLANK_COLS}

        # ── Row 1: title ───────────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=ncols)
        title = ws.cell(row=1, column=1,
                        value="IFMIS AP Invoice Loader  —  NT_DL  |  "
                              "Fill AMBER cells then save. One row = one invoice.")
        title.font      = Font(bold=True, size=11, color=WHITE_TXT)
        title.fill      = _fill(BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

        # ── Row 2: headers ─────────────────────────────────────────────────────
        hdr_border_s = _side("medium", "FFFFFF")
        hdr_border   = Border(left=hdr_border_s, right=hdr_border_s,
                              top=hdr_border_s, bottom=hdr_border_s)
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=ci, value=col.replace("_", " "))
            cell.fill      = _fill(BLUE)
            cell.font      = Font(bold=True, color=WHITE_TXT, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = hdr_border
        ws.row_dimensions[2].height = 28

        # ── Row 3: hints ───────────────────────────────────────────────────────
        for ci, col in enumerate(COLUMNS, start=1):
            bg = NEED_FILL if ci in blank_ci else HINT_BG
            cell = ws.cell(row=3, column=ci, value=COLUMN_HINTS.get(col, ""))
            cell.fill      = _fill(bg)
            cell.font      = Font(italic=True, color="1F5C8B", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[3].height = 15

        # ── Row 4: sample ──────────────────────────────────────────────────────
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=4, column=ci, value=COLUMN_SAMPLE.get(col, ""))
            cell.fill      = _fill(SAMPLE_BG)
            cell.font      = Font(italic=True, color="1F5C8B", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
        ws.row_dimensions[4].height = 15

        # ── Rows 5+: pre-filled data ───────────────────────────────────────────
        for ri, row_dict in enumerate(rows, start=5):
            bg = ROW_ODD if ri % 2 == 1 else ROW_EVEN
            for ci, col in enumerate(COLUMNS, start=1):
                val = row_dict.get(col, "")
                cell_bg = NEED_FILL if (ci in blank_ci and not val) else bg
                cell = ws.cell(row=ri, column=ci, value=val if val else None)
                cell.fill      = _fill(cell_bg)
                cell.font      = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = _border()
            ws.row_dimensions[ri].height = 15

        # ── Blank rows after data (up to row 104) ─────────────────────────────
        last_data = 4 + len(rows)
        for ri in range(last_data + 1, max(last_data + 2, 105)):
            bg = ROW_ODD if ri % 2 == 1 else ROW_EVEN
            for ci in range(1, ncols + 1):
                cell_bg = NEED_FILL if ci in blank_ci else bg
                cell = ws.cell(row=ri, column=ci)
                cell.fill      = _fill(cell_bg)
                cell.font      = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = _border()
            ws.row_dimensions[ri].height = 15

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [14, 14, 14, 14, 30, 16, 14, 14, 12, 18, 68, 18]
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A5"
        wb.save(filepath)
        return ""
    except Exception as exc:
        return f"Export failed: {exc}"


# ── DL keystroke export ───────────────────────────────────────────────────────

def _build_dl_keystroke_row(row: dict) -> list:
    """
    Build an 82-cell DataLoad-format (backslash-macro) row for direct use in DataLoad.
    Mirrors build_keystroke_row but uses \\{TAB} instead of {Tab} for DL compatibility.
    """
    _DT  = "\\{TAB}"
    _BS  = "\\{BACKSPACE}"
    _ENT = "\\{ENTER}"
    _A2E = "\\%2\\{ESC}"
    _AD  = "\\%d"
    _CS  = "\\^s"
    _CF4 = "\\^{F4}"
    _ALT = "\\%"
    _DN  = "\\{DOWN}"
    _SHT = "\\+{TAB}"

    row = _normalize_invoice_row(row)
    sup   = row.get("Supplier_Num", "")
    idate = row.get("Invoice_Date", "")
    inum  = row.get("Invoice_Num", "")
    amt   = row.get("Invoice_Amount", "")
    desc  = row.get("Description", "")
    pmeth = row.get("Payment_Method", "") or "CHECK"
    gldt  = row.get("GL_Date", "")
    auth  = row.get("Auth_Ref_No", "")
    admc  = row.get("Administrative_Code", "")
    dist  = row.get("Distribution_Account", "")

    return [
        # C1–C10
        _DT, _DT, _BS, _DT, "Standard", _DT, _BS, _DT, _BS, _DT,
        # C11–C20
        sup, _DT, _ENT, "Provisional", _DT, idate, _DT, inum, _DT, _DT,
        # C21–C30
        amt, _DT, _DT, _DT, _DT, _DT, _DT, _DT, desc, _DT,
        # C31–C40
        _DT, _DT, "IMMEDIATE", _DT, pmeth, _DT, _DT, _DT, _DT, _DT,
        # C41–C50
        _DT, _DT, _DT, _DT, _DT, _DT, _DT, _DT, _DT, _DT,
        # C51–C60
        _DT, _DT, auth, _DT, admc, _DT, _ENT, _A2E, _DT, _DT,
        # C61–C70
        amt, _DT, _AD, _DT, _DT, amt, _DT, gldt, _DT, dist,
        # C71–C82
        _DT, _CS, _CF4, _ALT, _DN, _DN, _DN, _DN, _ENT, _SHT, _SHT, _SHT,
    ]


def _build_dl_keystroke_row(row: dict) -> list:
    """Updated 109-cell backslash-macro row for DataLoad export."""
    row = _normalize_invoice_row(row)
    sup = row.get("Supplier_Num", "")
    idate = row.get("Invoice_Date", "")
    inum = row.get("Invoice_Num", "")
    amt = row.get("Invoice_Amount", "")
    apply_amt = (amt or "").replace(",", "")
    desc = row.get("Description", "")
    pmeth = row.get("Payment_Method", "") or "CHECK"
    gldt = row.get("GL_Date", "")
    auth = row.get("Auth_Ref_No", "")
    admc = row.get("Administrative_Code", "")
    dist = row.get("Distribution_Account", "")
    old_imp = row.get("Old_Imprest_No", "")

    return [
        "\\{TAB}", "\\{TAB}", "\\{BACKSPACE}", "\\{TAB}", "Standard", "\\{TAB}", "\\{BACKSPACE}", "\\{TAB}", "\\{BACKSPACE}", "\\{TAB}",
        sup, "\\{TAB}", "\\{ENTER}", "Provisional", "\\{TAB}", idate, "\\{TAB}", inum, "\\{TAB}", "\\{TAB}",
        amt, "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", desc, "\\{TAB}",
        "\\{TAB}", "\\{TAB}", "IMMEDIATE", "\\{TAB}", pmeth, "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}",
        "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}",
        "\\{TAB}", "\\{TAB}", auth, "\\{TAB}", admc, "\\{TAB}", "\\{ENTER}", "\\%2\\{ESC}", "\\{TAB}", "\\{TAB}",
        apply_amt, "\\{TAB}", "\\%d", "\\{TAB}", "\\{TAB}", apply_amt, "\\{TAB}", gldt, "\\{TAB}", dist,
        "\\{TAB}", "\\^s", "\\^{F4}", "\\%c", "\\%u", "\\%k", "\\%v", "\\{DOWN}", "\\{DOWN}", "\\{ENTER}",
        old_imp, "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", "\\{TAB}", apply_amt, "\\{TAB}",
        "\\{TAB}", "\\{ENTER}", "\\{SPACE}", "\\{TAB}", apply_amt, "\\{TAB}", gldt, "\\{TAB}", "\\^s", "\\^{F4}",
        "\\%", "\\{DOWN}", "\\{DOWN}", "\\{DOWN}", "\\{DOWN}", "\\{ENTER}", "\\+{TAB}", "\\+{TAB}", "\\+{TAB}",
    ]


def export_keystroke_file(filepath: str, rows: list) -> str:
    """
    Export an 82-column DataLoad keystroke grid to filepath.
    Load directly in DataLoad (Per Cell mode, 'Use Alternate Method' ticked)
    as a fallback if SendInput mode fails.
    Returns an error string, or "" on success.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl is not installed."

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DL_Keystrokes"
        _write_keystroke_sheet(ws, rows)
        wb.save(filepath)
        return ""

        BLUE      = "0070C0"
        WHITE_TXT = "FFFFFF"
        GREY_BG   = "F2F2F2"
        ncols     = 82

        # ── Row 1: title ──────────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title = ws.cell(
            row=1, column=1,
            value="NT_DL Imprest Surrender — DataLoad Keystroke Fallback  "
                  "| Load in Per Cell mode  |  'Use Alternate Method' must be ticked in DL settings")
        title.font      = Font(bold=True, size=10, color=WHITE_TXT)
        title.fill      = PatternFill(fill_type="solid", fgColor=BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

        # ── Row 2: column headers C1–C82 ─────────────────────────────────────
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=2, column=ci, value=f"C{ci}")
            cell.font      = Font(bold=True, color=WHITE_TXT, size=9)
            cell.fill      = PatternFill(fill_type="solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 15

        # ── Rows 3+: one DL keystroke row per invoice ─────────────────────────
        for ri, row_dict in enumerate(rows, start=3):
            bg = GREY_BG if ri % 2 == 1 else "FFFFFF"
            ks_row = _build_dl_keystroke_row(row_dict)
            for ci, val in enumerate(ks_row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = Font(size=9)
                cell.fill      = PatternFill(fill_type="solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left")
            ws.row_dimensions[ri].height = 14

        # ── Column widths: narrow for macro cols, wider for data cols ─────────
        # Data is at C11 (sup), C16 (idate), C18 (inum), C21 (amt),
        # C29 (desc), C35 (pmeth), C53 (auth), C55 (admc), C61 (amt),
        # C66 (amt), C68 (gldt), C70 (dist)
        data_cols = {11: 12, 16: 12, 18: 12, 21: 12, 29: 28,
                     35: 12, 53: 8, 55: 14, 61: 10, 66: 10, 68: 12, 70: 52}
        for ci in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                data_cols.get(ci, 10))

        ws.freeze_panes = "A3"
        wb.save(filepath)
        return ""
    except Exception as exc:
        return f"Export failed: {exc}"


# ── Helper ────────────────────────────────────────────────────────────────────

def export_keystroke_sheet_to_workbook(source_path: str, save_path: str, rows: list) -> str:
    """
    Save a copy of the completed workbook with a DL_Keystrokes sheet added.
    Replaces any existing DL_Keystrokes sheet in the saved copy.
    Returns an error string, or "" on success.
    """
    try:
        import openpyxl
    except ImportError:
        return "openpyxl is not installed."

    wb = None
    try:
        keep_vba = source_path.lower().endswith(".xlsm") or save_path.lower().endswith(".xlsm")
        wb = openpyxl.load_workbook(
            source_path,
            keep_vba=keep_vba,
            keep_links=False,
        )
        if "DL_Keystrokes" in wb.sheetnames:
            del wb["DL_Keystrokes"]
        ws = wb.create_sheet("DL_Keystrokes")
        _write_keystroke_sheet(ws, rows)
        wb.save(save_path)
        return ""
    except Exception as exc:
        return f"Export failed: {exc}"
    finally:
        if wb is not None:
            close_wb = getattr(wb, "close", None)
            if callable(close_wb):
                close_wb()


def _write_keystroke_sheet(ws, rows: list) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    blue = "0070C0"
    white_text = "FFFFFF"
    grey_bg = "F2F2F2"
    ncols = 109
    title_font = Font(bold=True, size=10, color=white_text)
    header_font = Font(bold=True, color=white_text, size=9)
    body_font = Font(size=9)
    blue_fill = PatternFill(fill_type="solid", fgColor=blue)
    odd_fill = PatternFill(fill_type="solid", fgColor=grey_bg)
    even_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    title_alignment = Alignment(horizontal="center", vertical="center")
    header_alignment = Alignment(horizontal="center")
    body_alignment = Alignment(horizontal="left")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(
        row=1,
        column=1,
        value=(
            "NT_DL Imprest Surrender - DataLoad Keystroke Fallback  "
            "| Load in Per Cell mode  |  'Use Alternate Method' must be ticked in DL settings"
        ),
    )
    title.font = title_font
    title.fill = blue_fill
    title.alignment = title_alignment
    ws.row_dimensions[1].height = 18

    for ci in range(1, ncols + 1):
        cell = ws.cell(row=2, column=ci, value=f"C{ci}")
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = header_alignment
    ws.row_dimensions[2].height = 15

    for ri, row_dict in enumerate(rows, start=3):
        fill = odd_fill if ri % 2 == 1 else even_fill
        ks_row = _build_dl_keystroke_row(row_dict)
        for ci, val in enumerate(ks_row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = body_alignment
        ws.row_dimensions[ri].height = 14

    data_cols = {
        11: 12,
        16: 12,
        18: 12,
        21: 12,
        29: 28,
        35: 12,
        53: 8,
        55: 14,
        61: 10,
        66: 10,
        68: 12,
        70: 52,
        81: 14,
        89: 10,
        95: 10,
        97: 12,
    }
    for ci in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = data_cols.get(ci, 10)

    ws.freeze_panes = "A3"


def build_row_summary(row: dict) -> str:
    """One-line human-readable summary of an invoice row."""
    desc = (row.get("Description") or "")[:32]
    return (f"{row.get('Invoice_Num', '?')} | "
            f"Supplier {row.get('Supplier_Num', '?')} | "
            f"Amt {row.get('Invoice_Amount', '?')} | {desc}")


_INTER_ACTION_DELAY = 0.2   # Default non-fast-send delay between actions


def _mid_row_popup_check(sender, popup_fn) -> bool:
    """
    Quick popup check for use inside execute_row_for_loader.
    Returns True to continue, False to abort the row.
    """
    if popup_fn is None:
        return True
    from kdl.window.window_manager import WindowManager
    popup = WindowManager.detect_blocking_popup(
        sender.target_hwnd, sender.target_title)
    if popup:
        return popup_fn(popup)
    return True


def execute_row_for_loader(sender, row_dict: dict, is_stop_requested,
                           actions=None, popup_fn=None,
                           inter_action_delay=None, is_last_row=False) -> bool:
    """
    Execute the AP invoice template for one row using an existing DataSender.
    Used by the main LoaderThread when load_mode is 'imprest_surrender'.
      sender              – a configured DataSender instance (use_fast_send=True)
      row_dict            – {col_name: value} for the 11 AP invoice columns
      is_stop_requested   – callable() -> bool
      actions             – action tuple sequence (default: TEMPLATE_ACTIONS)
      popup_fn            – optional callable(popup_title: str) -> bool
                            called when a blocking popup is detected mid-row
      inter_action_delay  – delay in seconds between keystrokes; defaults to
                            _INTER_ACTION_DELAY (0.2) when None
      is_last_row         – when True, adds a 500 ms settle delay before the
                            Ctrl+S save to ensure the last row is fully saved
    """
    if actions is None:
        actions = TEMPLATE_ACTIONS
    if inter_action_delay is None:
        inter_action_delay = _INTER_ACTION_DELAY
    row_dict = _normalize_invoice_row(row_dict)
    row_dict["Application_Amount"] = (row_dict.get("Invoice_Amount", "") or "").replace(",", "")

    from kdl.engine.data_sender import _SI_VK_MAP
    vk_tab = _SI_VK_MAP.get("tab", 0x09)

    def _wait_after_action(delay_seconds: float) -> bool:
        if sender.load_control:
            return sender._wait_for_ready()
        if not sender._sleep_interruptible(delay_seconds):
            sender.last_error = "imprest action interrupted"
            return False
        return sender._wait_if_hourglass()

    for action in actions:
        if is_stop_requested():
            return False

        kind = action[0]

        if kind == "tab":
            for _ in range(action[1]):
                if not sender._si_send_vk(vk_tab):
                    return False
                if not _wait_after_action(inter_action_delay):
                    return False
            # NOTE: no popup check here.  When a Supplier Site (or similar) LOV
            # opens after a Tab, the *next* action in the template types the
            # required value directly into the LOV Find field.  The popup check
            # that fires after that text/field action then dismisses the LOV with
            # Down+Enter so the correct entry is accepted.  Dismissing here (before
            # the value is typed) would select the wrong default entry.

        elif kind == "key":
            vk = _SI_VK_MAP.get(action[1], 0)
            if vk and not sender._si_send_vk(vk):
                return False
            if not _wait_after_action(inter_action_delay):
                return False

        elif kind == "hotkey":
            # On the last row, add a 500 ms settle before saving to ensure
            # the row is written before the macro moves on.
            if is_last_row and action[1] == ["ctrl"] and action[2] == "s":
                if not _wait_after_action(0.5):
                    return False
            if not sender._si_send_hotkey(action[1], action[2]):
                return False
            if not _wait_after_action(inter_action_delay):
                return False

        elif kind == "delay":
            if not sender._sleep_interruptible(action[1] / 1000.0):
                sender.last_error = "imprest delay interrupted"
                return False

        elif kind == "field":
            value = (row_dict.get(action[1]) or "").strip()
            if value and not sender._si_send_unicode(value):
                return False
            if not _wait_after_action(inter_action_delay):
                return False
            if not _mid_row_popup_check(sender, popup_fn):
                return False

        elif kind == "text":
            if action[1] and not sender._si_send_unicode(action[1]):
                return False
            if not _wait_after_action(inter_action_delay):
                return False
            if not _mid_row_popup_check(sender, popup_fn):
                return False

    return True


# ── Background loader thread ──────────────────────────────────────────────────

class ImprestSurrenderThread(QThread):
    """Sends invoice keystrokes to IFMIS for each row using SendInput."""

    progress       = Signal(int, int)   # (current_row, total_rows)
    status_update  = Signal(str)        # status text
    finished_signal = Signal(int, str)  # (rows_loaded, message)
    stopped_signal  = Signal()

    def __init__(self, rows: list, target_hwnd, target_title: str, parent=None):
        super().__init__(parent)
        self._rows         = rows
        self._target_hwnd  = target_hwnd
        self._target_title = target_title
        self._stop         = False

    def request_stop(self):
        self._stop = True

    # ── run ──────────────────────────────────────────────────────────────────

    def run(self):
        from kdl.engine.data_sender import DataSender, _SI_VK_MAP

        sender = DataSender()
        sender.use_fast_send = True
        sender.set_target(self._target_hwnd, self._target_title)
        sender.set_stop_checker(lambda: self._stop)

        if not sender.activate_target():
            self.finished_signal.emit(
                0, f"Cannot activate target window: {self._target_title}")
            return

        total  = len(self._rows)
        loaded = 0

        for i, row in enumerate(self._rows):
            if self._stop:
                self.stopped_signal.emit()
                return

            inv = row.get("Invoice_Num") or f"row {i + 1}"
            self.status_update.emit(f"Loading {i + 1}/{total}: {inv} …")

            ok = execute_row_for_loader(
                sender,
                row,
                lambda: self._stop,
                is_last_row=(i == total - 1),
            )
            if not ok:
                if self._stop:
                    self.stopped_signal.emit()
                    return
                self.finished_signal.emit(
                    loaded,
                    f"Error at row {i + 1} ({inv}): "
                    f"{sender.last_error or 'unknown error'}")
                return

            loaded += 1
            self.progress.emit(i + 1, total)

        self.finished_signal.emit(
            loaded, f"Done — {loaded} invoice(s) loaded into IFMIS.")

    # ── row executor ─────────────────────────────────────────────────────────

    def _execute_row(self, sender, row: dict, vk_map: dict) -> bool:
        vk_tab = vk_map.get("tab", 0x09)

        for action in TEMPLATE_ACTIONS:
            if self._stop:
                return False

            kind = action[0]

            if kind == "tab":
                for _ in range(action[1]):
                    if not sender._si_send_vk(vk_tab):
                        return False
                    time.sleep(_INTER_ACTION_DELAY)

            elif kind == "key":
                vk = vk_map.get(action[1], 0)
                if vk and not sender._si_send_vk(vk):
                    return False
                time.sleep(_INTER_ACTION_DELAY)

            elif kind == "hotkey":
                if not sender._si_send_hotkey(action[1], action[2]):
                    return False
                time.sleep(_INTER_ACTION_DELAY)

            elif kind == "delay":
                deadline = time.monotonic() + action[1] / 1000.0
                while time.monotonic() < deadline:
                    if self._stop:
                        return False
                    time.sleep(0.05)

            elif kind == "field":
                value = (row.get(action[1]) or "").strip()
                if value and not sender._si_send_unicode(value):
                    return False
                time.sleep(_INTER_ACTION_DELAY)

            elif kind == "text":
                if action[1] and not sender._si_send_unicode(action[1]):
                    return False
                time.sleep(_INTER_ACTION_DELAY)

        return True
