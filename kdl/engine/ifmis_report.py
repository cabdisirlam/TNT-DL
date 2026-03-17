"""
GOK IFMIS Notes to Financial Statements Engine (v8)
Pure Python / openpyxl implementation of the VBA macro.

Reads an IFMIS Notes worksheet (openpyxl Worksheet) and produces an
openpyxl Workbook with 5 sheets:
  Notes · Performance · Position · Net Assets · Cash Flow

All values in the statement sheets are cross-sheet Excel formulas that
reference the Notes sheet totals so the workbook stays live-linked.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Sheet names ────────────────────────────────────────────────────────────
SH_NOTES = "Notes"
SH_PERF  = "Performance"
SH_NA    = "Net Assets"
SH_POS   = "Position"
SH_CF    = "Cash Flow"

# ── Formatting constants ───────────────────────────────────────────────────
NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'
TNR     = "Times New Roman"
_BLUE   = "0070C0"
_WHITE  = "FFFFFF"
_GRAY   = "F2F2F2"
_GREEN  = "008000"
_RED_C  = "FF0000"


# ── Data classes ──────────────────────────────────────────────────────────
@dataclass
class NoteItem:
    desc: str = ""
    code: str = ""
    cur:  float = 0.0
    prev: float = 0.0


@dataclass
class ParsedNote:
    key:        str   = ""
    title:      str   = ""
    items:      list  = field(default_factory=list)
    total_cur:  float = 0.0
    total_prev: float = 0.0
    seq_num:    object = 0   # int or str e.g. "15(a)"


@dataclass
class ReportResult:
    success:  bool            = False
    message:  str             = ""
    workbook: Optional[object] = None   # openpyxl.Workbook when success=True


# ── Public entry point ─────────────────────────────────────────────────────
def generate_ifmis_report(source_ws) -> ReportResult:
    """
    Generate financial statements from a parsed IFMIS Notes worksheet.

    Args:
        source_ws: an openpyxl Worksheet containing the Notes data.

    Returns:
        ReportResult with success flag, message, and workbook.
    """
    if not _HAS_OPENPYXL:
        return ReportResult(
            False,
            "openpyxl is not installed.\n\nRun:  pip install openpyxl"
        )
    try:
        engine = _IFMISEngine(source_ws)
        engine.parse()
        wb = engine.build()
        return ReportResult(True, "Financial statements generated successfully.", wb)
    except Exception as exc:
        import traceback
        return ReportResult(False, f"Generation failed: {exc}\n\n{traceback.format_exc()}")


# ── Engine ─────────────────────────────────────────────────────────────────
class _IFMISEngine:
    def __init__(self, source_ws):
        self._ws     = source_ws
        self._notes: list[ParsedNote] = []
        self._entity      = ""
        self._cur_period  = ""
        self._prev_period = ""

        # Aggregates computed in _compute_aggregates
        self._surplus_cur   = 0.0
        self._surplus_prev  = 0.0
        self._net_assets_cur  = 0.0
        self._net_assets_prev = 0.0
        self._na_prior_opening = 0.0
        self._na_close_prev    = 0.0
        self._na_open_cur      = 0.0
        self._na_close_cur     = 0.0
        self._rte_cur          = 0.0
        # Revenue subtotals
        self._rev_ne_cur  = 0.0
        self._rev_ne_prev = 0.0
        self._rev_ex_cur  = 0.0
        self._rev_ex_prev = 0.0
        self._total_rev_cur  = 0.0
        self._total_rev_prev = 0.0
        self._total_exp_cur  = 0.0
        self._total_exp_prev = 0.0
        # Asset/liability subtotals
        self._total_ca_cur   = 0.0
        self._total_ca_prev  = 0.0
        self._total_nca_cur  = 0.0
        self._total_nca_prev = 0.0
        self._total_assets_cur   = 0.0
        self._total_assets_prev  = 0.0
        self._total_liab_cur  = 0.0
        self._total_liab_prev = 0.0

        # ── Linking maps (populated during build) ──────────────────────
        # Notes sheet: note_key → row number of that note's TOTAL line
        self._note_rows: dict[str, int] = {}

        # Performance sheet key rows (for cross-sheet references)
        self._perf_rows: dict[str, int] = {}

        # Position sheet key rows
        self._pos_rows: dict[str, int] = {}

        # Net Assets sheet key rows
        self._na_rows: dict[str, int] = {}

    # ── Parse ──────────────────────────────────────────────────────────────
    def parse(self):
        self._parse_header()
        self._parse_notes()
        self._combine_two("6",   "7",   "6_7",  "Transfers from Domestic and Foreign Partners")
        self._combine_two("22A", "22B", "CASH", "Cash and Cash Equivalents")
        self._make_ppe_cumulative()
        self._make_wc_note()
        self._assign_seq_nums()
        self._compute_aggregates()

    def _parse_header(self):
        for r in range(1, 11):
            raw = self._ws.cell(row=r, column=2).value
            t   = str(raw or "").replace("\xa0", " ")
            if "Entity:" in t:
                self._entity = t[t.index("Entity:") + 7:].strip()
            elif "Current Period:" in t:
                self._cur_period = t[t.index("Current Period:") + 15:].strip()
            elif "Compare With:" in t:
                self._prev_period = t[t.index("Compare With:") + 13:].strip()

    def _parse_notes(self):
        max_row  = self._ws.max_row or 0
        in_items = False
        ci       = -1
        for r in range(1, max_row + 1):
            raw = self._ws.cell(row=r, column=1).value
            c1  = str(raw or "").replace("\xa0", " ").strip()
            if not c1:
                continue

            nk, nt = _is_note_title(c1)
            if nk:
                self._notes.append(ParsedNote(key=nk, title=nt))
                ci       = len(self._notes) - 1
                in_items = False
                continue

            if c1 == "Item Description":
                in_items = True
                continue
            if c1 == "Kshs":
                continue
            if c1 == "TOTAL" and ci >= 0:
                self._notes[ci].total_cur  = _to_float(self._ws.cell(row=r, column=3).value)
                self._notes[ci].total_prev = _to_float(self._ws.cell(row=r, column=4).value)
                in_items = False
                continue
            if in_items and ci >= 0:
                cv = _to_float(self._ws.cell(row=r, column=3).value)
                pv = _to_float(self._ws.cell(row=r, column=4).value)
                if cv != 0 or pv != 0:
                    code = str(self._ws.cell(row=r, column=2).value or "")
                    self._notes[ci].items.append(NoteItem(desc=c1, code=code, cur=cv, prev=pv))

    def _combine_two(self, k1: str, k2: str, nk: str, nt: str):
        n1 = self._find(k1)
        n2 = self._find(k2)
        if n1 is None and n2 is None:
            return
        merged = ParsedNote(key=nk, title=nt)
        if n1:
            merged.items.extend(n1.items)
            merged.total_cur  += n1.total_cur
            merged.total_prev += n1.total_prev
            n1.key = f"SKIP_{k1}"
        if n2:
            merged.items.extend(n2.items)
            merged.total_cur  += n2.total_cur
            merged.total_prev += n2.total_prev
            n2.key = f"SKIP_{k2}"
        self._notes.append(merged)

    def _make_ppe_cumulative(self):
        n = self._find("18")
        if n is None or not n.items:
            return
        opening = NoteItem(
            desc="Opening Balance (Prior Period PPE)",
            code="",
            cur=n.total_prev,
            prev=0.0,
        )
        n.items.insert(0, opening)
        n.total_cur = n.total_prev + n.total_cur
        n.title = "Property, Plant and Equipment"

    def _make_wc_note(self):
        n23 = self._find("23")
        n24 = self._find("24")
        wc  = ParsedNote(key="WC", title="Changes in Working Capital")
        if n23 and (n23.total_cur != 0 or n23.total_prev != 0):
            item = NoteItem(
                desc="Movement in Receivables (Imprest & Clearance)",
                code="Note 23",
                cur=n23.total_prev - n23.total_cur,
                prev=0.0,
            )
            wc.items.append(item)
            wc.total_cur += item.cur
        if n24 and (n24.total_cur != 0 or n24.total_prev != 0):
            item = NoteItem(
                desc="Movement in Trade and Other Payables",
                code="Note 24",
                cur=n24.total_cur - n24.total_prev,
                prev=0.0,
            )
            wc.items.append(item)
            wc.total_cur += item.cur
        self._notes.append(wc)

    def _assign_seq_nums(self):
        order = [
            "4", "6_7", "3", "1", "2", "9", "8", "11",
            "12", "13", "14", "15", "16", "17", "19", "21",
            "20", "CASH", "23", "18", "24", "26", "WC",
        ]
        sq    = 6
        t_num = 0
        for k in order:
            n = self._find(k)
            if n is None:
                continue
            if not (n.total_cur != 0 or n.total_prev != 0):
                continue
            if not n.items:
                continue
            if k == "WC":
                n.seq_num = f"{t_num}(a)" if t_num > 0 else sq
                if not t_num:
                    sq += 1
            else:
                n.seq_num = sq
                if k == "15":
                    t_num = sq
                sq += 1

    def _compute_aggregates(self):
        # Revenue
        rev_ne_keys = ["4", "6_7", "3", "1", "2", "9", "8"]
        rev_ne_cur  = sum(self._nc(k) for k in rev_ne_keys)
        rev_ne_prev = sum(self._np(k) for k in rev_ne_keys)
        rev_ex_cur  = self._nc("11")
        rev_ex_prev = self._np("11")
        self._rev_ne_cur  = rev_ne_cur
        self._rev_ne_prev = rev_ne_prev
        self._rev_ex_cur  = rev_ex_cur
        self._rev_ex_prev = rev_ex_prev
        self._total_rev_cur  = rev_ne_cur + rev_ex_cur
        self._total_rev_prev = rev_ne_prev + rev_ex_prev

        # Expenses
        exp_keys = ["12", "13", "14", "15", "16", "17", "19", "21"]
        self._total_exp_cur  = sum(self._nc(k) for k in exp_keys)
        self._total_exp_prev = sum(self._np(k) for k in exp_keys)

        # Surplus / Deficit
        self._surplus_cur  = self._total_rev_cur  - self._total_exp_cur
        self._surplus_prev = self._total_rev_prev - self._total_exp_prev

        # Assets
        self._total_ca_cur   = self._nc("CASH") + self._nc("23")
        self._total_ca_prev  = self._np("CASH") + self._np("23")
        self._total_nca_cur  = self._nc("18")
        self._total_nca_prev = self._np("18")
        self._total_assets_cur  = self._total_ca_cur  + self._total_nca_cur
        self._total_assets_prev = self._total_ca_prev + self._total_nca_prev

        # Liabilities
        self._total_liab_cur  = self._nc("24")
        self._total_liab_prev = self._np("24")

        # Net Assets
        self._net_assets_cur  = self._total_assets_cur  - self._total_liab_cur
        self._net_assets_prev = self._total_assets_prev - self._total_liab_prev

        # Return to Exchequer (Note 26, current period only)
        self._rte_cur = self._nc("26")

        # Net Assets statement values
        self._na_prior_opening = self._net_assets_prev - self._surplus_prev
        self._na_close_prev    = self._net_assets_prev
        self._na_open_cur      = self._net_assets_prev   # = prior closing
        self._na_close_cur     = self._na_open_cur + self._surplus_cur - self._rte_cur

    # ── Build workbook ─────────────────────────────────────────────────────
    def build(self) -> object:
        wb = Workbook()
        wb.remove(wb.active)
        self._build_notes_sheet(wb)    # Must be first — populates _note_rows
        self._build_perf_sheet(wb)     # Populates _perf_rows
        self._build_pos_sheet(wb)      # Populates _pos_rows
        self._build_na_sheet(wb)       # Populates _na_rows; back-fills Position
        self._build_cf_sheet(wb)
        return wb

    # ── Notes sheet ────────────────────────────────────────────────────────
    def _build_notes_sheet(self, wb):
        ws = wb.create_sheet(SH_NOTES)
        r  = _write_title(ws, 1, 4, "NOTES TO THE FINANCIAL STATEMENTS",
                          self._entity, self._cur_period)
        r += 1

        order = [
            "4", "6_7", "3", "1", "2", "9", "8", "11",
            "12", "13", "14", "15", "WC", "16", "17", "19", "21",
            "20", "CASH", "23", "18", "24", "26",
        ]
        for k in order:
            n = self._find(k)
            if n is None or not n.seq_num or not n.items:
                continue

            # Note title (merged A:D)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=f"Note {n.seq_num}. {n.title}")
            c.font = Font(name=TNR, size=12, bold=True)
            r += 1

            # Header row 1
            for col, val in enumerate(
                ["Item Description", "Item Code", "Current Period", "Previous Period"], 1
            ):
                _blue_hdr(ws.cell(row=r, column=col, value=val))
            r += 1

            # Header row 2 (Kshs)
            for col in range(1, 5):
                val = "Kshs" if col in (3, 4) else None
                _blue_hdr(ws.cell(row=r, column=col, value=val))
            r += 1

            # Items
            for i, item in enumerate(n.items):
                ws.cell(row=r, column=1, value=item.desc).font = Font(name=TNR, size=11)
                ws.cell(row=r, column=2, value=item.code).font = Font(name=TNR, size=11)
                _num(ws.cell(row=r, column=3), item.cur)
                _num(ws.cell(row=r, column=4), item.prev)
                if i % 2 == 0:
                    gf = PatternFill(start_color=_GRAY, end_color=_GRAY, fill_type="solid")
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = gf
                r += 1

            # Total row — record this row for cross-sheet formula linking
            tc = sum(it.cur  for it in n.items)
            tp = sum(it.prev for it in n.items)
            ws.cell(row=r, column=1, value="TOTAL").font = Font(name=TNR, size=11, bold=True)
            _num(ws.cell(row=r, column=3), tc, bold=True)
            _num(ws.cell(row=r, column=4), tp, bold=True)
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = Border(bottom=Side(style="thin"))
            self._note_rows[k] = r   # ← record for cross-sheet linking
            r += 2

        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22

    # ── Performance sheet ──────────────────────────────────────────────────
    def _build_perf_sheet(self, wb):
        ws = wb.create_sheet(SH_PERF)
        r  = _write_title(ws, 2, 5, "STATEMENT OF FINANCIAL PERFORMANCE",
                          self._entity, self._cur_period)
        _blue_hdr(ws.cell(row=r, column=2))
        for col, val in [(3, "Notes"), (4, self._cur_period), (5, self._prev_period)]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1
        for col, val in [(2, None), (3, None), (4, "Kshs"), (5, "Kshs")]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1

        # Non-exchange revenue
        ws.cell(row=r, column=2,
                value="Revenue from non-exchange transactions").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ne_items = [
            ("Transfers from Exchequer",                       "4"),
            ("Transfers from Domestic and Foreign Partners",   "6_7"),
            ("Proceeds from Domestic and Foreign Grants",      "3"),
            ("Tax Receipts",                                   "1"),
            ("Social Security Contributions",                  "2"),
            ("Reimbursements and Refunds",                     "9"),
            ("Proceeds from Sale of Assets",                   "8"),
        ]
        ne_start = r
        for lbl, k in ne_items:
            if self._has(k):
                self._line(ws, r, lbl, k)
                r += 1
        ne_end = r - 1
        total_ne_row = 0
        if ne_start <= ne_end:
            _sum_total(ws, r, "Total non-exchange revenue", ne_start, ne_end)
            total_ne_row = r
            r += 2

        # Exchange revenue
        ws.cell(row=r, column=2,
                value="Revenue from exchange transactions").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ex_start = r
        if self._has("11"):
            self._line(ws, r, "Other Receipts", "11")
            r += 1
        ex_end = r - 1
        total_ex_row = 0
        if ex_start <= ex_end:
            _sum_total(ws, r, "Total exchange revenue", ex_start, ex_end)
            total_ex_row = r
            r += 2

        # Total Revenue — formula linking the two subtotals
        ws.cell(row=r, column=2,
                value="Total Revenue").font = Font(name=TNR, size=11, bold=True)
        if total_ne_row and total_ex_row:
            rev_f_cur  = f"=D{total_ne_row}+D{total_ex_row}"
            rev_f_prev = f"=E{total_ne_row}+E{total_ex_row}"
        elif total_ne_row:
            rev_f_cur  = f"=D{total_ne_row}"
            rev_f_prev = f"=E{total_ne_row}"
        elif total_ex_row:
            rev_f_cur  = f"=D{total_ex_row}"
            rev_f_prev = f"=E{total_ex_row}"
        else:
            rev_f_cur  = self._total_rev_cur
            rev_f_prev = self._total_rev_prev
        _fnum(ws.cell(row=r, column=4), rev_f_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5), rev_f_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        total_rev_row = r
        self._perf_rows['total_rev'] = r
        r += 2

        # Expenses
        ws.cell(row=r, column=2,
                value="Expenses").font = Font(name=TNR, size=11, bold=True)
        r += 1
        exp_items = [
            ("Compensation of Employees",          "12"),
            ("Use of Goods and Services",           "13"),
            ("Subsidies",                           "14"),
            ("Transfers to Other Government Units", "15"),
            ("Other Grants and Transfers",          "16"),
            ("Social Security Benefits",            "17"),
            ("Finance Costs",                       "19"),
            ("Other Payments",                      "21"),
        ]
        exp_start = r
        for lbl, k in exp_items:
            if self._has(k):
                self._line(ws, r, lbl, k)
                r += 1
        exp_end = r - 1
        total_exp_row = 0
        if exp_start <= exp_end:
            _sum_total(ws, r, "Total Expenses", exp_start, exp_end)
            total_exp_row = r
            self._perf_rows['total_exp'] = r
            r += 2

        # Surplus / Deficit — formula linking Total Revenue - Total Expenses
        r += 1
        ws.cell(row=r, column=2,
                value="Surplus/(Deficit) for the Period").font = Font(name=TNR, size=11, bold=True)
        if total_rev_row and total_exp_row:
            sur_f_cur  = f"=D{total_rev_row}-D{total_exp_row}"
            sur_f_prev = f"=E{total_rev_row}-E{total_exp_row}"
        else:
            sur_f_cur  = self._surplus_cur
            sur_f_prev = self._surplus_prev
        _fnum(ws.cell(row=r, column=4), sur_f_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5), sur_f_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="double"))
        self._perf_rows['surplus'] = r
        _finalize_sheet(ws)

    # ── Position sheet ─────────────────────────────────────────────────────
    def _build_pos_sheet(self, wb):
        ws = wb.create_sheet(SH_POS)
        r  = _write_title(ws, 2, 5, "STATEMENT OF FINANCIAL POSITION",
                          self._entity, self._cur_period)
        _blue_hdr(ws.cell(row=r, column=2))
        for col, val in [(3, "Notes"), (4, self._cur_period), (5, self._prev_period)]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1
        for col, val in [(2, None), (3, None), (4, "Kshs"), (5, "Kshs")]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1

        # ASSETS
        ws.cell(row=r, column=2, value="ASSETS").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ws.cell(row=r, column=2, value="Current Assets").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ca_start = r
        if self._has("CASH"):
            self._line(ws, r, "Cash and Cash Equivalents",        "CASH")
            r += 1
        if self._has("23"):
            self._line(ws, r, "Receivables - Imprest & Clearance", "23")
            r += 1
        ca_end = r - 1
        total_ca_row = 0
        if ca_start <= ca_end:
            _sum_total(ws, r, "Total Current Assets", ca_start, ca_end)
            total_ca_row = r
            r += 2

        ws.cell(row=r, column=2,
                value="Non-Current Assets").font = Font(name=TNR, size=11, bold=True)
        r += 1
        nca_start = r
        if self._has("18"):
            self._line(ws, r, "Property, Plant and Equipment", "18")
            r += 1
        nca_end = r - 1
        total_nca_row = 0
        if nca_start <= nca_end:
            _sum_total(ws, r, "Total Non-Current Assets", nca_start, nca_end)
            total_nca_row = r
            r += 2

        # Total Assets — formula sum of the two asset totals
        ws.cell(row=r, column=2,
                value="Total Assets (a)").font = Font(name=TNR, size=11, bold=True)
        if total_ca_row and total_nca_row:
            ta_f_cur  = f"=D{total_ca_row}+D{total_nca_row}"
            ta_f_prev = f"=E{total_ca_row}+E{total_nca_row}"
        elif total_ca_row:
            ta_f_cur  = f"=D{total_ca_row}"
            ta_f_prev = f"=E{total_ca_row}"
        elif total_nca_row:
            ta_f_cur  = f"=D{total_nca_row}"
            ta_f_prev = f"=E{total_nca_row}"
        else:
            ta_f_cur  = self._total_assets_cur
            ta_f_prev = self._total_assets_prev
        _fnum(ws.cell(row=r, column=4), ta_f_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5), ta_f_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        total_assets_row = r
        r += 2

        # LIABILITIES
        ws.cell(row=r, column=2, value="LIABILITIES").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ws.cell(row=r, column=2,
                value="Current Liabilities").font = Font(name=TNR, size=11, bold=True)
        r += 1
        cl_start = r
        if self._has("24"):
            self._line(ws, r, "Accounts Payable", "24")
            r += 1
        cl_end = r - 1
        total_cl_row = 0
        if cl_start <= cl_end:
            _sum_total(ws, r, "Total Current Liabilities", cl_start, cl_end)
            total_cl_row = r
            r += 2

        # Total Liabilities
        ws.cell(row=r, column=2,
                value="Total Liabilities (b)").font = Font(name=TNR, size=11, bold=True)
        tl_f_cur  = f"=D{total_cl_row}" if total_cl_row else self._total_liab_cur
        tl_f_prev = f"=E{total_cl_row}" if total_cl_row else self._total_liab_prev
        _fnum(ws.cell(row=r, column=4), tl_f_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5), tl_f_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        total_liab_row = r
        r += 2

        # Net Assets — formula: Total Assets - Total Liabilities
        ws.cell(row=r, column=2,
                value="Net Assets (a - b)").font = Font(name=TNR, size=11, bold=True)
        if total_assets_row and total_liab_row:
            na_f_cur  = f"=D{total_assets_row}-D{total_liab_row}"
            na_f_prev = f"=E{total_assets_row}-E{total_liab_row}"
        else:
            na_f_cur  = self._net_assets_cur
            na_f_prev = self._net_assets_prev
        _fnum(ws.cell(row=r, column=4), na_f_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5), na_f_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="double"))
        self._pos_rows['net_assets'] = r
        r += 2

        # Represented by
        ws.cell(row=r, column=2,
                value="Represented by:").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ws.cell(row=r, column=2,
                value="Accumulated Surplus/(Deficit)").font = Font(name=TNR, size=11)
        # Values filled in by _build_na_sheet once Net Assets closing rows are known
        self._pos_rows['acc_surplus'] = r
        r += 1

        ws.cell(row=r, column=2,
                value="Total Net Assets").font = Font(name=TNR, size=11, bold=True)
        self._pos_rows['total_na'] = r
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="double"))
        _finalize_sheet(ws)

    # ── Net Assets sheet ───────────────────────────────────────────────────
    def _build_na_sheet(self, wb):
        ws = wb.create_sheet(SH_NA)
        r  = _write_title(ws, 1, 3, "STATEMENT OF CHANGES IN NET ASSETS",
                          self._entity, self._cur_period)
        _blue_hdr(ws.cell(row=r, column=1))
        _blue_hdr(ws.cell(row=r, column=2, value="Accumulated Surplus"))
        _blue_hdr(ws.cell(row=r, column=3, value="Total"))
        r += 1

        surplus_row = self._perf_rows.get('surplus', 0)

        # ── Prior year section ──
        prior_open_row = r
        ws.cell(row=r, column=1,
                value="Balance as at 1st July").font = Font(name=TNR, size=11)
        # Prior opening = prior net assets - prior surplus (computed; no direct note link)
        _num(ws.cell(row=r, column=2), self._na_prior_opening, color=_GREEN)
        _num(ws.cell(row=r, column=3), self._na_prior_opening, color=_GREEN)
        r += 1

        prior_surplus_row = r
        ws.cell(row=r, column=1,
                value="Surplus/(Deficit) for the year").font = Font(name=TNR, size=11)
        # Link to Performance sheet previous-period surplus (column E)
        if surplus_row:
            _fnum(ws.cell(row=r, column=2),
                  f"='{SH_PERF}'!E{surplus_row}", color=_GREEN)
            _fnum(ws.cell(row=r, column=3),
                  f"='{SH_PERF}'!E{surplus_row}", color=_GREEN)
        else:
            _num(ws.cell(row=r, column=2), self._surplus_prev, color=_GREEN)
            _num(ws.cell(row=r, column=3), self._surplus_prev, color=_GREEN)
        r += 1

        prior_close_row = r
        ws.cell(row=r, column=1,
                value="As at June 30, 20xx").font = Font(name=TNR, size=11, bold=True)
        # Prior closing = prior opening + prior surplus (formula)
        _fnum(ws.cell(row=r, column=2),
              f"=B{prior_open_row}+B{prior_surplus_row}", bold=True)
        _fnum(ws.cell(row=r, column=3),
              f"=C{prior_open_row}+C{prior_surplus_row}", bold=True)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        r += 2

        # ── Current year section ──
        curr_open_row = r
        ws.cell(row=r, column=1,
                value="As at July 1, 20xx").font = Font(name=TNR, size=11)
        # Current opening = prior closing (cross-row link)
        _fnum(ws.cell(row=r, column=2), f"=B{prior_close_row}")
        _fnum(ws.cell(row=r, column=3), f"=C{prior_close_row}")
        r += 1

        curr_surplus_row = r
        ws.cell(row=r, column=1,
                value="Surplus/(Deficit) for the year").font = Font(name=TNR, size=11)
        # Link to Performance sheet current-period surplus (column D)
        if surplus_row:
            _fnum(ws.cell(row=r, column=2),
                  f"='{SH_PERF}'!D{surplus_row}", color=_GREEN)
            _fnum(ws.cell(row=r, column=3),
                  f"='{SH_PERF}'!D{surplus_row}", color=_GREEN)
        else:
            _num(ws.cell(row=r, column=2), self._surplus_cur, color=_GREEN)
            _num(ws.cell(row=r, column=3), self._surplus_cur, color=_GREEN)
        r += 1

        # Return to Exchequer — link to Notes sheet (negated)
        rte_row = r
        n26_note_row = self._note_rows.get('26', 0)
        ws.cell(row=r, column=1,
                value="Return to Exchequer").font = Font(name=TNR, size=11)
        if n26_note_row:
            _fnum(ws.cell(row=r, column=2),
                  f"=-'{SH_NOTES}'!C{n26_note_row}", color=_GREEN)
            _fnum(ws.cell(row=r, column=3),
                  f"=-'{SH_NOTES}'!C{n26_note_row}", color=_GREEN)
        else:
            _num(ws.cell(row=r, column=2), -self._rte_cur, color=_GREEN)
            _num(ws.cell(row=r, column=3), -self._rte_cur, color=_GREEN)
        r += 1

        curr_close_row = r
        ws.cell(row=r, column=1,
                value="As at June 30, 20xx").font = Font(name=TNR, size=11, bold=True)
        # Current closing = opening + surplus + RTE (all already signed)
        _fnum(ws.cell(row=r, column=2),
              f"=B{curr_open_row}+B{curr_surplus_row}+B{rte_row}", bold=True)
        _fnum(ws.cell(row=r, column=3),
              f"=C{curr_open_row}+C{curr_surplus_row}+C{rte_row}", bold=True)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="double"))

        self._na_rows['curr_close'] = curr_close_row
        self._na_rows['prev_close'] = prior_close_row

        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 22

        # ── Back-fill Position sheet "Represented by" rows ───────────────
        pos_ws = wb[SH_POS]
        acc_r   = self._pos_rows.get('acc_surplus', 0)
        total_r = self._pos_rows.get('total_na', 0)
        if acc_r:
            # Link to Net Assets sheet closing balances
            _fnum(pos_ws.cell(row=acc_r, column=4),
                  f"='{SH_NA}'!B{curr_close_row}", color=_GREEN)
            _fnum(pos_ws.cell(row=acc_r, column=5),
                  f"='{SH_NA}'!B{prior_close_row}", color=_GREEN)
        if total_r:
            _fnum(pos_ws.cell(row=total_r, column=4),
                  f"='{SH_NA}'!B{curr_close_row}", bold=True)
            _fnum(pos_ws.cell(row=total_r, column=5),
                  f"='{SH_NA}'!B{prior_close_row}", bold=True)

    # ── Cash Flow sheet ────────────────────────────────────────────────────
    def _build_cf_sheet(self, wb):
        ws = wb.create_sheet(SH_CF)
        r  = _write_title(ws, 2, 5, "STATEMENT OF CASH FLOWS",
                          self._entity, self._cur_period)
        _blue_hdr(ws.cell(row=r, column=2))
        for col, val in [(3, "Notes"), (4, self._cur_period), (5, self._prev_period)]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1
        for col, val in [(2, None), (3, None), (4, "Kshs"), (5, "Kshs")]:
            _blue_hdr(ws.cell(row=r, column=col, value=val))
        r += 1

        # ── Operating activities ──
        ws.cell(row=r, column=2,
                value="Cash flows from operating activities").font = Font(name=TNR, size=11, bold=True)
        r += 1
        ws.cell(row=r, column=2,
                value="Receipts").font = Font(name=TNR, size=11, bold=True)
        r += 1

        rec_items = [
            ("Exchequer Releases",                             "4"),
            ("Transfers from Domestic and Foreign Partners",   "6_7"),
            ("Tax Receipts",                                   "1"),
            ("Social Security Contributions",                  "2"),
            ("Proceeds from Domestic and Foreign Grants",      "3"),
            ("Proceeds from Sale of Assets",                   "8"),
            ("Other Receipts",                                 "11"),
        ]
        rec_start = r
        rec_cur = rec_prev = 0.0
        for lbl, k in rec_items:
            if self._has(k):
                self._line(ws, r, lbl, k)
                rec_cur  += self._nc(k)
                rec_prev += self._np(k)
                r += 1
        rec_end = r - 1
        total_rec_row = 0
        if rec_start <= rec_end:
            _sum_total(ws, r, "Total Receipts", rec_start, rec_end)
            total_rec_row = r
            r += 2

        ws.cell(row=r, column=2,
                value="Payments").font = Font(name=TNR, size=11, bold=True)
        r += 1
        pay_items = [
            ("Compensation of Employees",          "12"),
            ("Use of Goods and Services",           "13"),
            ("Transfers to Other Government Units", "15"),
            ("Other Grants and Transfers",          "16"),
            ("Social Security Benefits",            "17"),
        ]
        pay_start = r
        pay_cur = pay_prev = 0.0
        for lbl, k in pay_items:
            if self._has(k):
                self._line(ws, r, lbl, k, negate=True)
                pay_cur  += -self._nc(k)
                pay_prev += -self._np(k)
                r += 1
        pay_end = r - 1
        total_pay_row = 0
        if pay_start <= pay_end:
            _sum_total(ws, r, "Total Payments", pay_start, pay_end)
            total_pay_row = r
            r += 1

        # Working Capital (synthesised — no direct Notes link)
        r += 1
        wc     = self._find("WC")
        wc_cur = wc.total_cur if wc else 0.0
        wc_row = r
        ws.cell(row=r, column=2,
                value="Changes in Working Capital").font = Font(name=TNR, size=11)
        if wc and wc.seq_num:
            ws.cell(row=r, column=3,
                    value=str(wc.seq_num)).font = Font(name=TNR, size=11)
        _num(ws.cell(row=r, column=4), wc_cur)
        _num(ws.cell(row=r, column=5), 0.0)
        r += 2

        # Net operating — formula linking Receipts + Payments + WC rows
        net_ops_cur  = rec_cur  + pay_cur  + wc_cur
        net_ops_prev = rec_prev + pay_prev
        ws.cell(row=r, column=2,
                value="Net cash flows from/(used in) operating activities"
                ).font = Font(name=TNR, size=11, bold=True)
        if total_rec_row and total_pay_row:
            _fnum(ws.cell(row=r, column=4),
                  f"=D{total_rec_row}+D{total_pay_row}+D{wc_row}", bold=True)
            _fnum(ws.cell(row=r, column=5),
                  f"=E{total_rec_row}+E{total_pay_row}", bold=True)
        else:
            _num(ws.cell(row=r, column=4), net_ops_cur,  bold=True)
            _num(ws.cell(row=r, column=5), net_ops_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        net_ops_row = r
        r += 2

        # ── Investing activities ──
        ws.cell(row=r, column=2,
                value="Cash flows from investing activities").font = Font(name=TNR, size=11, bold=True)
        r += 1
        inv_cur = inv_prev = 0.0
        net_inv_row = 0
        ppe = self._find("18")
        if ppe and ppe.items:
            opening_cur = ppe.items[0].cur
            acq_cur  = -(ppe.total_cur - opening_cur)
            acq_prev = -ppe.total_prev
            ws.cell(row=r, column=2,
                    value="Acquisition of Assets").font = Font(name=TNR, size=11)
            ws.cell(row=r, column=3,
                    value=str(ppe.seq_num) if ppe.seq_num else "").font = Font(name=TNR, size=11)
            _num(ws.cell(row=r, column=4), acq_cur)
            _num(ws.cell(row=r, column=5), acq_prev)
            inv_cur  = acq_cur
            inv_prev = acq_prev
            r += 1
        if inv_cur != 0 or inv_prev != 0:
            _total_line(ws, r,
                        "Net cash flows from/(used in) investing activities",
                        inv_cur, inv_prev)
            net_inv_row = r
            r += 2

        # ── Financing activities ──
        fin_cur = fin_prev = 0.0
        net_fin_row = 0
        n26 = self._find("26")
        n26_note_row = self._note_rows.get('26', 0)
        if n26 and n26.items and self._nc("26") != 0:
            ws.cell(row=r, column=2,
                    value="Cash flows from financing activities").font = Font(name=TNR, size=11, bold=True)
            r += 1
            ws.cell(row=r, column=2,
                    value="Return to Exchequer").font = Font(name=TNR, size=11)
            ws.cell(row=r, column=3,
                    value=str(n26.seq_num) if n26.seq_num else "").font = Font(name=TNR, size=11)
            fin_cur = -self._nc("26")
            if n26_note_row:
                _fnum(ws.cell(row=r, column=4), f"=-'{SH_NOTES}'!C{n26_note_row}")
                _fnum(ws.cell(row=r, column=5), f"=-'{SH_NOTES}'!D{n26_note_row}")
            else:
                _num(ws.cell(row=r, column=4), fin_cur)
                _num(ws.cell(row=r, column=5), 0.0)
            r += 1
            _total_line(ws, r,
                        "Net cash flows from/(used in) financing activities",
                        fin_cur, 0.0)
            net_fin_row = r
            r += 2

        # Net change in cash — formula linking net ops + net inv + net fin
        net_chg_cur  = net_ops_cur  + inv_cur  + fin_cur
        net_chg_prev = net_ops_prev + inv_prev + fin_prev
        ws.cell(row=r, column=2,
                value="Net increase/(decrease) in cash and cash equivalents"
                ).font = Font(name=TNR, size=11, bold=True)
        parts_cur  = [f"D{net_ops_row}"]
        parts_prev = [f"E{net_ops_row}"]
        if net_inv_row:
            parts_cur.append(f"D{net_inv_row}")
            parts_prev.append(f"E{net_inv_row}")
        if net_fin_row:
            parts_cur.append(f"D{net_fin_row}")
            parts_prev.append(f"E{net_fin_row}")
        _fnum(ws.cell(row=r, column=4),
              "=" + "+".join(parts_cur)  if net_ops_row else net_chg_cur,  bold=True)
        _fnum(ws.cell(row=r, column=5),
              "=" + "+".join(parts_prev) if net_ops_row else net_chg_prev, bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))
        net_chg_row = r
        r += 2

        # Opening / Closing cash
        cash_sn    = str(self._find("CASH").seq_num) if self._find("CASH") else ""
        cash_note_row = self._note_rows.get('CASH', 0)
        cash_prev  = self._np("CASH")
        cash_cur   = self._nc("CASH")
        cash_start_cur  = cash_prev
        cash_start_prev = cash_prev - net_chg_prev

        cash_open_row = r
        ws.cell(row=r, column=2,
                value="Cash and cash equivalents at start of period").font = Font(name=TNR, size=11)
        ws.cell(row=r, column=3, value=cash_sn).font = Font(name=TNR, size=11)
        _num(ws.cell(row=r, column=4), cash_start_cur)
        _num(ws.cell(row=r, column=5), cash_start_prev)
        r += 1

        cash_close_row = r
        ws.cell(row=r, column=2,
                value="Cash and cash equivalents at end of period"
                ).font = Font(name=TNR, size=11, bold=True)
        ws.cell(row=r, column=3, value=cash_sn).font = Font(name=TNR, size=11)
        # Closing = opening + net change (formula)
        _fnum(ws.cell(row=r, column=4),
              f"=D{cash_open_row}+D{net_chg_row}", bold=True)
        _fnum(ws.cell(row=r, column=5),
              f"=E{cash_open_row}+E{net_chg_row}", bold=True)
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = Border(bottom=Side(style="double"))
        r += 2

        # ── Controls ──
        def _red_ctrl(row, label, cur_v, prev_v):
            ws.cell(row=row, column=2,
                    value=label).font = Font(name=TNR, size=11, bold=True, color=_RED_C)
            _fnum(ws.cell(row=row, column=4), cur_v,  color=_RED_C)
            _fnum(ws.cell(row=row, column=5), prev_v, color=_RED_C)

        # Control: CF closing vs Note total for Cash
        if cash_note_row:
            ctrl_cur  = f"=D{cash_close_row}-'{SH_NOTES}'!C{cash_note_row}"
            ctrl_prev = f"=E{cash_close_row}-'{SH_NOTES}'!D{cash_note_row}"
            note_cur  = f"='{SH_NOTES}'!C{cash_note_row}"
            note_prev = f"='{SH_NOTES}'!D{cash_note_row}"
        else:
            cash_end_cur  = cash_start_cur  + net_chg_cur
            cash_end_prev = cash_start_prev + net_chg_prev
            ctrl_cur  = cash_end_cur  - cash_cur
            ctrl_prev = cash_end_prev - cash_prev
            note_cur  = cash_cur
            note_prev = cash_prev

        _red_ctrl(r, "CONTROL: Cash per Note", note_cur, note_prev)
        ws.cell(row=r, column=3, value=cash_sn).font = Font(name=TNR, size=11, color=_RED_C)
        r += 1
        _red_ctrl(r, "CONTROL: Difference (CF vs Note)", ctrl_cur, ctrl_prev)
        r += 2

        # Control: Position balance check (Net Assets vs Net Assets closing)
        pos_na_row = self._pos_rows.get('net_assets', 0)
        na_close   = self._na_rows.get('curr_close', 0)
        na_prev_cl = self._na_rows.get('prev_close', 0)
        if pos_na_row and na_close:
            bal_cur  = f"='{SH_POS}'!D{pos_na_row}-'{SH_NA}'!B{na_close}"
            bal_prev = f"='{SH_POS}'!E{pos_na_row}-'{SH_NA}'!B{na_prev_cl}" if na_prev_cl else 0.0
        else:
            bal_cur  = self._net_assets_cur  - self._na_close_cur
            bal_prev = self._net_assets_prev - self._na_close_prev
        _red_ctrl(r, "CONTROL: Position Balance Check", bal_cur, bal_prev)

        _finalize_sheet(ws)

    # ── Helpers ────────────────────────────────────────────────────────────
    def _find(self, key: str) -> Optional[ParsedNote]:
        for n in self._notes:
            if n.key == key:
                return n
        return None

    def _nc(self, key: str) -> float:
        n = self._find(key)
        return n.total_cur if n else 0.0

    def _np(self, key: str) -> float:
        n = self._find(key)
        return n.total_prev if n else 0.0

    def _has(self, key: str) -> bool:
        n = self._find(key)
        return n is not None and (n.total_cur != 0 or n.total_prev != 0)

    def _line(self, ws, r: int, label: str, key: str, negate: bool = False):
        """Write one line row; values are cross-sheet formulas when available."""
        n    = self._find(key)
        sign = -1 if negate else 1
        cv   = (n.total_cur  if n else 0.0) * sign
        pv   = (n.total_prev if n else 0.0) * sign
        seq  = str(n.seq_num) if n and n.seq_num else ""
        ws.cell(row=r, column=2, value=label).font = Font(name=TNR, size=11)
        ws.cell(row=r, column=3, value=seq).font   = Font(name=TNR, size=11)

        note_row = self._note_rows.get(key)
        if note_row:
            pfx = "-" if negate else ""
            _fnum(ws.cell(row=r, column=4), f"={pfx}'{SH_NOTES}'!C{note_row}")
            _fnum(ws.cell(row=r, column=5), f"={pfx}'{SH_NOTES}'!D{note_row}")
        else:
            _num(ws.cell(row=r, column=4), cv)
            _num(ws.cell(row=r, column=5), pv)


# ── Module-level formatting helpers ───────────────────────────────────────
def _write_title(ws, col_start: int, col_end: int, title: str,
                 entity: str, cur_period: str) -> int:
    """Write 3 title rows; return the next row number (4)."""
    rows = [entity, f"FOR THE PERIOD ENDED {cur_period}", title]
    for i, text in enumerate(rows, start=1):
        ws.merge_cells(start_row=i, start_column=col_start,
                       end_row=i,   end_column=col_end)
        c = ws.cell(row=i, column=col_start, value=text)
        c.font      = Font(name=TNR, bold=True, size=13 if i == 3 else 11)
        c.alignment = Alignment(horizontal="center")
    return 4


def _blue_hdr(cell):
    cell.font      = Font(name=TNR, bold=True, size=11, color=_WHITE)
    cell.fill      = PatternFill(start_color=_BLUE, end_color=_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _num(cell, value: float, bold: bool = False, color: str = None):
    """Write a hard-coded numeric value."""
    cell.value         = value
    cell.number_format = NUM_FMT
    kwargs = {"name": TNR, "size": 11, "bold": bold}
    if color:
        kwargs["color"] = color
    cell.font      = Font(**kwargs)
    cell.alignment = Alignment(horizontal="right")


def _fnum(cell, value, bold: bool = False, color: str = None):
    """Write a formula string or numeric value with number formatting."""
    cell.value         = value
    cell.number_format = NUM_FMT
    kwargs = {"name": TNR, "size": 11, "bold": bold}
    if color:
        kwargs["color"] = color
    cell.font      = Font(**kwargs)
    cell.alignment = Alignment(horizontal="right")


def _sum_total(ws, r: int, label: str, start_r: int, end_r: int):
    """Write a subtotal row using SUM formulas over the data rows above."""
    ws.cell(row=r, column=2, value=label).font = Font(name=TNR, size=11, bold=True)
    for col, src in [(4, "D"), (5, "E")]:
        c = ws.cell(row=r, column=col)
        c.value         = f"=SUM({src}{start_r}:{src}{end_r})"
        c.number_format = NUM_FMT
        c.font          = Font(name=TNR, size=11, bold=True)
        c.alignment     = Alignment(horizontal="right")
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = Border(bottom=Side(style="thin"))


def _total_line(ws, r: int, label: str, cur_val: float, prev_val: float):
    """Write a total row with hard-coded values (fallback for derived totals)."""
    ws.cell(row=r, column=2, value=label).font = Font(name=TNR, size=11, bold=True)
    _num(ws.cell(row=r, column=4), cur_val,  bold=True)
    _num(ws.cell(row=r, column=5), prev_val, bold=True)
    for c in range(2, 6):
        ws.cell(row=r, column=c).border = Border(bottom=Side(style="thin"))


def _finalize_sheet(ws):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22


def _to_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _is_note_title(txt: str):
    """
    Parse a note title line like '4. Transfers from Exchequer'
    Returns (key, title) or ('', '') if not a note title.
    """
    if not txt:
        return "", ""
    i = 0
    while i < len(txt) and txt[i].isdigit():
        i += 1
    if i == 0:
        return "", ""
    # Optional letter suffix (A / B)
    if i < len(txt) and txt[i].upper() in ("A", "B"):
        i += 1
    num  = txt[:i]
    rest = txt[i:]
    rest_stripped = rest.strip()
    if rest_stripped.startswith("."):
        title = rest_stripped[1:].strip()
    elif len(rest) >= 2 and rest[:2] == "  ":
        title = rest_stripped
    else:
        return "", ""
    if len(title) < 3:
        return "", ""
    return num, title
