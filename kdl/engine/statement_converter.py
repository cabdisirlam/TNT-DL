"""
Bank Statement Converter.
Direct Python port of the Build_Statement_Output VBA macro logic.
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    openpyxl = None


HEADER_ROW = 13
DATA_START_ROW = 14
AUDIT_DETAIL_HEADER_ROW = 9
AUDIT_DETAIL_FIRST_ROW = 10
DN_PREFIX_CELL = r"\*s"
OUTPUT_LAST_COL = 16
OUTPUT_FIRST_ROW = 2


@dataclass
class ConversionResult:
    success: bool
    message: str
    output_rows: int = 0
    skipped_rows: int = 0
    total_debits: float = 0.0
    total_credits: float = 0.0
    opening_balance: Optional[float] = None
    closing_balance_stmt: Optional[float] = None
    closing_balance_calc: Optional[float] = None
    variance: Optional[float] = None
    output_data: List[List] = field(default_factory=list)
    audit_rows: List[List] = field(default_factory=list)
    saved_path: str = ""


def _normalize_header_text(value) -> str:
    text = str(value or "").replace("\xa0", " ").replace("\r", " ").replace("\n", " ").strip()
    while "  " in text:
        text = text.replace("  ", " ")
    return text.upper()


def _find_header_col(ws: Worksheet, hdr_row: int, header_text: str) -> int:
    want = _normalize_header_text(header_text)
    for col in range(1, ws.max_column + 1):
        if _normalize_header_text(ws.cell(row=hdr_row, column=col).value) == want:
            return col
    return 0


def _find_txn_header_row(ws: Worksheet) -> int:
    last_row = max(1, ws.max_row)
    for row_num in range(1, last_row + 1):
        if (
            _find_header_col(ws, row_num, "Date") > 0
            and _find_header_col(ws, row_num, "Transaction Details") > 0
            and _find_header_col(ws, row_num, "Debit") > 0
            and _find_header_col(ws, row_num, "Credit") > 0
        ):
            return row_num
    return 0


def _normalize_row13_headers(ws: Worksheet):
    ws.cell(row=HEADER_ROW, column=1, value="Date")
    ws.cell(row=HEADER_ROW, column=2, value="Transaction Reference")
    ws.cell(row=HEADER_ROW, column=3, value="Transaction Details")
    ws.cell(row=HEADER_ROW, column=4, value="Transaction Type")
    ws.cell(row=HEADER_ROW, column=5, value="Originator Reference")
    ws.cell(row=HEADER_ROW, column=6, value="Debit")
    ws.cell(row=HEADER_ROW, column=7, value="Credit")
    ws.cell(row=HEADER_ROW, column=8, value="Closing Balance")
    ws.row_dimensions[HEADER_ROW].font = Font(bold=True)


def _safe_double(value) -> float:
    parsed = _parse_number(value)
    return float(parsed) if parsed is not None else 0.0


def _parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if len(text) >= 2 and text[0] == "(" and text[-1] == ")":
        text = "-" + text[1:-1]
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _cell_display_text(cell) -> str:
    text_value = getattr(cell, "text", None)
    if text_value not in (None, ""):
        return str(text_value).strip()
    return str(cell.value or "").strip()


def _month_text_to_num(mon: str) -> int:
    mon = str(mon or "").strip().lower()
    mapping = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "may": 5,
        "jun": 6,
        "jul": 7,
        "aug": 8,
        "sep": 9,
        "oct": 10,
        "nov": 11,
        "dec": 12,
    }
    return mapping.get(mon[:3], 0)


def _parse_mdy(text: str, sep: str):
    parts = text.strip().split(sep)
    if len(parts) != 3:
        return None
    try:
        mm = int(parts[0])
        dd = int(parts[1])
        yy = int(parts[2])
        if yy < 100:
            yy += 2000
        if mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return None
        return date(yy, mm, dd)
    except (TypeError, ValueError):
        return None


def _parse_dmon_y(text: str, sep: str):
    parts = text.strip().split(sep)
    if len(parts) != 3:
        return None
    try:
        dd = int(parts[0])
        mm = _month_text_to_num(parts[1])
        yy = int(parts[2])
        if yy < 100:
            yy += 2000
        if mm < 1 or mm > 12 or dd < 1 or dd > 31:
            return None
        return date(yy, mm, dd)
    except (TypeError, ValueError):
        return None


def _parse_date_cell(cell) -> Optional[date]:
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(cell.value).strip() if cell.value is not None else ""
    if text in ("", "####"):
        text = _cell_display_text(cell)
    if not text:
        return None
    if " " in text:
        text = text.split(" ", 1)[0]

    if "/" in text:
        return _parse_mdy(text, "/")
    if "-" in text:
        return _parse_dmon_y(text, "-")
    return None


def _strip_leading_zeros(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""
    stripped = value.lstrip("0")
    return stripped if stripped else ""


def _extract_doc_no_10(details: str) -> str:
    details = str(details or "").strip()
    if not details:
        return ""

    run = ""
    best10 = ""
    for ch in details:
        if ch.isdigit():
            run += ch
        else:
            if len(run) >= 10:
                best10 = run[-10:]
            run = ""
    if len(run) >= 10:
        best10 = run[-10:]

    if not best10:
        return ""
    return _strip_leading_zeros(best10)


def _fmt_date(dt_value: date) -> str:
    return dt_value.strftime("%d-%b-%Y")


def _make_payment_row(doc_no, dt_value: date, amount: float) -> list:
    return [
        "tab",
        "tab",
        "trfd",
        "tab",
        str(doc_no) if doc_no else "",
        "tab",
        _fmt_date(dt_value),
        "tab",
        _fmt_date(dt_value),
        "tab",
        amount,
        "tab",
        DN_PREFIX_CELL,
        "*dn",
    ]


def _make_receipt_row(doc_no: str, dt_value: date, amount: float) -> list:
    return [
        "tab",
        "*dn",
        "r",
        "tab",
        "trfc",
        "tab",
        doc_no,
        "tab",
        _fmt_date(dt_value),
        "tab",
        _fmt_date(dt_value),
        "tab",
        amount,
        "tab",
        DN_PREFIX_CELL,
        "*dn",
    ]


def _write_rows_to_sheet(ws_out: Worksheet, rows: List[List]):
    for index, row_vals in enumerate(rows, start=OUTPUT_FIRST_ROW):
        for col_index, value in enumerate(row_vals, start=1):
            out_value = value
            if row_vals[:3] == ["tab", "tab", "trfd"]:
                if col_index in (7, 9) and isinstance(value, str):
                    parsed = _parse_dmon_y(value, "-")
                    out_value = parsed if parsed is not None else value
                elif col_index == 11:
                    parsed_num = _parse_number(value)
                    out_value = parsed_num if parsed_num is not None else value
                elif col_index == 5 and str(value).strip().isdigit():
                    out_value = int(str(value).strip())
            elif len(row_vals) >= 5 and row_vals[:5] == ["tab", "*dn", "r", "tab", "trfc"]:
                if col_index in (9, 11) and isinstance(value, str):
                    parsed = _parse_dmon_y(value, "-")
                    out_value = parsed if parsed is not None else value
                elif col_index == 13:
                    parsed_num = _parse_number(value)
                    out_value = parsed_num if parsed_num is not None else value
            ws_out.cell(row=index, column=col_index, value=out_value)


def _write_audit_header(ws_audit: Worksheet):
    headers = [
        "Row#",
        "Date (Raw)",
        "Reference",
        "Transaction Type",
        "Transaction Details",
        "Debit (Raw)",
        "Credit (Raw)",
        "Debit (Parsed)",
        "Credit (Parsed)",
        "Skip Reason",
    ]
    for col_index, header in enumerate(headers, start=1):
        cell = ws_audit.cell(row=AUDIT_DETAIL_HEADER_ROW, column=col_index, value=header)
        cell.font = Font(bold=True)


def _add_audit(
    ws_audit: Worksheet,
    audit_row: int,
    row_num: int,
    raw_date: str,
    raw_ref,
    txn_type: str,
    details: str,
    raw_debit,
    raw_credit,
    reason: str,
):
    values = [
        row_num,
        raw_date,
        raw_ref,
        txn_type,
        details,
        raw_debit,
        raw_credit,
        _safe_double(raw_debit),
        _safe_double(raw_credit),
        reason,
    ]
    for col_index, value in enumerate(values, start=1):
        ws_audit.cell(row=audit_row, column=col_index, value=value)


def _write_audit_summary(
    ws_audit: Worksheet,
    opening_balance,
    additions_cred: float,
    outs_deb: float,
    closing_calc,
    closing_stmt,
    variance,
):
    labels = [
        "BANK RECONCILIATION SUMMARY",
        "Opening Balance",
        "Additions (Receipts/Credits Captured)",
        "Less: Outs (Payments/Debits Captured)",
        "Closing Balance (Calculated)",
        "Closing Balance (Statement)",
        "Closing Match (Statement - Calculated)",
    ]
    for row_index, label in enumerate(labels, start=1):
        cell = ws_audit.cell(row=row_index, column=1, value=label)
        if row_index == 1:
            cell.font = Font(bold=True)

    if opening_balance is None:
        ws_audit.cell(row=2, column=2, value="N/A (No Closing Balance column detected)")
        ws_audit.cell(row=3, column=2, value=additions_cred)
        ws_audit.cell(row=4, column=2, value=outs_deb)
    else:
        ws_audit.cell(row=2, column=2, value=float(opening_balance))
        ws_audit.cell(row=3, column=2, value=additions_cred)
        ws_audit.cell(row=4, column=2, value=outs_deb)
        ws_audit.cell(row=5, column=2, value=float(closing_calc))
        ws_audit.cell(row=6, column=2, value="N/A" if closing_stmt is None else float(closing_stmt))
        ws_audit.cell(row=7, column=2, value="N/A" if variance is None else float(variance))


def _get_or_create_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(title=sheet_name)


def _clear_sheet(ws: Worksheet):
    ws.delete_rows(1, ws.max_row or 1)
    ws.delete_cols(1, ws.max_column or 1)


def _add_row_to_amount_map(mapping: Dict[str, List[int]], ref10: str, amount: float, row_num: int):
    key = f"{ref10}|{amount:.2f}"
    mapping.setdefault(key, []).append(row_num)


def _sort_key(row_vals: List):
    if len(row_vals) > 8 and row_vals[8]:
        return _parse_dmon_y(str(row_vals[8]), "-") or date.min
    if len(row_vals) > 6 and row_vals[6]:
        return _parse_dmon_y(str(row_vals[6]), "-") or date.min
    return date.min


def convert_statement(wb: Workbook, sheet_name: str, skip_contra: bool = True) -> ConversionResult:
    if openpyxl is None:
        return ConversionResult(success=False, message="openpyxl not installed.")
    if sheet_name not in wb.sheetnames:
        return ConversionResult(success=False, message=f'Sheet "{sheet_name}" not found.')

    ws = wb[sheet_name]

    hdr_row = _find_txn_header_row(ws)
    if hdr_row == 0:
        _normalize_row13_headers(ws)
        hdr_row = HEADER_ROW

    col_date = _find_header_col(ws, hdr_row, "Date")
    col_ref = _find_header_col(ws, hdr_row, "Transaction Reference")
    col_details = _find_header_col(ws, hdr_row, "Transaction Details")
    col_type = _find_header_col(ws, hdr_row, "Transaction Type")
    col_debit = _find_header_col(ws, hdr_row, "Debit")
    col_credit = _find_header_col(ws, hdr_row, "Credit")
    col_bal = _find_header_col(ws, hdr_row, "Closing Balance")
    if col_bal == 0:
        col_bal = _find_header_col(ws, hdr_row, "Balance")

    if col_date == 0 or col_details == 0 or col_debit == 0 or col_credit == 0:
        return ConversionResult(
            success=False,
            message="Required columns not found (Date, Transaction Details, Debit, Credit).",
        )

    last_row = ws.cell(row=ws.max_row, column=col_date).row
    while last_row >= DATA_START_ROW and not str(ws.cell(row=last_row, column=col_date).value or "").strip():
        last_row -= 1
    if last_row < DATA_START_ROW:
        return ConversionResult(success=False, message="No data rows found.")

    ws_audit = _get_or_create_sheet(wb, "Audit_Skipped")
    ws_out = _get_or_create_sheet(wb, "Output")
    _clear_sheet(ws_audit)
    _clear_sheet(ws_out)
    _write_audit_header(ws_audit)

    audit_row = AUDIT_DETAIL_FIRST_ROW
    output_data: List[List] = []
    receipts: List[List] = []
    audit_rows: List[List] = []
    skip_row: Dict[int, str] = {}
    skip_tally: Dict[str, int] = {}
    total_debits = 0.0
    total_credits = 0.0

    opening_balance = None
    closing_balance_stmt = None
    if col_bal > 0:
        first_bal_row = 0
        last_bal_row = 0
        for row_num in range(DATA_START_ROW, last_row + 1):
            if _parse_number(ws.cell(row=row_num, column=col_bal).value) is not None:
                first_bal_row = row_num
                break
        for row_num in range(last_row, DATA_START_ROW - 1, -1):
            if _parse_number(ws.cell(row=row_num, column=col_bal).value) is not None:
                last_bal_row = row_num
                break

        if first_bal_row > 0:
            first_bal = float(_parse_number(ws.cell(row=first_bal_row, column=col_bal).value))
            first_debit = _safe_double(ws.cell(row=first_bal_row, column=col_debit).value)
            first_credit = _safe_double(ws.cell(row=first_bal_row, column=col_credit).value)
            opening_balance = first_bal - first_credit + first_debit

        if last_bal_row > 0:
            closing_balance_stmt = float(_parse_number(ws.cell(row=last_bal_row, column=col_bal).value))

    if skip_contra:
        debit_map: Dict[str, List[int]] = {}
        credit_map: Dict[str, List[int]] = {}
        for row_num in range(DATA_START_ROW, last_row + 1):
            details = str(ws.cell(row=row_num, column=col_details).value or "")
            ref10 = _extract_doc_no_10(details)
            if not ref10:
                continue
            debit_value = _safe_double(ws.cell(row=row_num, column=col_debit).value)
            credit_value = _safe_double(ws.cell(row=row_num, column=col_credit).value)
            if debit_value > 0 and credit_value == 0:
                _add_row_to_amount_map(debit_map, ref10, debit_value, row_num)
            elif credit_value > 0 and debit_value == 0:
                _add_row_to_amount_map(credit_map, ref10, credit_value, row_num)

        for key in debit_map.keys():
            if key not in credit_map:
                continue
            pairs = min(len(debit_map[key]), len(credit_map[key]))
            for index in range(pairs):
                skip_row[debit_map[key][index]] = "CONTRA_MATCHED_DETAILS10_AMOUNT"
                skip_row[credit_map[key][index]] = "CONTRA_MATCHED_DETAILS10_AMOUNT"

    for row_num in range(DATA_START_ROW, last_row + 1):
        raw_date_text = _cell_display_text(ws.cell(row=row_num, column=col_date))
        raw_ref = ws.cell(row=row_num, column=col_ref).value if col_ref > 0 else ""
        raw_debit = ws.cell(row=row_num, column=col_debit).value
        raw_credit = ws.cell(row=row_num, column=col_credit).value
        details = str(ws.cell(row=row_num, column=col_details).value or "")
        txn_type = str(ws.cell(row=row_num, column=col_type).value or "") if col_type > 0 else ""
        ref10 = _extract_doc_no_10(details)

        def add_audit(reason: str, reference_value):
            nonlocal audit_row
            record = [
                row_num,
                raw_date_text,
                reference_value,
                txn_type,
                details,
                raw_debit,
                raw_credit,
                _safe_double(raw_debit),
                _safe_double(raw_credit),
                reason,
            ]
            audit_rows.append(record)
            _add_audit(
                ws_audit,
                audit_row,
                row_num,
                raw_date_text,
                reference_value,
                txn_type,
                details,
                raw_debit,
                raw_credit,
                reason,
            )
            audit_row += 1
            skip_tally[reason] = skip_tally.get(reason, 0) + 1

        if row_num in skip_row:
            add_audit(skip_row[row_num], ref10)
            continue

        parsed_date = _parse_date_cell(ws.cell(row=row_num, column=col_date))
        if parsed_date is None:
            add_audit("BLANK_OR_INVALID_DATE", raw_ref)
            continue

        debit_value = _safe_double(raw_debit)
        credit_value = _safe_double(raw_credit)

        if debit_value > 0 and credit_value > 0:
            add_audit("BOTH_DEBIT_AND_CREDIT_PRESENT", raw_ref)
            continue
        if debit_value == 0 and credit_value == 0:
            add_audit("ZERO_OR_NON_NUMERIC_AMOUNT", raw_ref)
            continue

        if debit_value > 0 and credit_value == 0:
            output_data.append(_make_payment_row(ref10, parsed_date, debit_value))
            total_debits += debit_value
        elif credit_value > 0 and debit_value == 0:
            doc_no_receipt = str(raw_ref).strip() if raw_ref is not None else ""
            if not doc_no_receipt:
                doc_no_receipt = "BNK is Blank"
            receipts.append(_make_receipt_row(doc_no_receipt, parsed_date, credit_value))
            total_credits += credit_value

    for receipt_row in receipts:
        output_data.append(receipt_row)

    output_data.sort(key=_sort_key)
    _write_rows_to_sheet(ws_out, output_data)

    closing_balance_calc = None
    variance = None
    if opening_balance is not None:
        closing_balance_calc = float(opening_balance) + total_credits - total_debits
        if closing_balance_stmt is not None:
            variance = float(closing_balance_stmt) - float(closing_balance_calc)

    _write_audit_summary(
        ws_audit,
        opening_balance,
        total_credits,
        total_debits,
        closing_balance_calc,
        closing_balance_stmt,
        variance,
    )

    ws_out.column_dimensions["A"].width = ws_out.column_dimensions["A"].width or 12

    message_lines = [
        "Conversion complete.",
        f"Output rows: {len(output_data)}",
        f"Skipped rows: {len(audit_rows)}",
        f"Total Debits: {total_debits:,.2f}",
        f"Total Credits: {total_credits:,.2f}",
    ]
    if opening_balance is not None:
        message_lines.append(f"Opening Balance: {opening_balance:,.2f}")
    if closing_balance_stmt is not None:
        message_lines.append(f"Statement Closing: {closing_balance_stmt:,.2f}")
    if closing_balance_calc is not None:
        message_lines.append(f"Calculated Closing: {closing_balance_calc:,.2f}")
    if variance is not None:
        message_lines.append(f"Variance: {variance:,.2f}")
    if skip_tally:
        message_lines.append("")
        message_lines.append("Skip breakdown:")
        for reason, count in sorted(skip_tally.items(), key=lambda item: (-item[1], item[0])):
            message_lines.append(f"  {reason}: {count}")

    return ConversionResult(
        success=True,
        message="\n".join(message_lines),
        output_rows=len(output_data),
        skipped_rows=len(audit_rows),
        total_debits=total_debits,
        total_credits=total_credits,
        opening_balance=opening_balance,
        closing_balance_stmt=closing_balance_stmt,
        closing_balance_calc=closing_balance_calc,
        variance=variance,
        output_data=output_data,
        audit_rows=audit_rows,
    )
