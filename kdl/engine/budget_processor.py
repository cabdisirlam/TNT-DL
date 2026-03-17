"""
GOK IFMIS Statement of Budget Execution - Python Processor.

Python equivalent of the Budget VBA macro (Version 3.0).
Reads one or more IFMIS budget worksheets and produces a formatted
output workbook (budget.xlsx) with the same layout and formulas as
the original VBA macro.
"""

from dataclasses import dataclass
from typing import Optional


@dataclass
class BudgetResult:
    success: bool
    message: str
    workbook: Optional[object] = None  # openpyxl Workbook on success


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _clean_cell(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def _side(style="thin", color="BFBFBF"):
    from openpyxl.styles import Side
    return Side(style=style, color=color)


def _border_thin():
    from openpyxl.styles import Border
    s = _side("thin", "BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _border_medium_bottom():
    from openpyxl.styles import Border
    return Border(bottom=_side("medium", "000000"))


def _border_medium_all():
    from openpyxl.styles import Border
    s = _side("medium", "000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _blue_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="0070C0")


def _grey_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="F2F2F2")


def _white_fill():
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor="FFFFFF")


def _font(bold=False, size=9, color="000000", name="Calibri"):
    from openpyxl.styles import Font
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h="general", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Row formatters  (mirror VBA FmtData / FmtTotal / FmtBalance / WriteSection)
# ---------------------------------------------------------------------------

_N_FMT = '#,##0.00_);[Red](#,##0.00);"-"'
_P_FMT = '0.00%;[Red](0.00%);"-"'


def _fmt_section_header(ws, r: int, text: str):
    """Merge A:I, blue bg, white bold text – mirrors WriteSection."""
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(r, 1)
    c.value = text
    c.font = _font(bold=True, color="FFFFFF")
    c.fill = _blue_fill()
    c.alignment = _align("center")
    ws.row_dimensions[r].height = 15


def _fmt_data_row(ws, r: int, use_grey: bool):
    """Alternating grey/white data row – mirrors VBA FmtData."""
    fill = _grey_fill() if use_grey else _white_fill()
    border = _border_thin()
    for j in range(1, 10):
        c = ws.cell(r, j)
        c.fill = fill
        c.font = _font()
        c.border = border
    ws.cell(r, 1).alignment = _align("left")
    ws.cell(r, 2).alignment = _align("center")
    for j in range(3, 9):
        c = ws.cell(r, j)
        c.alignment = _align("right")
        c.number_format = _N_FMT
    ws.cell(r, 9).alignment = _align("right")
    ws.cell(r, 9).number_format = _P_FMT
    ws.row_dimensions[r].height = 15


def _fmt_total_row(ws, r: int):
    """Bold total row with medium bottom border – mirrors VBA FmtTotal."""
    for j in range(1, 10):
        c = ws.cell(r, j)
        c.fill = _white_fill()
        c.font = _font(bold=True)
        c.border = _border_medium_bottom()
    ws.cell(r, 1).alignment = _align("center")
    for j in range(3, 9):
        c = ws.cell(r, j)
        c.alignment = _align("right")
        c.number_format = _N_FMT
    ws.cell(r, 9).alignment = _align("right")
    ws.cell(r, 9).number_format = _P_FMT
    ws.row_dimensions[r].height = 15.75


def _fmt_balance_row(ws, r: int):
    """Grey bg, blue bold text, medium border – mirrors VBA FmtBalance."""
    border = _border_medium_all()
    for j in range(1, 10):
        c = ws.cell(r, j)
        c.fill = _grey_fill()
        c.font = _font(bold=True, color="0070C0")
        c.alignment = _align("center")
        c.border = border
    for j in range(3, 9):
        ws.cell(r, j).number_format = _N_FMT
    ws.cell(r, 9).number_format = _P_FMT
    ws.row_dimensions[r].height = 15.75


# ---------------------------------------------------------------------------
# Single-sheet processor
# ---------------------------------------------------------------------------

def _process_one_sheet(ws_in, ws_out) -> str:
    """
    Process one IFMIS budget input worksheet (ws_in) and write formatted
    output to ws_out (an empty sheet in the output workbook).
    Returns a one-line summary string.
    Raises ValueError if the sheet is not a recognisable budget sheet.
    """
    last_row = ws_in.max_row or 1

    # ── STEP 1: Read header info ──────────────────────────────────────
    title_text = entity_text = period_text = ""
    for i in range(1, 7):
        for j in range(1, 10):
            v = ws_in.cell(i, j).value
            if v is None:
                continue
            s = _clean_cell(v)
            sl = s.lower()
            if "statment" in sl or "statement" in sl:
                title_text = s
            if "entity:" in sl:
                entity_text = s
            if "current period" in sl:
                period_text = s

    # ── Locate RECEIPTS / PAYMENTS sections ───────────────────────────
    receipts_row = payments_row = rec_total_row = pay_total_row = 0
    in_r = in_p = False

    for i in range(1, last_row + 1):
        cell_val = _clean_cell(ws_in.cell(i, 1).value)
        uval = cell_val.upper()
        uval_nospace = uval.replace(" ", "")

        if uval == "RECEIPTS":
            receipts_row = i; in_r = True; in_p = False
        if uval == "PAYMENTS":
            payments_row = i; in_p = True; in_r = False
        if in_r and uval_nospace == "TOTAL":
            rec_total_row = i; in_r = False
        if in_p and uval_nospace == "TOTAL":
            pay_total_row = i; in_p = False

    if not receipts_row or not payments_row:
        raise ValueError("No RECEIPTS/PAYMENTS sections found – not a budget sheet.")
    if not rec_total_row:
        rec_total_row = payments_row      # fallback
    if not pay_total_row:
        pay_total_row = last_row          # fallback

    # ── Collect RECEIPT items ──────────────────────────────────────────
    r_items = []
    for i in range(receipts_row + 1, rec_total_row):
        cell_val = _clean_cell(ws_in.cell(i, 1).value)
        if not cell_val:
            continue
        is_exch = "exchequer" in cell_val.lower()
        if not is_exch:
            all_zero = True
            for j in range(3, 10):
                v = ws_in.cell(i, j).value
                if v is not None and v != 0 and v != "":
                    all_zero = False
                    break
            if all_zero:
                continue
        r_items.append({
            "label": cell_val,
            "note":  ws_in.cell(i, 2).value,
            "C":     ws_in.cell(i, 3).value,
            "D":     ws_in.cell(i, 4).value,
            "E":     ws_in.cell(i, 5).value,
            "G":     ws_in.cell(i, 7).value,
            "is_exch": is_exch,
        })

    # ── Collect PAYMENT items ──────────────────────────────────────────
    p_items = []
    for i in range(payments_row + 1, pay_total_row):
        cell_val = _clean_cell(ws_in.cell(i, 1).value)
        if not cell_val:
            continue
        all_zero = True
        for j in range(3, 10):
            v = ws_in.cell(i, j).value
            if v is not None and v != 0 and v != "":
                all_zero = False
                break
        if all_zero:
            continue
        p_items.append({
            "label": cell_val,
            "note":  ws_in.cell(i, 2).value,
            "C":     ws_in.cell(i, 3).value,
            "D":     ws_in.cell(i, 4).value,
            "E":     ws_in.cell(i, 5).value,
            "G":     ws_in.cell(i, 7).value,
        })

    # ── Collect footer lines ───────────────────────────────────────────
    footer_items = []
    for i in range(pay_total_row + 1, last_row + 1):
        cell_val = _clean_cell(ws_in.cell(i, 1).value)
        if not cell_val:
            continue
        cv_lower = cell_val.lower()
        if any(k in cv_lower for k in [
            "prepared", "reviewed by", "approved by", "printed on", "printed by"
        ]):
            col_e = _clean_cell(ws_in.cell(i, 2).value)
            if not col_e:
                col_e = _clean_cell(ws_in.cell(i, 5).value)
            footer_items.append({
                "text":  cell_val,
                "col_d": _clean_cell(ws_in.cell(i, 4).value),
                "col_e": col_e,
            })

    # ── STEP 2: Build output sheet ─────────────────────────────────────

    # Rows 1-3: Titles  (merged A:I, centred bold 11pt)
    for idx, txt in enumerate([title_text, entity_text, period_text]):
        r = idx + 1
        ws_out.merge_cells(f"A{r}:I{r}")
        c = ws_out.cell(r, 1)
        c.value = txt
        c.font = _font(bold=True, size=11)
        c.alignment = _align("center")
        ws_out.row_dimensions[r].height = 18

    # Row 5: Column headers (blue bg, white bold, wrap)
    cr = 5
    headers = [
        "Description",
        "Note",
        "Printed Estimate",
        "Reallocation /\nTransfer",
        "Supplementary\nEstimates",
        "Final Approved\nEstimate (Net)",
        "Actual",
        "Budget Utilization\nDifferences",
        "% of\nUtilization",
    ]
    for j, hdr in enumerate(headers, 1):
        c = ws_out.cell(cr, j)
        c.value = hdr
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _blue_fill()
        c.alignment = _align("center", wrap=True)
    ws_out.row_dimensions[cr].height = 36

    # Row 6: Sub-headers
    cr = 6
    sub_hdrs = ["", "", "a", "b", "c", "d = a+b+c", "e", "f = d - e", "g = e/d"]
    for j, sh in enumerate(sub_hdrs, 1):
        c = ws_out.cell(cr, j)
        c.value = sh
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _blue_fill()
        c.alignment = _align("center")
    ws_out.row_dimensions[cr].height = 13.5

    # Row 7: RECEIPTS section header
    cr = 7
    _fmt_section_header(ws_out, cr, "RECEIPTS")

    # Receipt data rows
    first_r = cr + 1
    exch_r = 0
    alt = True
    for item in r_items:
        cr += 1
        ws_out.cell(cr, 1).value = item["label"]
        if item["note"] is not None:
            ws_out.cell(cr, 2).value = item["note"]
        if not item["is_exch"]:
            if item["C"] is not None:
                ws_out.cell(cr, 3).value = item["C"]
            if item["D"] is not None:
                ws_out.cell(cr, 4).value = item["D"]
            if item["E"] is not None:
                ws_out.cell(cr, 5).value = item["E"]
        if item["G"] is not None:
            ws_out.cell(cr, 7).value = item["G"]
        if item["is_exch"]:
            exch_r = cr
        _fmt_data_row(ws_out, cr, alt)
        alt = not alt
    last_r = cr

    # Total Receipts
    cr += 1
    rec_tot = cr
    ws_out.cell(cr, 1).value = "Total Receipts"
    _fmt_total_row(ws_out, cr)

    # Blank row then PAYMENTS
    cr += 2
    _fmt_section_header(ws_out, cr, "PAYMENTS")

    # Payment data rows
    first_p = cr + 1
    alt = True
    for item in p_items:
        cr += 1
        ws_out.cell(cr, 1).value = item["label"]
        if item["note"] is not None:
            ws_out.cell(cr, 2).value = item["note"]
        if item["C"] is not None:
            ws_out.cell(cr, 3).value = item["C"]
        if item["D"] is not None:
            ws_out.cell(cr, 4).value = item["D"]
        if item["E"] is not None:
            ws_out.cell(cr, 5).value = item["E"]
        if item["G"] is not None:
            ws_out.cell(cr, 7).value = item["G"]
        _fmt_data_row(ws_out, cr, alt)
        alt = not alt
    last_p = cr

    # Total Payments
    cr += 1
    pay_tot = cr
    ws_out.cell(cr, 1).value = "Total Payments"
    _fmt_total_row(ws_out, cr)

    # Blank row then Balance Check
    cr += 2
    bal_r = cr
    ws_out.cell(cr, 1).value = "Balance Check  (Receipts = Payments)"
    _fmt_balance_row(ws_out, cr)

    # ── STEP 3: Formulas ──────────────────────────────────────────────

    # Payment rows: F = C+D+E,  H = F-G,  I = G/F
    for r in range(first_p, last_p + 1):
        ws_out.cell(r, 6).value = f"=C{r}+D{r}+E{r}"
        ws_out.cell(r, 8).value = f"=F{r}-G{r}"
        ws_out.cell(r, 9).value = f"=IFERROR(G{r}/F{r},0)"

    # Receipt rows (non-Exchequer): same formulas
    for r in range(first_r, last_r + 1):
        if r != exch_r:
            ws_out.cell(r, 6).value = f"=C{r}+D{r}+E{r}"
            ws_out.cell(r, 8).value = f"=F{r}-G{r}"
            ws_out.cell(r, 9).value = f"=IFERROR(G{r}/F{r},0)"

    # Payments Total row
    ws_out.cell(pay_tot, 3).value = f"=SUM(C{first_p}:C{last_p})"
    ws_out.cell(pay_tot, 4).value = f"=SUM(D{first_p}:D{last_p})"
    ws_out.cell(pay_tot, 5).value = f"=SUM(E{first_p}:E{last_p})"
    ws_out.cell(pay_tot, 6).value = f"=C{pay_tot}+D{pay_tot}+E{pay_tot}"
    ws_out.cell(pay_tot, 7).value = f"=SUM(G{first_p}:G{last_p})"
    ws_out.cell(pay_tot, 8).value = f"=F{pay_tot}-G{pay_tot}"
    # col I intentionally blank on Total Payments (mirrors VBA)

    # Exchequer row: balancing formula (C = PayTot_C minus other receipt C's)
    if exch_r > 0:
        other_rows = [r for r in range(first_r, last_r + 1) if r != exch_r]
        if other_rows:
            sC = "+".join(f"C{r}" for r in other_rows)
            sD = "+".join(f"D{r}" for r in other_rows)
            sE = "+".join(f"E{r}" for r in other_rows)
        else:
            sC = sD = sE = "0"
        ws_out.cell(exch_r, 3).value = f"=C{pay_tot}-({sC})"
        ws_out.cell(exch_r, 4).value = f"=D{pay_tot}-({sD})"
        ws_out.cell(exch_r, 5).value = f"=E{pay_tot}-({sE})"
        ws_out.cell(exch_r, 6).value = f"=C{exch_r}+D{exch_r}+E{exch_r}"
        ws_out.cell(exch_r, 8).value = f"=F{exch_r}-G{exch_r}"
        ws_out.cell(exch_r, 9).value = f"=IFERROR(G{exch_r}/F{exch_r},0)"

    # Receipts Total = Payments Total (budget must balance)
    ws_out.cell(rec_tot, 3).value = f"=C{pay_tot}"
    ws_out.cell(rec_tot, 4).value = f"=D{pay_tot}"
    ws_out.cell(rec_tot, 5).value = f"=E{pay_tot}"
    ws_out.cell(rec_tot, 6).value = f"=F{pay_tot}"
    ws_out.cell(rec_tot, 7).value = f"=SUM(G{first_r}:G{last_r})"
    ws_out.cell(rec_tot, 8).value = f"=F{rec_tot}-G{rec_tot}"
    ws_out.cell(rec_tot, 9).value = f"=IFERROR(G{rec_tot}/F{rec_tot},0)"

    # Balance Check row: ✓ OK or difference
    check = "\u2713"
    for k, col in enumerate(["C", "D", "E", "F", "G", "H", "I"]):
        ws_out.cell(bal_r, 3 + k).value = (
            f'=IF(ABS({col}{rec_tot}-{col}{pay_tot})<0.01,"{check} OK",'
            f'{col}{rec_tot}-{col}{pay_tot})'
        )

    # ── STEP 4: Footer ────────────────────────────────────────────────
    cr = bal_r + 2

    for fi in footer_items:
        if "prepared" in fi["text"].lower():
            cr += 1
            ws_out.merge_cells(f"A{cr}:I{cr}")
            c = ws_out.cell(cr, 1)
            c.value = fi["text"]
            c.font = _font(bold=True)
            ws_out.row_dimensions[cr].height = 13.5

    for sig_label in ["Prepared By:", "Reviewed By:", "Approved By:"]:
        cr += 1
        ws_out.cell(cr, 1).value = sig_label
        ws_out.cell(cr, 1).font = _font(bold=True)
        ws_out.cell(cr, 4).value = "Date:"
        ws_out.cell(cr, 4).font = _font(bold=True)
        ws_out.row_dimensions[cr].height = 24
        cr += 1  # blank row for signature space

    cr += 1
    for fi in footer_items:
        if "printed on" in fi["text"].lower():
            ws_out.cell(cr, 1).value = fi["text"]
            ws_out.cell(cr, 1).font = _font(size=9)
            if fi["col_e"]:
                ws_out.merge_cells(f"E{cr}:I{cr}")
                ws_out.cell(cr, 5).value = fi["col_e"]
                ws_out.cell(cr, 5).font = _font(size=9)

    cr += 1
    for fi in footer_items:
        if "printed by" in fi["text"].lower():
            ws_out.cell(cr, 1).value = fi["text"]
            ws_out.cell(cr, 1).font = _font(size=9)

    # ── STEP 5: Column widths + page setup ────────────────────────────
    widths = {"A": 35, "B": 6, "C": 18, "D": 16, "E": 18,
              "F": 22, "G": 18, "H": 22, "I": 14}
    for col, w in widths.items():
        ws_out.column_dimensions[col].width = w

    ws_out.freeze_panes = "A7"

    ws_out.page_setup.orientation = "landscape"
    ws_out.page_setup.fitToPage = True
    ws_out.page_setup.fitToWidth = 1
    ws_out.page_setup.fitToHeight = 0
    ws_out.print_title_rows = "1:6"

    # ── STEP 6: Name the sheet ────────────────────────────────────────
    new_name = "Combined"
    tl = title_text.lower()
    if "recurrent" in tl:
        new_name = "Recurrent Expenditure"
    elif "development" in tl:
        new_name = "Development Expenditure"

    # Avoid duplicate names in the output workbook
    existing = {s.title for s in ws_out.parent.worksheets if s is not ws_out}
    candidate = new_name
    suffix = 2
    while candidate in existing:
        candidate = f"{new_name} ({suffix})"
        suffix += 1
    try:
        ws_out.title = candidate
    except Exception:
        pass

    return (
        f"{len(r_items)} receipt item(s), {len(p_items)} payment item(s) "
        f"\u2192 \"{ws_out.title}\""
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def process_budget_sheets(wb_in, sheet_names: list) -> BudgetResult:
    """
    Process one or more IFMIS budget sheets from *wb_in* and return
    a BudgetResult containing a formatted openpyxl Workbook.
    """
    try:
        import openpyxl
    except ImportError:
        return BudgetResult(False, "openpyxl is not installed.")

    from openpyxl import Workbook
    wb_out = Workbook()
    wb_out.remove(wb_out.active)   # start with no sheets

    processed = []
    errors = []

    for sname in sheet_names:
        try:
            ws_in = wb_in[sname]
            ws_out = wb_out.create_sheet(title=sname)
            summary = _process_one_sheet(ws_in, ws_out)
            processed.append(f"  \u2022 {sname}: {summary}")
        except Exception as exc:
            import traceback
            errors.append(f"  \u2022 {sname}: {exc}\n{traceback.format_exc()}")

    if not processed:
        return BudgetResult(
            success=False,
            message="No sheets processed successfully.\n\n" + "\n".join(errors),
        )

    parts = [f"Processed {len(processed)} sheet(s):\n" + "\n".join(processed)]
    if errors:
        parts.append(f"\n{len(errors)} sheet(s) failed:\n" + "\n".join(errors))

    return BudgetResult(
        success=True,
        message="\n".join(parts),
        workbook=wb_out,
    )
