"""Helpers for importing CSV/HTML/XLS tabular files into openpyxl workbooks."""

from __future__ import annotations

import csv
import io
import os
import re
import tempfile
from datetime import date, datetime
from html.parser import HTMLParser


class _HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self._table_depth = 0
        self._current_table: list[list[str]] | None = None
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            if self._table_depth == 0:
                self._current_table = []
            self._table_depth += 1
            return

        if self._table_depth != 1:
            return

        if tag == "tr":
            self._current_row = []
        elif tag in ("td", "th"):
            self._current_cell = []
        elif tag == "br" and self._current_cell is not None:
            self._current_cell.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == "table":
            if self._table_depth == 1 and self._current_table:
                if any(row for row in self._current_table):
                    self.tables.append(self._current_table)
                self._current_table = None
            if self._table_depth > 0:
                self._table_depth -= 1
            return

        if self._table_depth != 1:
            return

        if tag in ("td", "th") and self._current_row is not None and self._current_cell is not None:
            text = "".join(self._current_cell)
            text = "\n".join(part.strip() for part in text.splitlines() if part.strip())
            self._current_row.append(text)
            self._current_cell = None
        elif tag == "tr" and self._current_table is not None and self._current_row is not None:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = None

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell.append(data)


def _sheet_safe_name(name: str, fallback: str) -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\\\' else ch for ch in str(name).strip())
    cleaned = cleaned.strip("'")
    return (cleaned or fallback)[:31]


def _base_name(filepath: str) -> str:
    return os.path.splitext(os.path.basename(filepath))[0]


def detect_source_format(filepath: str) -> str:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return "csv"
    if ext in (".html", ".htm"):
        return "html"

    try:
        with open(filepath, "rb") as f:
            sample = f.read(16384)
    except OSError:
        return "excel"

    if sample.startswith(b"PK\x03\x04"):
        return "excel"
    if sample.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "excel"

    lowered = sample.decode("latin-1", errors="ignore").lower()
    html_markers = (
        "<html",
        "<table",
        "<tr",
        "<td",
        "<th",
        "<!doctype html",
        "urn:schemas-microsoft-com:office:excel",
        "content-type",
    )
    if any(marker in lowered for marker in html_markers):
        return "html"

    return "excel"


def read_text_with_fallbacks(filepath: str) -> str:
    with open(filepath, "rb") as f:
        raw = f.read()

    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="replace")


def iter_csv_rows(filepath: str):
    text = read_text_with_fallbacks(filepath)
    # Auto-detect delimiter (comma, semicolon, tab, pipe) from the first 4 KB
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t|')
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ','
    with io.StringIO(text, newline="") as stream:
        yield from csv.reader(stream, delimiter=delimiter)


def load_html_tables(filepath: str) -> list[list[list[str]]]:
    text = read_text_with_fallbacks(filepath)
    parser = _HTMLTableParser()
    parser.feed(text)
    parser.close()
    return parser.tables


_NUMERIC_RE = re.compile(r"^\(?-?\d[\d,]*\.?\d*\)?%?$")
_DATE_FORMATS = (
    "%d-%b-%Y",
    "%d-%B-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y-%m-%d",
)


def _coerce_tabular_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float, datetime, date)):
        return value

    text = str(value).strip()
    if text == "":
        return ""

    compact = text.replace("\u00a0", " ").strip()
    candidate = compact.replace(",", "")

    if _NUMERIC_RE.match(compact):
        is_percent = compact.endswith("%")
        if is_percent:
            candidate = candidate[:-1]
        if candidate.startswith("(") and candidate.endswith(")"):
            candidate = "-" + candidate[1:-1]
        try:
            number = float(candidate)
            if is_percent:
                return number / 100.0
            if "." not in candidate and "e" not in candidate.lower():
                return int(number)
            return number
        except ValueError:
            pass

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(compact, fmt)
        except ValueError:
            continue

    return compact


def _append_tabular_row(ws, row_vals):
    ws.append([_coerce_tabular_value(val) for val in row_vals])


def list_source_sheet_names(filepath: str) -> list[str]:
    source_kind = detect_source_format(filepath)
    if source_kind == "csv":
        return [_sheet_safe_name(_base_name(filepath), "Sheet1")]
    if source_kind == "html":
        tables = load_html_tables(filepath)
        if not tables:
            raise RuntimeError("No HTML tables found in the selected file.")
        return [_sheet_safe_name(_base_name(filepath), "Sheet1")]
    raise ValueError(f"Unsupported tabular source: {filepath}")


def build_workbook_from_source(filepath: str):
    import openpyxl

    source_kind = detect_source_format(filepath)
    if source_kind == "csv":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = list_source_sheet_names(filepath)[0]
        for row_vals in iter_csv_rows(filepath):
            _append_tabular_row(ws, row_vals)
        return wb

    if source_kind == "html":
        tables = load_html_tables(filepath)
        if not tables:
            raise RuntimeError("No HTML tables found in the selected file.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = list_source_sheet_names(filepath)[0]
        for idx, rows in enumerate(tables):
            for row_vals in rows:
                _append_tabular_row(ws, row_vals)
            if idx < len(tables) - 1:
                ws.append([])
        return wb

    raise ValueError(f"Unsupported tabular source: {filepath}")


def build_filtered_workbook_from_excel(
    filepath: str,
    *,
    sheet_names: list[str] | None = None,
    data_only: bool = True,
    keep_links: bool = False,
):
    import openpyxl

    source_wb = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=data_only,
        keep_links=keep_links,
    )
    target_wb = openpyxl.Workbook()
    target_wb.remove(target_wb.active)

    wanted = []
    seen = set()
    if sheet_names:
        for name in sheet_names:
            if name in source_wb.sheetnames and name not in seen:
                wanted.append(name)
                seen.add(name)
    if not wanted:
        wanted = list(source_wb.sheetnames)

    try:
        for sheet_name in wanted:
            src_ws = source_wb[sheet_name]
            dst_ws = target_wb.create_sheet(title=sheet_name)
            for row_vals in src_ws.iter_rows(values_only=True):
                dst_ws.append(list(row_vals))
    finally:
        source_wb.close()

    return target_wb


def _iter_legacy_xls_connection_strings(filepath: str):
    abs_path = os.path.abspath(filepath)
    yield (
        "Microsoft.ACE.OLEDB.12.0",
        f'Provider=Microsoft.ACE.OLEDB.12.0;Data Source={abs_path};'
        'Extended Properties="Excel 8.0;HDR=NO;IMEX=1";',
    )
    yield (
        "Microsoft.Jet.OLEDB.4.0",
        f'Provider=Microsoft.Jet.OLEDB.4.0;Data Source={abs_path};'
        'Extended Properties="Excel 8.0;HDR=NO;IMEX=1";',
    )


def _open_legacy_xls_connection(filepath: str):
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("pywin32 is required for legacy .xls support.") from exc

    errors = []
    for provider_name, conn_str in _iter_legacy_xls_connection_strings(filepath):
        conn = None
        try:
            conn = win32com.client.Dispatch("ADODB.Connection")
            conn.Open(conn_str)
            return conn, provider_name
        except Exception as exc:
            errors.append(f"{provider_name}: {exc}")
            if conn is not None:
                try:
                    conn.Close()
                except Exception:
                    pass

    raise RuntimeError(
        "Could not open this legacy .xls workbook via ACE/Jet. "
        "Install the Microsoft Access Database Engine or open it in Excel and save as .xlsx."
        + (f" Details: {' | '.join(errors)}" if errors else "")
    )


def _extract_ado_sheet_name(table_name: str) -> str:
    name = str(table_name or "").strip()
    if name.startswith("'") and name.endswith("'") and len(name) >= 2:
        name = name[1:-1]
    if name.endswith("$"):
        name = name[:-1]
    return name


def _iter_legacy_xls_tables(conn):
    recordset = None
    try:
        recordset = conn.OpenSchema(20)
        while not recordset.EOF:
            raw_name = recordset.Fields("TABLE_NAME").Value
            table_type = str(recordset.Fields("TABLE_TYPE").Value or "").upper()
            table_name = str(raw_name or "")
            if (
                table_type == "TABLE"
                and "$" in table_name
                and not table_name.startswith("_xlnm")
            ):
                yield table_name
            recordset.MoveNext()
    finally:
        if recordset is not None:
            try:
                recordset.Close()
            except Exception:
                pass


def list_legacy_xls_sheet_names(filepath: str) -> list[str]:
    conn = None
    try:
        conn, _ = _open_legacy_xls_connection(filepath)
        names = []
        seen: set[str] = set()
        for table_name in _iter_legacy_xls_tables(conn):
            name = _extract_ado_sheet_name(table_name)
            if name and name not in seen:
                seen.add(name)
                names.append(name)
        if not names:
            raise RuntimeError("No worksheets were found in the selected .xls workbook.")
        return names
    finally:
        if conn is not None:
            try:
                conn.Close()
            except Exception:
                pass


def _ado_cell_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%b-%y", "%d-%b-%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return value


def build_workbook_from_legacy_xls(filepath: str, sheet_names: list[str] | None = None):
    import openpyxl
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("pywin32 is required for legacy .xls support.") from exc

    conn = None
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wanted = {name.strip().lower() for name in (sheet_names or []) if str(name).strip()}
    try:
        conn, _ = _open_legacy_xls_connection(filepath)
        for table_name in _iter_legacy_xls_tables(conn):
            sheet_name = _extract_ado_sheet_name(table_name)
            if not sheet_name:
                continue
            if wanted and sheet_name.strip().lower() not in wanted:
                continue
            ws = wb.create_sheet(title=_sheet_safe_name(sheet_name, "Sheet1"))
            rs = None
            try:
                rs = win32com.client.Dispatch("ADODB.Recordset")
                rs.Open(f"SELECT * FROM [{table_name}]", conn, 1, 1)
                field_count = rs.Fields.Count
                while not rs.EOF:
                    row_vals = [_ado_cell_value(rs.Fields(i).Value) for i in range(field_count)]
                    while row_vals and row_vals[-1] is None:
                        row_vals.pop()
                    ws.append(row_vals)
                    rs.MoveNext()
            finally:
                if rs is not None:
                    try:
                        rs.Close()
                    except Exception:
                        pass
        if not wb.sheetnames:
            raise RuntimeError("No worksheets were found in the selected .xls workbook.")
        return wb
    finally:
        if conn is not None:
            try:
                conn.Close()
            except Exception:
                pass


def convert_legacy_xls_to_temp_xlsx(filepath: str) -> str:
    excel_app = None
    excel_wb = None
    pythoncom = None
    temp_path = ""
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        excel_wb = excel_app.Workbooks.Open(
            os.path.abspath(filepath),
            UpdateLinks=0,
            ReadOnly=True,
        )
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        os.unlink(temp_path)
        excel_wb.SaveAs(temp_path, 51)
        return temp_path
    except Exception as exc:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass
        raise RuntimeError(
            "Could not convert this legacy .xls workbook. Open it in Excel and save as .xlsx, "
            "or make sure Excel is installed."
        ) from exc
    finally:
        if excel_wb is not None:
            try:
                excel_wb.Close(False)
            except Exception:
                pass
        if excel_app is not None:
            try:
                excel_app.Quit()
            except Exception:
                pass
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def load_workbook_from_source(
    filepath: str,
    *,
    sheet_names: list[str] | None = None,
    data_only: bool = True,
    read_only: bool = False,
    keep_links: bool = False,
    keep_vba: bool = False,
):
    import openpyxl

    source_kind = detect_source_format(filepath)
    if source_kind in ("csv", "html"):
        return build_workbook_from_source(filepath)

    if os.path.splitext(filepath)[1].lower() == ".xls":
        try:
            return build_workbook_from_legacy_xls(filepath, sheet_names=sheet_names)
        except Exception:
            tmp_path = convert_legacy_xls_to_temp_xlsx(filepath)
            try:
                wb = openpyxl.load_workbook(
                    tmp_path,
                    data_only=data_only,
                    read_only=read_only,
                    keep_links=keep_links,
                    keep_vba=False,
                )
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
            return wb

    return openpyxl.load_workbook(
        filepath,
        data_only=data_only,
        read_only=read_only,
        keep_links=keep_links,
        keep_vba=keep_vba,
    )
