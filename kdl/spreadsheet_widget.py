"""
KDL Spreadsheet Widget
Custom spreadsheet grid built on QTableWidget with cell type highlighting,
Excel/CSV import, and row/column selection.
"""

import csv
import os
from typing import Dict, List, Optional, Tuple
from PySide6.QtWidgets import (
    QTableWidget, QTableWidgetItem, QHeaderView, QMenu,
    QFileDialog, QApplication, QMessageBox, QTableWidgetSelectionRange,
    QDialog, QVBoxLayout, QGridLayout, QLabel, QSpinBox, QDialogButtonBox
)
from PySide6.QtCore import Qt, Signal, QItemSelectionModel
from PySide6.QtGui import QColor, QFont, QAction, QKeySequence, QBrush, QPainter, QPen


# Cell highlighting colors
COLOR_DATA = QColor("#FFFFFF")           # Plain data cells
COLOR_KEYSTROKE = QColor("#E6F3FD")      # Keystrokes
COLOR_COMMAND = QColor("#EEF7FE")        # Shortcuts/commands
COLOR_MOUSE = QColor("#FFE082")          # Mouse actions (yellow)
COLOR_KEY_COLUMN = QColor("#E3F1FC")     # Key columns
COLOR_ACTIVE_ROW = QColor("#CFE8FA")     # Current loading row
COLOR_ACTIVE_CELL = QColor("#9CCFF3")    # Current loading cell
COLOR_GUIDE_CELL = QColor("#DAEEFC")     # Selection guide cell
COLOR_CURSOR_CELL = QColor(255, 224, 130, 120)  # Yellow overlay for current cell
COLOR_CURSOR_BORDER = QColor("#D89A00")
COLOR_ACTIVE_HEADER_BG = QColor("#F5A623")       # Orange header highlight (matches DataLoad)
COLOR_ACTIVE_HEADER_FG = QColor("#FFFFFF")        # White text on orange header
BRUSH_ACTIVE_HEADER_BG = QBrush(COLOR_ACTIVE_HEADER_BG)
BRUSH_ACTIVE_HEADER_FG = QBrush(COLOR_ACTIVE_HEADER_FG)


def apply_spreadsheet_theme(dark: bool = False):
    """Update module-level cell colors for the given theme."""
    global COLOR_DATA, COLOR_KEYSTROKE, COLOR_COMMAND, COLOR_KEY_COLUMN
    global COLOR_ACTIVE_ROW, COLOR_ACTIVE_CELL, COLOR_GUIDE_CELL
    if dark:
        COLOR_DATA.setNamedColor("#1E2330")
        COLOR_KEYSTROKE.setNamedColor("#1A2A40")
        COLOR_COMMAND.setNamedColor("#1E2E44")
        COLOR_KEY_COLUMN.setNamedColor("#1A2C42")
        COLOR_ACTIVE_ROW.setNamedColor("#1A3A50")
        COLOR_ACTIVE_CELL.setNamedColor("#1A4A60")
        COLOR_GUIDE_CELL.setNamedColor("#1A3550")
    else:
        COLOR_DATA.setNamedColor("#FFFFFF")
        COLOR_KEYSTROKE.setNamedColor("#E6F3FD")
        COLOR_COMMAND.setNamedColor("#EEF7FE")
        COLOR_KEY_COLUMN.setNamedColor("#E3F1FC")
        COLOR_ACTIVE_ROW.setNamedColor("#CFE8FA")
        COLOR_ACTIVE_CELL.setNamedColor("#9CCFF3")
        COLOR_GUIDE_CELL.setNamedColor("#DAEEFC")


class _HighlightHeaderView(QHeaderView):
    """QHeaderView that respects model ForegroundRole even when QSS sets color."""

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)

    def paintSection(self, painter, rect, logicalIndex):
        model = self.model()
        if model is None:
            super().paintSection(painter, rect, logicalIndex)
            return

        bg_brush = model.headerData(logicalIndex, self.orientation(), Qt.BackgroundRole)
        if bg_brush is None:
            # No custom styling — use default QSS painting
            super().paintSection(painter, rect, logicalIndex)
            return

        # Custom painting for highlighted (active) sections
        painter.save()

        # Background fill
        painter.fillRect(rect, bg_brush)

        # Borders to match the default header look
        border_light = QColor("#D7E7F6")
        border_med = QColor("#B6D3EE")
        if self.orientation() == Qt.Horizontal:
            painter.setPen(QPen(border_light, 1))
            painter.drawLine(rect.right(), rect.top(), rect.right(), rect.bottom())
            painter.setPen(QPen(border_med, 1))
            painter.drawLine(rect.left(), rect.bottom(), rect.right(), rect.bottom())
        else:
            painter.setPen(QPen(border_light, 1))
            painter.drawLine(rect.left(), rect.bottom(), rect.right(), rect.bottom())
            painter.setPen(QPen(border_med, 1))
            painter.drawLine(rect.right(), rect.top(), rect.right(), rect.bottom())

        # Text with custom foreground color
        fg_brush = model.headerData(logicalIndex, self.orientation(), Qt.ForegroundRole)
        text = model.headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
        if text is not None:
            if fg_brush is not None:
                painter.setPen(fg_brush.color())
            else:
                painter.setPen(QColor("#4E6E8F"))
            font = painter.font()
            font.setWeight(QFont.DemiBold)
            font.setPixelSize(15)
            painter.setFont(font)
            painter.drawText(rect, Qt.AlignCenter, str(text))

        painter.restore()


class SpreadsheetWidget(QTableWidget):
    """
    Editable spreadsheet grid with keystroke syntax highlighting.
    """

    data_changed = Signal()
    cell_count_changed = Signal(int, int)  # rows, cols
    paste_completed = Signal(int, int)  # pasted_rows, pasted_cols

    DEFAULT_ROWS = 100
    DEFAULT_COLS = 20

    def __init__(self, parent=None):
        super().__init__(parent)
        self.key_columns = set()  # Column indices that are key columns
        self._active_row = -1
        self._active_col = -1
        self._active_cell = None
        self._guide_row = -1
        self._guide_col = -1
        self._guide_cell = None
        self._guide_enabled = True

        # History (undo/redo) and navigation (back/forward)
        self._undo_stack: List[dict] = []
        self._redo_stack: List[dict] = []
        self._history_limit = 80
        self._history_restoring = False
        self._history_pending_before: Optional[dict] = None
        self._cell_cache: Dict[Tuple[int, int], str] = {}
        self._dirty_cells: set = set()  # Track changed cells for incremental snapshots
        self._last_snapshot: Optional[dict] = None
        self._last_snapshot_cells: Dict[Tuple[int, int], str] = {}  # Fast lookup for incremental snapshots
        self._nav_back_stack: List[Tuple[int, int]] = []
        self._nav_forward_stack: List[Tuple[int, int]] = []
        self._nav_internal = False
        self._shift_anchor: Optional[Tuple[int, int]] = None

        # Initialize grid
        self.setRowCount(self.DEFAULT_ROWS)
        self.setColumnCount(self.DEFAULT_COLS)

        # Use custom header views that respect model ForegroundRole over QSS
        self.setHorizontalHeader(_HighlightHeaderView(Qt.Horizontal, self))
        self.setVerticalHeader(_HighlightHeaderView(Qt.Vertical, self))

        # Headers
        self._update_headers()

        # Appearance
        self.setAlternatingRowColors(False)
        self.setGridStyle(Qt.SolidLine)
        self.horizontalHeader().setDefaultSectionSize(100)
        self.verticalHeader().setDefaultSectionSize(24)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

        # Selection: allow contiguous range selection for "selected rows/cols" loads,
        # while still keeping a single active current cell for guides.
        self.setSelectionMode(QTableWidget.ContiguousSelection)
        self.setSelectionBehavior(QTableWidget.SelectItems)

        # Font
        font = QFont("Consolas", 13)
        self.setFont(font)

        # Signals
        self.cellChanged.connect(self._on_cell_changed)
        self.currentCellChanged.connect(self._on_current_cell_changed)

        # Context menu
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

        self._last_snapshot = self._capture_snapshot()

    @staticmethod
    def _column_label(index: int) -> str:
        """Convert zero-based index to Excel column label (A, Z, AA, AB, ..., AAA...)."""
        n = int(index) + 1
        chars = []
        while n > 0:
            n, rem = divmod(n - 1, 26)
            chars.append(chr(65 + rem))
        return "".join(reversed(chars))

    def _update_headers(self):
        """Set column headers as A, B, C... and row headers as 1, 2, 3..."""
        headers = [self._column_label(i) for i in range(self.columnCount())]
        self.setHorizontalHeaderLabels(headers)

    def _on_cell_changed(self, row, col):
        """Handle cell content change - apply syntax highlighting."""
        item = self.item(row, col)
        if item is None:
            return

        old_value = self._cell_cache.get((row, col), "")
        new_value = item.text()
        changed = old_value != new_value

        self._apply_cell_style(row, col, item, item.text())
        if new_value != "":
            self._cell_cache[(row, col)] = new_value
        else:
            self._cell_cache.pop((row, col), None)

        if self._active_cell == (row, col):
            item.setBackground(COLOR_ACTIVE_CELL)

        if changed and not self._history_restoring:
            self._dirty_cells.add((row, col))
            self._push_undo_snapshot(self._last_snapshot)
            self._redo_stack.clear()
            self._last_snapshot = self._capture_snapshot_incremental()

        if not self._history_restoring:
            self.data_changed.emit()

    def _on_current_cell_changed(self, row, col, prev_row, prev_col):
        """Track navigation history and keep row/column guides visible."""
        if row >= 0 and col >= 0:
            if not self._nav_internal and prev_row >= 0 and prev_col >= 0 and (prev_row, prev_col) != (row, col):
                if not self._nav_back_stack or self._nav_back_stack[-1] != (prev_row, prev_col):
                    self._nav_back_stack.append((prev_row, prev_col))
                    if len(self._nav_back_stack) > 300:
                        self._nav_back_stack = self._nav_back_stack[-300:]
                self._nav_forward_stack.clear()

            self.set_selection_guides(row, col)
            if prev_row >= 0 and prev_col >= 0:
                self.viewport().update(self.visualRect(self.model().index(prev_row, prev_col)))
            self.viewport().update(self.visualRect(self.model().index(row, col)))

    def _apply_cell_style(self, row: int, col: int, item: QTableWidgetItem, value: str):
        """Apply background style for one cell without emitting signals."""
        if item is None:
            return

        value = (value or "").strip()
        if not value:
            item.setBackground(COLOR_DATA)
            return

        if col in self.key_columns:
            item.setBackground(COLOR_KEY_COLUMN)
        elif value.startswith("\\"):
            item.setBackground(COLOR_KEYSTROKE)
        elif value.startswith("*MC") or value.startswith("*MR"):
            item.setBackground(COLOR_MOUSE)
        elif value.startswith("*"):
            item.setBackground(COLOR_COMMAND)
        else:
            item.setBackground(COLOR_DATA)

    def _ensure_item(self, row: int, col: int) -> Optional[QTableWidgetItem]:
        """Ensure a QTableWidgetItem exists for the given cell."""
        if row < 0 or col < 0 or row >= self.rowCount() or col >= self.columnCount():
            return None
        item = self.item(row, col)
        if item is None:
            item = QTableWidgetItem("")
            self.setItem(row, col, item)
        return item

    def set_key_columns(self, columns: set):
        """Set which columns are key columns and re-highlight."""
        self.key_columns = columns
        self._refresh_highlighting()

    def _refresh_highlighting(self):
        """Re-apply highlighting to cells with content (uses cache for speed)."""
        self.blockSignals(True)
        # Only visit cells known to have content via the cache
        for (row, col), text in self._cell_cache.items():
            if row >= self.rowCount() or col >= self.columnCount():
                continue
            item = self.item(row, col)
            if item:
                self._apply_cell_style(row, col, item, text)
        if self._active_cell:
            a_item = self._ensure_item(self._active_cell[0], self._active_cell[1])
            if a_item is not None:
                a_item.setBackground(COLOR_ACTIVE_CELL)
        self.blockSignals(False)
        if not self._history_restoring:
            self.data_changed.emit()

    def _capture_snapshot(self) -> dict:
        """Capture sparse table state for undo/redo (full scan)."""
        cells: Dict[Tuple[int, int], str] = {}
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if not item:
                    continue
                text = item.text()
                if text != "":
                    cells[(row, col)] = text
        self._last_snapshot_cells = dict(cells)
        self._dirty_cells.clear()
        return {
            "rows": self.rowCount(),
            "cols": self.columnCount(),
            "cells": cells,
            "key_columns": set(self.key_columns),
        }

    def _capture_snapshot_incremental(self) -> dict:
        """Fast incremental snapshot using only dirty cells."""
        if not self._last_snapshot_cells and not self._dirty_cells:
            return self._capture_snapshot()

        cells = dict(self._last_snapshot_cells)
        for (row, col) in self._dirty_cells:
            item = self.item(row, col)
            if item and item.text() != "":
                cells[(row, col)] = item.text()
            else:
                cells.pop((row, col), None)
        self._last_snapshot_cells = dict(cells)
        self._dirty_cells.clear()
        return {
            "rows": self.rowCount(),
            "cols": self.columnCount(),
            "cells": cells,
            "key_columns": set(self.key_columns),
        }

    def _snapshots_equal(self, a: Optional[dict], b: Optional[dict]) -> bool:
        if a is None or b is None:
            return False
        return (
            a.get("rows") == b.get("rows")
            and a.get("cols") == b.get("cols")
            and a.get("key_columns", set()) == b.get("key_columns", set())
            and a.get("cells", {}) == b.get("cells", {})
        )

    def _push_undo_snapshot(self, snap: Optional[dict]):
        if snap is None:
            return
        self._undo_stack.append(snap)
        if len(self._undo_stack) > self._history_limit:
            self._undo_stack = self._undo_stack[-self._history_limit:]

    def _begin_history_action(self):
        if self._history_restoring:
            return
        self._history_pending_before = self._capture_snapshot()

    def _end_history_action(self):
        if self._history_restoring:
            self._history_pending_before = None
            return
        if self._history_pending_before is None:
            return
        after = self._capture_snapshot()
        if not self._snapshots_equal(self._history_pending_before, after):
            self._push_undo_snapshot(self._history_pending_before)
            self._redo_stack.clear()
            self._last_snapshot = after
            self._rebuild_cell_cache()
        self._history_pending_before = None

    def _rebuild_cell_cache(self):
        cache: Dict[Tuple[int, int], str] = {}
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item is None:
                    continue
                text = item.text()
                if text != "":
                    cache[(row, col)] = text
        self._cell_cache = cache
        self._dirty_cells.clear()

    def _apply_snapshot(self, snap: dict):
        self._history_restoring = True
        self.blockSignals(True)
        self.setUpdatesEnabled(False)
        try:
            rows = max(1, int(snap.get("rows", self.DEFAULT_ROWS)))
            cols = max(1, int(snap.get("cols", self.DEFAULT_COLS)))
            self.clear()
            self.setRowCount(rows)
            self.setColumnCount(cols)
            self._update_headers()

            for (row, col), text in snap.get("cells", {}).items():
                if row < 0 or col < 0 or row >= rows or col >= cols:
                    continue
                item = QTableWidgetItem(text)
                self.setItem(row, col, item)

            self.key_columns = set(snap.get("key_columns", set()))
            self._refresh_highlighting()
        finally:
            self.setUpdatesEnabled(True)
            self.blockSignals(False)
            self._history_restoring = False

        self._rebuild_cell_cache()
        self._dirty_cells.clear()
        self._last_snapshot = self._capture_snapshot()
        self.data_changed.emit()

    def undo(self) -> bool:
        if not self._undo_stack:
            return False
        current = self._capture_snapshot()
        target = self._undo_stack.pop()
        self._redo_stack.append(current)
        self._apply_snapshot(target)
        return True

    def redo(self) -> bool:
        if not self._redo_stack:
            return False
        current = self._capture_snapshot()
        target = self._redo_stack.pop()
        self._undo_stack.append(current)
        self._apply_snapshot(target)
        return True

    def go_back(self) -> bool:
        if not self._nav_back_stack:
            return False
        target = self._nav_back_stack.pop()
        cur_row, cur_col = self.currentRow(), self.currentColumn()
        if cur_row >= 0 and cur_col >= 0:
            self._nav_forward_stack.append((cur_row, cur_col))
        self._nav_internal = True
        try:
            self.setCurrentCell(target[0], target[1])
            item = self.item(target[0], target[1])
            if item is not None:
                self.scrollToItem(item)
        finally:
            self._nav_internal = False
        return True

    def go_forward(self) -> bool:
        if not self._nav_forward_stack:
            return False
        target = self._nav_forward_stack.pop()
        cur_row, cur_col = self.currentRow(), self.currentColumn()
        if cur_row >= 0 and cur_col >= 0:
            self._nav_back_stack.append((cur_row, cur_col))
        self._nav_internal = True
        try:
            self.setCurrentCell(target[0], target[1])
            item = self.item(target[0], target[1])
            if item is not None:
                self.scrollToItem(item)
        finally:
            self._nav_internal = False
        return True

    def _set_header_active(self, orientation, index: int, active: bool):
        """Apply/reset highlight on a header section."""
        if index < 0:
            return
        model = self.model()
        if active:
            model.setHeaderData(index, orientation, BRUSH_ACTIVE_HEADER_BG, Qt.BackgroundRole)
            model.setHeaderData(index, orientation, BRUSH_ACTIVE_HEADER_FG, Qt.ForegroundRole)
        else:
            model.setHeaderData(index, orientation, None, Qt.BackgroundRole)
            model.setHeaderData(index, orientation, None, Qt.ForegroundRole)

    def set_selection_guides(self, row: int, col: int):
        """Highlight selected row/column headers to guide the user."""
        if row < 0 or col < 0:
            return

        if self._guide_row != row:
            self._set_header_active(Qt.Vertical, self._guide_row, False)
        if self._guide_col != col:
            self._set_header_active(Qt.Horizontal, self._guide_col, False)

        previous_guide = self._guide_cell
        if previous_guide and previous_guide != (row, col):
            # Temporarily clear guide pointer so previous cell can fully reset.
            self._guide_cell = None
            self._restore_cell_background(previous_guide[0], previous_guide[1], force_plain=True)

        self._guide_row = row
        self._guide_col = col
        self._guide_cell = (row, col)
        self._set_header_active(Qt.Vertical, self._guide_row, True)
        self._set_header_active(Qt.Horizontal, self._guide_col, True)

    def _restore_cell_background(self, row: int, col: int, force_plain: bool = False):
        """Restore cell style based on value/key-column rules."""
        if row < 0 or col < 0:
            return
        item = self.item(row, col)
        if item is None:
            return
        if force_plain:
            self._apply_cell_style(row, col, item, item.text())
            return
        if self._active_cell == (row, col):
            item.setBackground(COLOR_ACTIVE_CELL)
            return
        self._apply_cell_style(row, col, item, item.text())

    def _is_cell_visible(self, row: int, col: int) -> bool:
        """Return True when the given cell is already visible in the viewport."""
        model = self.model()
        if model is None:
            return False
        index = model.index(row, col)
        if not index.isValid():
            return False
        rect = self.visualRect(index)
        return rect.isValid() and not rect.isEmpty() and self.viewport().rect().contains(rect.center())

    def set_loading_position(self, row: int, col: int, keep_visible: bool = False):
        """
        Highlight active loading position:
        - row header
        - column header
        - exact cell (yellow)
        """
        if row < 0 or col < 0 or row >= self.rowCount() or col >= self.columnCount():
            return

        # Clear previous header highlights
        if self._active_row != row:
            self._set_header_active(Qt.Vertical, self._active_row, False)
        if self._active_col != col:
            self._set_header_active(Qt.Horizontal, self._active_col, False)

        # Restore previous active cell background
        if self._active_cell and self._active_cell != (row, col):
            self._restore_cell_background(self._active_cell[0], self._active_cell[1])

        self._active_row = row
        self._active_col = col
        self._active_cell = (row, col)

        # Highlight current row and column headers
        self._set_header_active(Qt.Vertical, row, True)
        self._set_header_active(Qt.Horizontal, col, True)

        # Make the tracked cell obvious without creating empty items during load tracking.
        item = self.item(row, col)
        if item is not None:
            item.setBackground(COLOR_ACTIVE_CELL)
            if keep_visible and not self._is_cell_visible(row, col):
                self.scrollToItem(item)
        elif keep_visible and not self._is_cell_visible(row, col):
            index = self.model().index(row, col)
            if index.isValid():
                self.scrollTo(index)

    def clear_loading_position(self):
        """Clear load tracker highlights and restore original cell style."""
        self._set_header_active(Qt.Vertical, self._active_row, False)
        self._set_header_active(Qt.Horizontal, self._active_col, False)
        if self._active_cell:
            self._restore_cell_background(self._active_cell[0], self._active_cell[1])
        self._active_row = -1
        self._active_col = -1
        self._active_cell = None
        # Keep user guidance highlight on the current selection.
        row = self.currentRow()
        col = self.currentColumn()
        if row >= 0 and col >= 0:
            self.set_selection_guides(row, col)

    def highlight_loading_row(self, row: int, keep_visible: bool = False):
        """
        Backward compatible row-highlighter.
        Keeps existing behavior by highlighting the row header and first column.
        """
        col = self.currentColumn() if self.currentColumn() >= 0 else 0
        self.set_loading_position(row, col, keep_visible=keep_visible)

    def get_grid_data(self) -> list:
        """Get all cell data as a 2D list, trimmed to used range."""
        data = []
        last_row = self._find_last_row()
        last_col = self._find_last_col()
        if last_row < 0 or last_col < 0:
            return data

        for row in range(last_row + 1):
            row_data = []
            for col in range(last_col + 1):
                item = self.item(row, col)
                row_data.append(item.text() if item else "")
            data.append(row_data)

        return data

    def get_selected_range(self):
        """Get the selected range as (start_row, end_row, start_col, end_col)."""
        selection = self.selectedRanges()
        if not selection:
            return None
        r = selection[0]
        return (r.topRow(), r.bottomRow(), r.leftColumn(), r.rightColumn())

    def _find_last_row(self) -> int:
        """Find the last row that contains data (uses cache for speed)."""
        max_row = -1
        for (row, col), text in self._cell_cache.items():
            if text.strip():
                max_row = max(max_row, row)
        return max_row

    def _find_last_col(self) -> int:
        """Find the last column that contains data (uses cache for speed)."""
        max_col = -1
        for (row, col), text in self._cell_cache.items():
            if text.strip():
                max_col = max(max_col, col)
        return max_col

    def get_row_count_with_data(self) -> int:
        """Get the number of rows that contain data."""
        return max(0, self._find_last_row() + 1)

    # ── File Operations ──

    def import_csv(self, filepath: str):
        """Import data from a CSV file."""
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                return

            self._begin_history_action()

            # Resize grid if needed
            if len(rows) > self.rowCount():
                self.setRowCount(len(rows) + 10)
            max_cols = max(len(row) for row in rows)
            if max_cols > self.columnCount():
                self.setColumnCount(max_cols + 5)
                self._update_headers()

            # Populate
            self.blockSignals(True)
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    item = QTableWidgetItem(val)
                    self.setItem(r, c, item)
            self.blockSignals(False)

            self._rebuild_cell_cache()
            self._refresh_highlighting()
            self._end_history_action()
            self.data_changed.emit()

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import CSV:\n{str(e)}")

    def import_excel(self, filepath: str):
        """Import data from an Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return

            self._begin_history_action()

            # Resize grid if needed
            if len(rows) > self.rowCount():
                self.setRowCount(len(rows) + 10)
            max_cols = max(len(row) for row in rows)
            if max_cols > self.columnCount():
                self.setColumnCount(max_cols + 5)
                self._update_headers()

            # Populate
            self.blockSignals(True)
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    text = str(val) if val is not None else ""
                    item = QTableWidgetItem(text)
                    self.setItem(r, c, item)
            self.blockSignals(False)

            self._rebuild_cell_cache()
            self._refresh_highlighting()
            self._end_history_action()
            self.data_changed.emit()

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import Excel:\n{str(e)}")

    def load_from_rows(self, rows: list):
        """Load data from a list-of-lists directly into the grid (no file I/O)."""
        if not rows:
            return
        self._begin_history_action()
        if len(rows) > self.rowCount():
            self.setRowCount(len(rows) + 10)
        max_cols = max((len(r) for r in rows), default=0)
        if max_cols > self.columnCount():
            self.setColumnCount(max_cols + 5)
            self._update_headers()
        self.blockSignals(True)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                text = str(val) if val is not None else ''
                self.setItem(r, c, QTableWidgetItem(text))
        self.blockSignals(False)
        self._rebuild_cell_cache()
        self._refresh_highlighting()
        self._end_history_action()
        self.data_changed.emit()

    def export_csv(self, filepath: str):
        """Export grid data to a CSV file."""
        try:
            data = self.get_grid_data()
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(data)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export CSV:\n{str(e)}")

    def export_excel(self, filepath: str):
        """Export grid data to an Excel (.xlsx) file."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            data = self.get_grid_data()
            for r, row_data in enumerate(data, 1):
                for c, val in enumerate(row_data, 1):
                    ws.cell(row=r, column=c, value=val)
            wb.save(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel:\n{str(e)}")

    def clear_all(self):
        """Clear all cells in the spreadsheet."""
        self._begin_history_action()
        self.blockSignals(True)
        self.clearContents()
        self.blockSignals(False)
        self._cell_cache.clear()
        self._dirty_cells.clear()
        self._last_snapshot_cells.clear()
        self._end_history_action()
        self.data_changed.emit()

    # ── Context Menu ──

    def _show_context_menu(self, pos):
        """Show right-click context menu."""
        idx = self.indexAt(pos)
        if idx.isValid() and not self.selectionModel().isSelected(idx):
            self.clearSelection()
            self.setCurrentCell(idx.row(), idx.column())

        menu = QMenu(self)

        undo_action = QAction("Undo", self)
        undo_action.setShortcut(QKeySequence.Undo)
        undo_action.triggered.connect(self.undo)
        menu.addAction(undo_action)

        redo_action = QAction("Redo", self)
        redo_action.setShortcut(QKeySequence.Redo)
        redo_action.triggered.connect(self.redo)
        menu.addAction(redo_action)

        menu.addSeparator()

        cut_action = QAction("Cut", self)
        cut_action.setShortcut(QKeySequence.Cut)
        cut_action.triggered.connect(self._cut_cells)
        menu.addAction(cut_action)

        copy_action = QAction("Copy", self)
        copy_action.setShortcut(QKeySequence.Copy)
        copy_action.triggered.connect(self._copy_cells)
        menu.addAction(copy_action)

        paste_action = QAction("Paste", self)
        paste_action.setShortcut(QKeySequence.Paste)
        paste_action.triggered.connect(self._paste_cells)
        menu.addAction(paste_action)

        menu.addSeparator()

        select_all_action = QAction("Select All", self)
        select_all_action.setShortcut(QKeySequence.SelectAll)
        select_all_action.triggered.connect(self.selectAll)
        menu.addAction(select_all_action)

        menu.addSeparator()

        insert_row_action = QAction("Insert Row Below", self)
        insert_row_action.triggered.connect(self._insert_row)
        menu.addAction(insert_row_action)

        insert_col_action = QAction("Insert Column Right", self)
        insert_col_action.triggered.connect(self._insert_column_right)
        menu.addAction(insert_col_action)

        delete_row_action = QAction("Delete Selected Row(s)", self)
        delete_row_action.triggered.connect(self._delete_selected_rows)
        menu.addAction(delete_row_action)

        delete_row_range_action = QAction("Delete Row Range...", self)
        delete_row_range_action.triggered.connect(self._delete_row_range)
        menu.addAction(delete_row_range_action)

        delete_col_action = QAction("Delete Selected Column(s)", self)
        delete_col_action.triggered.connect(self._delete_selected_columns)
        menu.addAction(delete_col_action)

        menu.addSeparator()

        clear_action = QAction("Clear Selection", self)
        clear_action.triggered.connect(self._clear_selection)
        menu.addAction(clear_action)

        # Key column toggle
        col = self.currentColumn()
        menu.addSeparator()
        if col >= 0:
            col_label = self._column_label(col)
            if col in self.key_columns:
                unmark = QAction(f"Unmark Column {col_label} as Key Column", self)
                unmark.triggered.connect(lambda: self._toggle_key_column(col))
                menu.addAction(unmark)
            else:
                mark = QAction(f"Mark Column {col_label} as Key Column", self)
                mark.triggered.connect(lambda: self._toggle_key_column(col))
                menu.addAction(mark)
        else:
            no_col = QAction("Select a cell to mark key column", self)
            no_col.setEnabled(False)
            menu.addAction(no_col)

        menu.exec(self.mapToGlobal(pos))

    def _cut_cells(self):
        self._copy_cells()
        self._clear_selection()

    def _copy_cells(self):
        """Copy selected cells to clipboard in tab-separated format."""
        selection = self.selectedRanges()
        if not selection:
            return
        r = selection[0]
        text_rows = []
        for row in range(r.topRow(), r.bottomRow() + 1):
            row_data = []
            for col in range(r.leftColumn(), r.rightColumn() + 1):
                item = self.item(row, col)
                row_data.append(item.text() if item else "")
            text_rows.append("\t".join(row_data))
        QApplication.clipboard().setText("\n".join(text_rows))

    def _paste_cells(self):
        """Paste from clipboard into the grid."""
        text = QApplication.clipboard().text()
        if not text:
            return

        start_row = self.currentRow()
        start_col = self.currentColumn()
        if start_row < 0:
            start_row = 0
        if start_col < 0:
            start_col = 0

        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        rows = normalized.split("\n")
        if rows and rows[-1] == "":
            rows = rows[:-1]
        if not rows:
            return

        parsed_rows = [row_text.split("\t") for row_text in rows]
        paste_rows = len(parsed_rows)
        paste_cols = max((len(r) for r in parsed_rows), default=0)
        if paste_cols == 0:
            return

        req_rows = start_row + paste_rows
        req_cols = start_col + paste_cols
        self._begin_history_action()
        if req_rows > self.rowCount():
            self.setRowCount(req_rows + 10)
        if req_cols > self.columnCount():
            self.setColumnCount(req_cols + 5)
            self._update_headers()

        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.setUpdatesEnabled(False)
        self.blockSignals(True)
        try:
            for r_offset, cols in enumerate(parsed_rows):
                target_row = start_row + r_offset
                for c_offset, val in enumerate(cols):
                    target_col = start_col + c_offset
                    item = self.item(target_row, target_col)
                    if item is None:
                        item = QTableWidgetItem(val)
                        self.setItem(target_row, target_col, item)
                    else:
                        item.setText(val)
                    self._apply_cell_style(target_row, target_col, item, val)

                # Keep UI responsive on very large paste operations
                if (r_offset + 1) % 200 == 0:
                    QApplication.processEvents()
        finally:
            self.blockSignals(False)
            self.setUpdatesEnabled(True)
            QApplication.restoreOverrideCursor()

        self._end_history_action()
        self.data_changed.emit()
        self.paste_completed.emit(paste_rows, paste_cols)

    def _insert_row(self):
        row = self.currentRow()
        self._begin_history_action()
        self.insertRow(row + 1)
        self._end_history_action()
        self.data_changed.emit()

    def _insert_column_right(self):
        col = self.currentColumn()
        if col < 0:
            col = 0

        insert_at = col + 1
        self._begin_history_action()
        self.insertColumn(insert_at)

        if self.key_columns:
            shifted = set()
            for key_col in self.key_columns:
                if key_col >= insert_at:
                    shifted.add(key_col + 1)
                else:
                    shifted.add(key_col)
            self.key_columns = shifted

        self._update_headers()
        self._end_history_action()
        self._refresh_highlighting()

    def _delete_row(self):
        row = self.currentRow()
        if row < 0:
            return
        self._begin_history_action()
        self.removeRow(row)
        if self.rowCount() == 0:
            self.setRowCount(1)
        self._end_history_action()
        self.data_changed.emit()

    def _selected_rows(self) -> List[int]:
        rows = set()
        for selected in self.selectedRanges():
            rows.update(range(selected.topRow(), selected.bottomRow() + 1))
        if not rows and self.currentRow() >= 0:
            rows.add(self.currentRow())
        return sorted(rows)

    def _selected_columns(self) -> List[int]:
        cols = set()
        for selected in self.selectedRanges():
            cols.update(range(selected.leftColumn(), selected.rightColumn() + 1))
        if not cols and self.currentColumn() >= 0:
            cols.add(self.currentColumn())
        return sorted(cols)

    def _delete_selected_rows(self):
        rows = self._selected_rows()
        if not rows:
            return

        self._begin_history_action()
        for row in reversed(rows):
            if 0 <= row < self.rowCount():
                self.removeRow(row)
        if self.rowCount() == 0:
            self.setRowCount(1)
        self._end_history_action()
        self.data_changed.emit()

    def _prompt_delete_row_range(self, total_rows: int) -> Optional[Tuple[int, int]]:
        rows = self._selected_rows()
        if rows:
            from_default = rows[0] + 1
            to_default = rows[-1] + 1
        else:
            current_row = self.currentRow()
            from_default = current_row + 1 if current_row >= 0 else 1
            to_default = from_default

        from_default = max(1, min(total_rows, from_default))
        to_default = max(from_default, min(total_rows, to_default))

        dialog = QDialog(self)
        dialog.setWindowTitle("Delete Row Range")
        dialog.setWindowFlag(Qt.WindowCloseButtonHint, True)
        dialog.setMinimumWidth(320)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)
        layout.addWidget(QLabel("Delete rows using 1-based numbers:"))

        grid = QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(6)

        from_spin = QSpinBox(dialog)
        from_spin.setRange(1, total_rows)
        from_spin.setValue(from_default)
        from_spin.setAlignment(Qt.AlignCenter)

        to_spin = QSpinBox(dialog)
        to_spin.setRange(1, total_rows)
        to_spin.setValue(to_default)
        to_spin.setAlignment(Qt.AlignCenter)

        def _sync_to_min(value: int):
            to_spin.setMinimum(max(1, int(value)))
            if to_spin.value() < to_spin.minimum():
                to_spin.setValue(to_spin.minimum())

        from_spin.valueChanged.connect(_sync_to_min)
        _sync_to_min(from_spin.value())

        grid.addWidget(QLabel("From Row:"), 0, 0)
        grid.addWidget(from_spin, 0, 1)
        grid.addWidget(QLabel("To Row:"), 1, 0)
        grid.addWidget(to_spin, 1, 1)
        layout.addLayout(grid)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=dialog)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.Accepted:
            return None
        return from_spin.value(), to_spin.value()

    def _delete_row_range(self):
        total_rows = self.rowCount()
        if total_rows <= 0:
            return

        selected = self._prompt_delete_row_range(total_rows)
        if not selected:
            return
        from_row, to_row = selected

        count = to_row - from_row + 1
        if count <= 0:
            return

        confirm = QMessageBox.question(
            self,
            "Confirm Delete Row Range",
            f"Delete {count} row(s): {from_row} to {to_row}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return

        start_idx = from_row - 1
        end_idx = to_row - 1

        self._begin_history_action()
        for row in range(end_idx, start_idx - 1, -1):
            if 0 <= row < self.rowCount():
                self.removeRow(row)
        if self.rowCount() == 0:
            self.setRowCount(1)
        self._end_history_action()

        target_row = min(start_idx, self.rowCount() - 1)
        target_col = self.currentColumn()
        if target_col < 0:
            target_col = 0
        target_col = min(target_col, self.columnCount() - 1)
        if target_row >= 0 and target_col >= 0:
            self.setCurrentCell(target_row, target_col)

        self.data_changed.emit()

    def _delete_selected_columns(self):
        cols = self._selected_columns()
        if not cols:
            return

        self._begin_history_action()

        removed = []
        for col in reversed(cols):
            if 0 <= col < self.columnCount():
                self.removeColumn(col)
                removed.append(col)
        if self.columnCount() == 0:
            self.setColumnCount(1)

        if removed:
            removed_set = set(removed)
            remapped = set()
            for key_col in self.key_columns:
                if key_col in removed_set:
                    continue
                shift = sum(1 for c in removed_set if c < key_col)
                remapped.add(key_col - shift)
            self.key_columns = remapped

        self._update_headers()
        self._end_history_action()
        self._refresh_highlighting()

    def _clear_selection(self):
        self._begin_history_action()
        for item in self.selectedItems():
            item.setText("")
            item.setBackground(COLOR_DATA)
        self._end_history_action()

    def _toggle_key_column(self, col):
        if col < 0:
            return
        self._begin_history_action()
        if col in self.key_columns:
            self.key_columns.discard(col)
        else:
            self.key_columns.add(col)
        self._refresh_highlighting()
        self._end_history_action()

    # ── Keyboard shortcuts ──

    def keyPressEvent(self, event):
        """Handle keyboard shortcuts."""
        if (
            event.modifiers() == Qt.ShiftModifier
            and event.key() in (
                Qt.Key_Up,
                Qt.Key_Down,
                Qt.Key_Left,
                Qt.Key_Right,
            )
        ):
            self._extend_selection_with_shift(event.key())
            return

        if event.modifiers() & Qt.AltModifier and event.key() == Qt.Key_Left:
            self.go_back()
            return
        if event.modifiers() & Qt.AltModifier and event.key() == Qt.Key_Right:
            self.go_forward()
            return
        if event.matches(QKeySequence.Undo):
            self.undo()
            return
        if event.matches(QKeySequence.Redo) or (
            event.modifiers() & Qt.ControlModifier
            and event.modifiers() & Qt.ShiftModifier
            and event.key() == Qt.Key_Z
        ):
            self.redo()
            return
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_Minus:
            self._delete_row()
            return
        if (
            event.modifiers() == (Qt.ControlModifier | Qt.ShiftModifier)
            and event.key() == Qt.Key_Up
        ):
            self._jump_to_data_row(-1)
            return
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_Up:
            self._jump_to_top_row()
            return
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_Down:
            self._jump_to_data_row(1)
            return

        if event.matches(QKeySequence.Copy):
            self._copy_cells()
            return
        if event.matches(QKeySequence.Paste):
            self._paste_cells()
            return
        if event.matches(QKeySequence.Cut):
            self._cut_cells()
            return
        if event.matches(QKeySequence.SelectAll):
            self.selectAll()
            return
        if event.key() == Qt.Key_Delete:
            self._clear_selection()
            return
        if event.key() in (Qt.Key_Up, Qt.Key_Down, Qt.Key_Left, Qt.Key_Right):
            self._shift_anchor = None
        super().keyPressEvent(event)

    def mousePressEvent(self, event):
        self._shift_anchor = None
        super().mousePressEvent(event)

    def _extend_selection_with_shift(self, key: int):
        if self.rowCount() <= 0 or self.columnCount() <= 0:
            return

        row = self.currentRow()
        col = self.currentColumn()
        if row < 0 or col < 0:
            row, col = 0, 0
            self.setCurrentCell(row, col)

        if self._shift_anchor is None:
            self._shift_anchor = (row, col)

        target_row, target_col = row, col
        if key == Qt.Key_Up:
            target_row = max(0, row - 1)
        elif key == Qt.Key_Down:
            target_row = min(self.rowCount() - 1, row + 1)
        elif key == Qt.Key_Left:
            target_col = max(0, col - 1)
        elif key == Qt.Key_Right:
            target_col = min(self.columnCount() - 1, col + 1)

        if (target_row, target_col) == (row, col):
            return

        anchor_row, anchor_col = self._shift_anchor
        top = min(anchor_row, target_row)
        bottom = max(anchor_row, target_row)
        left = min(anchor_col, target_col)
        right = max(anchor_col, target_col)

        self.clearSelection()
        self.setRangeSelected(QTableWidgetSelectionRange(top, left, bottom, right), True)
        self.setCurrentCell(target_row, target_col, QItemSelectionModel.NoUpdate)
        self.scrollTo(self.model().index(target_row, target_col))

    def paintEvent(self, event):
        """Paint default grid then overlay active cursor cell."""
        super().paintEvent(event)
        row = self.currentRow()
        col = self.currentColumn()
        if row < 0 or col < 0:
            return

        model = self.model()
        if model is None:
            return

        painter = QPainter(self.viewport())
        painter.setRenderHint(QPainter.Antialiasing, False)

        # Draw focused cell with yellow overlay + border.
        idx = model.index(row, col)
        if idx.isValid():
            rect = self.visualRect(idx)
            if rect.isValid() and not rect.isEmpty():
                painter.fillRect(rect.adjusted(1, 1, -1, -1), COLOR_CURSOR_CELL)
                painter.setPen(QPen(COLOR_CURSOR_BORDER, 1))
                painter.drawRect(rect.adjusted(0, 0, -1, -1))

    def _row_has_data(self, row: int) -> bool:
        for col in range(self.columnCount()):
            item = self.item(row, col)
            if item and item.text().strip():
                return True
        return False

    def _jump_to_data_row(self, direction: int):
        """Jump to previous/next row that contains data."""
        if self.rowCount() <= 0:
            return

        row = self.currentRow()
        col = self.currentColumn()
        if row < 0:
            row = 0
        if col < 0:
            col = 0

        if direction < 0:
            r = row - 1
            while r >= 0 and not self._row_has_data(r):
                r -= 1
            target = max(0, r)
        else:
            last = self._find_last_row()
            r = row + 1
            while r <= last and not self._row_has_data(r):
                r += 1
            target = min(last, r) if last >= 0 else 0

        self.setCurrentCell(target, col)
        item = self.item(target, col)
        if item is not None:
            self.scrollToItem(item)

    def _jump_to_top_row(self):
        """Jump to the first row in the current column (Ctrl+Up)."""
        if self.rowCount() <= 0:
            return
        col = self.currentColumn()
        if col < 0:
            col = 0
        target = 0
        self.setCurrentCell(target, col)
        item = self.item(target, col)
        if item is not None:
            self.scrollToItem(item)
        else:
            self.scrollTo(self.model().index(target, col))

    def _search_positions(self, scope: str) -> List[Tuple[int, int]]:
        scope = (scope or "all").strip().lower()
        if scope not in {"all", "row", "column"}:
            scope = "all"

        if scope == "row":
            row = self.currentRow()
            if row < 0:
                row = 0
            last_col = self._find_last_col()
            if last_col < 0:
                last_col = self.columnCount() - 1
            return [(row, c) for c in range(max(0, last_col) + 1)]

        if scope == "column":
            col = self.currentColumn()
            if col < 0:
                col = 0
            last_row = self._find_last_row()
            if last_row < 0:
                last_row = self.rowCount() - 1
            return [(r, col) for r in range(max(0, last_row) + 1)]

        last_row = self._find_last_row()
        last_col = self._find_last_col()
        if last_row < 0 or last_col < 0:
            return []
        return [(r, c) for r in range(last_row + 1) for c in range(last_col + 1)]

    def find_next_match(self, query: str, scope: str = "all") -> Optional[Tuple[int, int]]:
        """Find next cell containing query text in the requested scope."""
        term = (query or "").strip()
        if not term:
            return None
        positions = self._search_positions(scope)
        if not positions:
            return None

        current = (self.currentRow(), self.currentColumn())
        start = 0
        try:
            idx = positions.index(current)
            start = (idx + 1) % len(positions)
        except ValueError:
            start = 0

        needle = term.casefold()
        for offset in range(len(positions)):
            row, col = positions[(start + offset) % len(positions)]
            item = self.item(row, col)
            text = item.text() if item else ""
            if needle in text.casefold():
                self.setCurrentCell(row, col)
                if item is not None:
                    self.scrollToItem(item)
                else:
                    self.scrollTo(self.model().index(row, col))
                return row, col
        return None
