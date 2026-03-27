"""
Bank Statement Converter Dialog.
Lets the user pick an Excel file, choose sheets, and run the conversion.
"""

import os
import zipfile
import xml.etree.ElementTree as ET
from html.parser import HTMLParser

from PySide6.QtCore import Qt, QThread, QTimer, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from kdl.dialogs.dialog_sizing import create_hint_button, fit_dialog_to_screen
from kdl.styles import accent_button_qss, dialog_qss, themed_button_qss
from kdl.tabular_import import (
    build_filtered_workbook_from_excel,
    detect_source_format,
    iter_csv_rows,
    list_legacy_xls_sheet_names,
    load_workbook_from_source,
)


def _default_browse_dir() -> str:
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    return downloads if os.path.isdir(downloads) else os.path.expanduser("~")


def _fast_xlsx_sheet_names(filepath: str) -> list[str] | None:
    """Read sheet names from an xlsx/xlsm file by parsing xl/workbook.xml directly.
    Returns a list of names or None if the zip approach fails."""
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            with zf.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = tree.findall(".//ns:sheet", ns)
        names = [s.get("name") for s in sheets if s.get("name")]
        return names if names else None
    except Exception:
        return None


class _HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[str]]] = []
        self._table_depth = 0
        self._current_table: list[list[str]] | None = None
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            if self._table_depth == 0:
                self._current_table = []
            self._table_depth += 1
            return

        if self._table_depth != 1:
            return

        if tag == "tr":
            self._current_row = []
        elif tag in ("td", "th"):
            self._current_cell = []
        elif tag == "br" and self._current_cell is not None:
            self._current_cell.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == "table":
            if self._table_depth == 1 and self._current_table:
                if any(row for row in self._current_table):
                    self.tables.append(self._current_table)
                self._current_table = None
            if self._table_depth > 0:
                self._table_depth -= 1
            return

        if self._table_depth != 1:
            return

        if tag in ("td", "th") and self._current_row is not None and self._current_cell is not None:
            text = "".join(self._current_cell)
            text = "\n".join(part.strip() for part in text.splitlines() if part.strip())
            self._current_row.append(text)
            self._current_cell = None
        elif tag == "tr" and self._current_table is not None and self._current_row is not None:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = None

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell.append(data)


def _sheet_safe_name(name: str, fallback: str) -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\\\' else ch for ch in str(name).strip())
    cleaned = cleaned.strip("'")
    return (cleaned or fallback)[:31]


def _load_html_tables(filepath: str) -> list[list[list[str]]]:
    with open(filepath, "rb") as f:
        raw = f.read()

    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    parser = _HTMLTableParser()
    parser.feed(text)
    parser.close()
    return parser.tables


def _build_workbook_from_html(filepath: str):
    import openpyxl

    tables = _load_html_tables(filepath)
    if not tables:
        raise RuntimeError("No HTML tables found in the selected file.")

    wb = openpyxl.Workbook()
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    ws = wb.active
    ws.title = _sheet_safe_name(base_name, "Sheet1")
    for idx, rows in enumerate(tables, start=1):
        for row_vals in rows:
            ws.append(row_vals)
        if idx < len(tables):
            ws.append([])

    return wb


class _SheetLoaderWorker(QThread):
    """Background worker that reads sheet names from an Excel file."""

    sheets_ready = Signal(list)   # list[str]
    load_error = Signal(str)

    def __init__(self, filepath: str):
        super().__init__()
        self.filepath = filepath

    def run(self):
        try:
            ext = os.path.splitext(self.filepath)[1].lower()
            source_kind = detect_source_format(self.filepath)

            # ── CSV: single implicit sheet ──
            if source_kind == "csv":
                name = os.path.splitext(os.path.basename(self.filepath))[0]
                self.sheets_ready.emit([name])
                return

            if source_kind == "html":
                tables = _load_html_tables(self.filepath)
                if not tables:
                    raise RuntimeError("No HTML tables found in the selected file.")
                base_name = os.path.splitext(os.path.basename(self.filepath))[0]
                self.sheets_ready.emit([_sheet_safe_name(base_name, "Sheet1")])
                return

            # ── Fast path: xlsx / xlsm via zipfile (no workbook load) ──
            if ext in (".xlsx", ".xlsm"):
                names = _fast_xlsx_sheet_names(self.filepath)
                if names:
                    self.sheets_ready.emit(names)
                    return

            if ext == ".xls":
                self.sheets_ready.emit(list_legacy_xls_sheet_names(self.filepath))
                return

            # ── Fallback: openpyxl read-only (slower but universal) ──
            import openpyxl
            wb = openpyxl.load_workbook(
                self.filepath, read_only=True, data_only=True, keep_links=False
            )
            try:
                self.sheets_ready.emit(list(wb.sheetnames))
            finally:
                wb.close()

        except Exception as exc:
            self.load_error.emit(str(exc))


class _ConverterWorker(QThread):
    def __init__(self, filepath: str, sheet_names: list[str], skip_contra: bool = True):
        super().__init__()
        self.filepath = filepath
        self.sheet_names = sheet_names if isinstance(sheet_names, list) else [sheet_names]
        self.skip_contra = skip_contra
        self.wb = None
        self.result = None
        self.error_message = ""
        self.save_as_copy = False

    def _convert_loaded_workbook(
        self,
        wb,
        convert_statement,
        ConversionResult,
        _get_or_create_sheet,
        _write_rows_to_sheet,
        _write_audit_header,
        AUDIT_DETAIL_FIRST_ROW,
    ):
        multi = len(self.sheet_names) > 1
        all_output_data = []
        all_audit_rows = []
        all_messages = []
        any_success = False

        for sheet_name in self.sheet_names:
            result = convert_statement(wb, sheet_name, skip_contra=self.skip_contra)
            if result.success:
                any_success = True
                all_output_data.extend(result.output_data)
                prefix = f"[{sheet_name}]\n" if multi else ""
                all_messages.append(f"{prefix}{result.message}")
                if multi:
                    all_audit_rows.extend(result.audit_rows)
            else:
                prefix = f"[{sheet_name}] FAILED\n" if multi else "FAILED\n"
                all_messages.append(f"{prefix}{result.message}")

        if any_success and multi:
            ws_out = _get_or_create_sheet(wb, "Output")
            _write_rows_to_sheet(ws_out, all_output_data)

            ws_audit = _get_or_create_sheet(wb, "Audit_Skipped")
            _write_audit_header(ws_audit)
            for i, row_vals in enumerate(all_audit_rows):
                for c, val in enumerate(row_vals, 1):
                    ws_audit.cell(row=AUDIT_DETAIL_FIRST_ROW + i, column=c, value=val)

        sep = "\n\n" + ("-" * 40) + "\n\n" if multi else ""
        self.result = ConversionResult(
            success=any_success,
            message=sep.join(all_messages),
            output_data=all_output_data,
        )
        if any_success:
            self.wb = wb

    def run(self):
        try:
            import openpyxl
            from kdl.engine.statement_converter import (
                AUDIT_DETAIL_FIRST_ROW,
                ConversionResult,
                _get_or_create_sheet,
                _write_audit_header,
                _write_rows_to_sheet,
                convert_statement,
            )

            source_ext = os.path.splitext(self.filepath)[1].lower()
            source_kind = detect_source_format(self.filepath)

            if source_kind in ("csv", "html"):
                if source_kind == "csv":
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    sheet_name = os.path.splitext(os.path.basename(self.filepath))[0]
                    ws.title = sheet_name
                    for row_vals in iter_csv_rows(self.filepath):
                        ws.append(row_vals)
                    self.sheet_names = [sheet_name]
                else:
                    wb = _build_workbook_from_html(self.filepath)
                    self.sheet_names = [name for name in self.sheet_names if name in wb.sheetnames]
                    if not self.sheet_names:
                        self.sheet_names = list(wb.sheetnames)
                self._convert_loaded_workbook(
                    wb,
                    convert_statement,
                    ConversionResult,
                    _get_or_create_sheet,
                    _write_rows_to_sheet,
                    _write_audit_header,
                    AUDIT_DETAIL_FIRST_ROW,
                )
                return

            keep_vba = source_ext == ".xlsm"
            full_wb = None
            try:
                full_wb = load_workbook_from_source(
                    self.filepath,
                    sheet_names=self.sheet_names,
                    data_only=True,
                    keep_links=False,
                    keep_vba=keep_vba,
                )
                self._convert_loaded_workbook(
                    full_wb,
                    convert_statement,
                    ConversionResult,
                    _get_or_create_sheet,
                    _write_rows_to_sheet,
                    _write_audit_header,
                    AUDIT_DETAIL_FIRST_ROW,
                )
                return
            except Exception as original_exc:
                close_wb = getattr(full_wb, "close", None)
                if callable(close_wb):
                    try:
                        close_wb()
                    except Exception:
                        pass
                if source_ext not in (".xlsx", ".xlsm"):
                    raise
                try:
                    wb = build_filtered_workbook_from_excel(
                        self.filepath,
                        sheet_names=self.sheet_names,
                        data_only=True,
                        keep_links=False,
                    )
                    self.save_as_copy = True
                    self._convert_loaded_workbook(
                        wb,
                        convert_statement,
                        ConversionResult,
                        _get_or_create_sheet,
                        _write_rows_to_sheet,
                        _write_audit_header,
                        AUDIT_DETAIL_FIRST_ROW,
                    )
                    if self.result is not None:
                        self.result.message += (
                            "\n\nLarge workbook fallback used. "
                            "The converted result will be saved as a separate "
                            "_converted.xlsx copy so the original workbook stays unchanged."
                        )
                    return
                except Exception as fallback_exc:
                    raise RuntimeError(
                        f"{original_exc}\n\nFallback conversion also failed:\n{fallback_exc}"
                    ) from fallback_exc
        except Exception as exc:
            if not self.error_message:
                self.error_message = str(exc)


class StatementConverterDialog(QDialog):
    load_into_grid = Signal(list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Bank Statement Converter")
        self.setMinimumWidth(440)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)
        self._worker = None
        self._sheet_loader = None
        self._result = None
        self._wb = None
        self._save_as_copy = False
        self._sheet_checks = []

        from kdl.config_store import get_dark_mode

        self.setStyleSheet(dialog_qss(dark=get_dark_mode()))
        self._build_ui()
        self._fit_to_screen()

    def _release_worker(self):
        worker = self._worker
        self._worker = None
        if worker is None:
            return
        try:
            worker.deleteLater()
        except Exception:
            pass

    def _build_ui(self):
        from kdl.config_store import get_dark_mode

        dark = get_dark_mode()
        primary_btn_qss = accent_button_qss(dark=dark)
        secondary_btn_qss = themed_button_qss(dark=dark)

        outer = QVBoxLayout(self)
        outer.setSpacing(0)
        outer.setContentsMargins(16, 16, 16, 16)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        _content = QWidget()
        layout = QVBoxLayout(_content)
        layout.setSpacing(12)
        layout.setContentsMargins(0, 0, 4, 0)

        intro_row = QHBoxLayout()
        intro_row.setSpacing(8)
        intro = QLabel("Convert bank statement sheets and load the cleaned output into the grid.")
        intro.setObjectName("DialogIntro")
        intro.setWordWrap(True)
        intro_row.addWidget(intro, 1)
        intro_row.addWidget(
            create_hint_button(
                "Choose the source workbook, select the sheets to convert, then review "
                "the summary before loading the output into the grid.",
                label="i",
            )
        )
        layout.addLayout(intro_row)

        file_group = QGroupBox("Source Workbook")
        file_layout = QHBoxLayout(file_group)
        file_layout.setSpacing(10)
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("Choose the bank statement workbook (.xlsx, .xls, .xlsm)...")
        self._file_edit.setReadOnly(True)
        browse_btn = QPushButton("Browse...")
        browse_btn.setMinimumWidth(112)
        browse_btn.setMinimumHeight(38)
        browse_btn.setStyleSheet(secondary_btn_qss)
        browse_btn.clicked.connect(self._browse_file)
        clear_btn = QPushButton("Clear")
        clear_btn.setMinimumWidth(96)
        clear_btn.setMinimumHeight(38)
        clear_btn.setStyleSheet(secondary_btn_qss)
        clear_btn.clicked.connect(self._clear_selected_source)
        file_layout.addWidget(self._file_edit)
        file_layout.addWidget(browse_btn)
        file_layout.addWidget(clear_btn)
        layout.addWidget(file_group)

        sheet_group = QGroupBox("Sheets")
        sheet_outer = QVBoxLayout(sheet_group)
        sheet_outer.setSpacing(8)

        sel_row = QHBoxLayout()
        sel_all_btn = QPushButton("Select All")
        sel_all_btn.setMinimumWidth(108)
        sel_all_btn.setMinimumHeight(36)
        sel_all_btn.setStyleSheet(secondary_btn_qss)
        sel_all_btn.clicked.connect(self._select_all_sheets)
        sel_none_btn = QPushButton("Clear All")
        sel_none_btn.setMinimumWidth(108)
        sel_none_btn.setMinimumHeight(36)
        sel_none_btn.setStyleSheet(secondary_btn_qss)
        sel_none_btn.clicked.connect(self._select_no_sheets)
        sel_row.addWidget(sel_all_btn)
        sel_row.addWidget(sel_none_btn)
        sel_row.addStretch()
        sheet_outer.addLayout(sel_row)

        self._sheet_check_container = QWidget()
        self._sheet_check_layout = QGridLayout(self._sheet_check_container)
        self._sheet_check_layout.setSpacing(8)
        self._sheet_check_layout.setContentsMargins(4, 2, 4, 2)
        self._sheet_scroll = QScrollArea()
        self._sheet_scroll.setWidgetResizable(True)
        self._sheet_scroll.setFrameShape(QFrame.NoFrame)
        self._sheet_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._sheet_scroll.setMinimumHeight(96)
        self._sheet_scroll.setMaximumHeight(220)
        self._sheet_scroll.setWidget(self._sheet_check_container)
        sheet_outer.addWidget(self._sheet_scroll)
        layout.addWidget(sheet_group)

        options_group = QGroupBox("Conversion Options")
        options_layout = QVBoxLayout(options_group)
        self._skip_contra_check = QCheckBox("Skip contra/duplicate matched rows (CONTRA_MATCHED)")
        self._skip_contra_check.setChecked(True)
        self._skip_contra_check.setToolTip(
            "When checked, rows that match as debit/credit pairs on the same reference "
            "and amount are excluded from output. Uncheck to include all rows."
        )
        options_layout.addWidget(self._skip_contra_check)
        layout.addWidget(options_group)

        btn_row = QHBoxLayout()
        self._convert_btn = QPushButton("Convert Workbook")
        self._convert_btn.setEnabled(False)
        self._convert_btn.setMinimumWidth(170)
        self._convert_btn.setMinimumHeight(38)
        self._convert_btn.setStyleSheet(primary_btn_qss)
        self._convert_btn.clicked.connect(self._run_conversion)
        btn_row.addStretch()
        btn_row.addWidget(self._convert_btn)
        layout.addLayout(btn_row)

        result_group = QGroupBox("Conversion Summary")
        result_layout = QVBoxLayout(result_group)
        self._result_text = QTextEdit()
        self._result_text.setReadOnly(True)
        self._result_text.setFixedHeight(128)
        self._result_text.setPlaceholderText("Conversion summary will appear here...")
        result_layout.addWidget(self._result_text)
        layout.addWidget(result_group)

        layout.addStretch()
        scroll.setWidget(_content)
        outer.addWidget(scroll, 1)

        # ── Action row (always visible outside scroll) ──
        action_row = QHBoxLayout()
        action_row.setSpacing(10)
        action_row.setContentsMargins(0, 8, 0, 0)
        self._load_grid_btn = QPushButton("Load Output to Grid")
        self._load_grid_btn.setMinimumWidth(158)
        self._load_grid_btn.setMinimumHeight(38)
        self._load_grid_btn.setEnabled(False)
        self._load_grid_btn.setStyleSheet(primary_btn_qss)
        self._load_grid_btn.setToolTip(
            "Load the converted Output sheet into the TNT DL grid and save the workbook."
        )
        self._load_grid_btn.clicked.connect(self._load_into_grid)

        self._close_btn = QPushButton("Close")
        self._close_btn.setMinimumWidth(104)
        self._close_btn.setMinimumHeight(38)
        self._close_btn.setStyleSheet(secondary_btn_qss)
        self._close_btn.clicked.connect(self.accept)

        action_row.addWidget(self._load_grid_btn)
        action_row.addStretch()
        action_row.addWidget(self._close_btn)
        outer.addLayout(action_row)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            _default_browse_dir(),
            "All Supported (*.xlsx *.xls *.xlsm *.csv *.html *.htm);;Excel Files (*.xlsx *.xls *.xlsm);;CSV Files (*.csv);;HTML Files (*.html *.htm);;All Files (*)",
        )
        if not path:
            return
        self._reset_source_state(clear_path=True)
        self._file_edit.setText(path)
        self._start_sheet_loader(path)

    def _clear_selected_source(self):
        self._reset_source_state(clear_path=True)

    def _reset_source_state(self, clear_path: bool):
        self._sheet_loader = None
        self._worker = None
        self._result = None
        self._wb = None
        self._save_as_copy = False
        for cb in self._sheet_checks:
            self._sheet_check_layout.removeWidget(cb)
            cb.deleteLater()
        self._sheet_checks.clear()
        self._convert_btn.setEnabled(False)
        self._convert_btn.setText("Convert Workbook")
        self._load_grid_btn.setEnabled(False)
        self._result_text.clear()
        if clear_path:
            self._file_edit.clear()

    def _start_sheet_loader(self, filepath: str):
        self._sheet_loader = None
        self._worker = None
        for cb in self._sheet_checks:
            self._sheet_check_layout.removeWidget(cb)
            cb.deleteLater()
        self._sheet_checks.clear()
        self._convert_btn.setEnabled(False)
        self._result = None
        self._wb = None
        self._save_as_copy = False
        self._load_grid_btn.setEnabled(False)
        self._result_text.setPlainText("Reading sheets\u2026")

        loader = _SheetLoaderWorker(filepath)
        loader.sheets_ready.connect(self._on_sheets_ready)
        loader.load_error.connect(self._on_sheets_error)
        loader.finished.connect(loader.deleteLater)
        self._sheet_loader = loader
        loader.start()

    def _on_sheets_ready(self, sheet_names: list):
        loader = self.sender()
        if loader is not self._sheet_loader:
            return
        self._result_text.clear()
        cols = 3
        for i, name in enumerate(sheet_names):
            row, col = divmod(i, cols)
            cb = QCheckBox(name)
            is_generated = name == "Output" or name.startswith("Audit_")
            cb.setChecked(not is_generated)
            cb.stateChanged.connect(self._update_convert_btn)
            self._sheet_check_layout.addWidget(cb, row, col)
            self._sheet_checks.append(cb)
        self._update_convert_btn()
        QTimer.singleShot(0, self._fit_to_screen)
        if sum(1 for cb in self._sheet_checks if cb.isChecked()) == 1:
            QTimer.singleShot(0, self._run_conversion)

    def _on_sheets_error(self, message: str):
        loader = self.sender()
        if loader is not self._sheet_loader:
            return
        self._result_text.clear()
        QMessageBox.warning(self, "File Error", f"Could not open file:\n{message}")

    def _select_all_sheets(self):
        for cb in self._sheet_checks:
            cb.setChecked(True)

    def _select_no_sheets(self):
        for cb in self._sheet_checks:
            cb.setChecked(False)

    def _update_convert_btn(self):
        has_file = bool(self._file_edit.text().strip())
        has_selection = any(cb.isChecked() for cb in self._sheet_checks)
        self._convert_btn.setEnabled(has_file and has_selection)

    def _run_conversion(self):
        filepath = self._file_edit.text().strip()
        selected = [cb.text() for cb in self._sheet_checks if cb.isChecked()]
        if not filepath or not selected:
            return

        self._convert_btn.setEnabled(False)
        self._convert_btn.setText("Converting...")
        self._result_text.setPlainText("Processing...")
        self._load_grid_btn.setEnabled(False)
        self._result = None
        self._wb = None
        self._save_as_copy = False

        self._worker = _ConverterWorker(
            filepath,
            selected,
            skip_contra=self._skip_contra_check.isChecked(),
        )
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _on_worker_finished(self):
        worker = self.sender()
        if worker is not self._worker:
            try:
                worker.deleteLater()
            except Exception:
                pass
            return

        self._convert_btn.setEnabled(True)
        self._convert_btn.setText("Convert Workbook")

        if worker.error_message:
            self._result = None
            self._wb = None
            self._save_as_copy = False
            self._result_text.setPlainText(f"ERROR:\n{worker.error_message}")
            QMessageBox.critical(self, "Conversion Error", worker.error_message)
            self._release_worker()
            return

        result = worker.result
        self._result = result
        self._save_as_copy = bool(getattr(worker, "save_as_copy", False))
        if result is not None and result.success:
            self._wb = worker.wb

        self._release_worker()

        if result is None:
            self._result_text.setPlainText("ERROR:\nConversion did not return a result.")
            return

        self._result_text.setPlainText(result.message)
        if result.success:
            self._load_grid_btn.setEnabled(bool(result.output_data))
            self._result_text.append(
                '\nClick "Load Output to Grid" to load the result and save the workbook.'
            )

    def _load_into_grid(self):
        if not (self._result and self._result.output_data):
            return

        if self._wb is not None:
            filepath = self._file_edit.text().strip()
            source_ext = os.path.splitext(filepath)[1].lower()
            save_path = filepath
            if self._save_as_copy or source_ext in (".xls", ".csv", ".html", ".htm"):
                save_path = os.path.splitext(filepath)[0] + "_converted.xlsx"
            saved_name = None
            try:
                self._wb.save(save_path)
                saved_name = os.path.basename(save_path)
            except PermissionError:
                try:
                    import win32com.client

                    abs_path = os.path.abspath(save_path)
                    xl = win32com.client.GetActiveObject("Excel.Application")
                    for wb_com in list(xl.Workbooks):
                        if os.path.abspath(wb_com.FullName).lower() == abs_path.lower():
                            wb_com.Save()
                            wb_com.Close(SaveChanges=False)
                            break
                except Exception:
                    pass
                try:
                    self._wb.save(save_path)
                    saved_name = os.path.basename(save_path)
                except Exception as exc:
                    self._result_text.append(f"\nSave failed: {exc}")
            except Exception as exc:
                self._result_text.append(f"\nSave failed: {exc}")

            if saved_name:
                if save_path != filepath:
                    self._result_text.append(
                        f"Saved to: {saved_name} (source file kept unchanged)"
                    )
                else:
                    self._result_text.append(f"Saved to: {saved_name}")

            close_wb = getattr(self._wb, "close", None)
            if callable(close_wb):
                try:
                    close_wb()
                except Exception:
                    pass

        self.load_into_grid.emit(self._result.output_data)
        self.accept()

    def closeEvent(self, event):
        if self._worker is not None and self._worker.isRunning():
            QMessageBox.warning(
                self,
                "Conversion In Progress",
                "Wait for the conversion to finish before closing this window.",
            )
            event.ignore()
            return
        self._release_worker()
        super().closeEvent(event)

    def _fit_to_screen(self):
        fit_dialog_to_screen(
            self,
            min_width=560,
            min_height=420,
            preferred_width=720,
            wide_width=860,
            margin_width=72,
            margin_height=72,
            extra_hint_width=32,
            extra_hint_height=28,
        )
