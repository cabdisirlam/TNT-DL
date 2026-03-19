"""
IFMIS Financial Statements Report Generator Dialog.

Imports a Notes Excel file, processes it entirely in Python using openpyxl,
and produces a formatted 5-sheet Excel workbook (Notes, Performance, Position,
Net Assets, Cash Flow).
"""

import os
import xml.etree.ElementTree as ET
import zipfile

from PySide6.QtCore import Qt, QThread, QTimer, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from kdl.styles import accent_button_qss, dialog_qss


def _default_dir() -> str:
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    return downloads if os.path.isdir(downloads) else os.path.expanduser("~")


def _fast_xlsx_sheet_names(filepath: str):
    """Read sheet names from an xlsx/xlsm file without loading the workbook."""
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            with zf.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = tree.findall(".//ns:sheet", ns)
        names  = [s.get("name") for s in sheets if s.get("name")]
        return names if names else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Background workers
# ---------------------------------------------------------------------------

class _SheetLoaderWorker(QThread):
    """Reads sheet names from an Excel file without loading cell data."""

    sheets_ready = Signal(list)
    load_error   = Signal(str)

    def __init__(self, filepath: str):
        super().__init__()
        self.filepath = filepath

    def run(self):
        try:
            ext = os.path.splitext(self.filepath)[1].lower()

            # ── Fast path: xlsx/xlsm via zipfile ──
            if ext in (".xlsx", ".xlsm"):
                names = _fast_xlsx_sheet_names(self.filepath)
                if names:
                    self.sheets_ready.emit(names)
                    return

            # ── .xls: try xlrd first ──
            if ext == ".xls":
                try:
                    import xlrd
                    wb = xlrd.open_workbook(self.filepath, on_demand=True)
                    try:
                        names = wb.sheet_names()
                    finally:
                        release = getattr(wb, "release_resources", None)
                        if callable(release):
                            release()
                    self.sheets_ready.emit(list(names))
                    return
                except Exception:
                    pass

            # ── Fallback: openpyxl ──
            import openpyxl
            wb = openpyxl.load_workbook(
                self.filepath, read_only=True, data_only=True, keep_links=False
            )
            try:
                self.sheets_ready.emit(list(wb.sheetnames))
            finally:
                wb.close()
        except Exception as exc:
            self.load_error.emit(str(exc))


class _ReportWorker(QThread):
    """Loads the selected sheet and runs the IFMIS report engine."""

    def __init__(self, filepath: str, sheet_name: str):
        super().__init__()
        self.filepath   = filepath
        self.sheet_name = sheet_name
        self.wb_out     = None   # openpyxl Workbook on success
        self.message    = ""
        self.success    = False

    def run(self):
        tmp_path = None
        excel_app = None
        excel_wb = None
        try:
            import openpyxl
            import traceback

            source_ext = os.path.splitext(self.filepath)[1].lower()
            load_path = self.filepath

            # ── Convert legacy .xls to temp .xlsx via win32com ──
            if source_ext == ".xls":
                try:
                    import tempfile
                    import pythoncom
                    import win32com.client
                    pythoncom.CoInitialize()
                    excel_app = win32com.client.DispatchEx("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    excel_wb = excel_app.Workbooks.Open(
                        os.path.abspath(self.filepath),
                        UpdateLinks=0,
                        ReadOnly=True,
                    )
                    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
                    os.close(fd)
                    os.unlink(tmp_path)
                    excel_wb.SaveAs(tmp_path, 51)
                    load_path = tmp_path
                except Exception as exc:
                    self.success = False
                    self.message = (
                        "Could not convert this legacy .xls workbook. "
                        "Open it in Excel and save as .xlsx, or make sure "
                        "Excel is installed.\n\n" + str(exc)
                    )
                    return
                finally:
                    if excel_wb is not None:
                        try:
                            excel_wb.Close(False)
                        except Exception:
                            pass
                    if excel_app is not None:
                        try:
                            excel_app.Quit()
                        except Exception:
                            pass
                    try:
                        import pythoncom
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass

            wb = openpyxl.load_workbook(load_path, data_only=True, keep_links=False)
            ws = wb[self.sheet_name]
            from kdl.engine.ifmis_report import generate_ifmis_report
            result = generate_ifmis_report(ws)
            wb.close()
            self.success = result.success
            self.message = result.message
            self.wb_out  = result.workbook
        except Exception as exc:
            import traceback
            self.success = False
            self.message = f"{exc}\n\n{traceback.format_exc()}"
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass


# ---------------------------------------------------------------------------
# Dialog
# ---------------------------------------------------------------------------

class FinancialReportDialog(QDialog):
    """Generate IFMIS Financial Statements from a Notes Excel file."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("IFMIS Financial Statements Generator")
        self.setMinimumWidth(560)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)
        self._worker       = None
        self._sheet_loader = None
        self._sheet_checks: list[QCheckBox] = []
        self._wb_out       = None

        from kdl.config_store import get_dark_mode
        self.setStyleSheet(dialog_qss(get_dark_mode()))
        self._build_ui()
        self._fit_to_screen()

    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── Notes file ──
        file_group  = QGroupBox("Notes Excel File")
        file_layout = QHBoxLayout(file_group)
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("Select the IFMIS Notes Excel file...")
        self._file_edit.setReadOnly(True)
        browse_btn = QPushButton("Browse...")
        browse_btn.setFixedWidth(90)
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(self._file_edit)
        file_layout.addWidget(browse_btn)
        layout.addWidget(file_group)

        # ── Sheet selector ──
        sheet_group = QGroupBox("Notes Sheet (select one)")
        sheet_outer = QVBoxLayout(sheet_group)
        sheet_outer.setSpacing(6)
        self._sheet_check_container = QWidget()
        self._sheet_check_layout    = QGridLayout(self._sheet_check_container)
        self._sheet_check_layout.setSpacing(4)
        self._sheet_check_layout.setContentsMargins(4, 2, 4, 2)
        sheet_outer.addWidget(self._sheet_check_container)
        layout.addWidget(sheet_group)

        # ── Generate button ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._generate_btn = QPushButton("  Generate  ")
        self._generate_btn.setEnabled(False)
        from kdl.config_store import get_dark_mode
        self._generate_btn.setStyleSheet(accent_button_qss(get_dark_mode()))
        self._generate_btn.clicked.connect(self._run_generation)
        btn_row.addWidget(self._generate_btn)
        layout.addLayout(btn_row)

        # ── Result ──
        result_group  = QGroupBox("Result")
        result_layout = QVBoxLayout(result_group)
        self._result_text = QTextEdit()
        self._result_text.setReadOnly(True)
        self._result_text.setFixedHeight(120)
        self._result_text.setPlaceholderText("Generation result will appear here...")
        result_layout.addWidget(self._result_text)
        layout.addWidget(result_group)

        # ── Save + Close ──
        action_row = QHBoxLayout()
        self._save_btn = QPushButton("Save Financial Statements...")
        self._save_btn.setEnabled(False)
        self._save_btn.setToolTip("Save the generated 5-sheet workbook to an Excel file")
        self._save_btn.clicked.connect(self._save_output)
        self._close_btn = QPushButton("Close")
        self._close_btn.clicked.connect(self.accept)
        action_row.addWidget(self._save_btn)
        action_row.addStretch()
        action_row.addWidget(self._close_btn)
        layout.addLayout(action_row)

    # ------------------------------------------------------------------
    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Notes Excel File",
            _default_dir(),
            "Excel Files (*.xlsx *.xlsm *.xls);;All Files (*)",
        )
        if not path:
            return
        self._file_edit.setText(path)
        self._start_sheet_loader(path)

    def _start_sheet_loader(self, filepath: str):
        if self._sheet_loader is not None and self._sheet_loader.isRunning():
            self._sheet_loader.quit()
            self._sheet_loader.wait()
        self._sheet_loader = None

        for cb in self._sheet_checks:
            self._sheet_check_layout.removeWidget(cb)
            cb.deleteLater()
        self._sheet_checks.clear()
        self._generate_btn.setEnabled(False)
        self._save_btn.setEnabled(False)
        self._wb_out = None
        self._result_text.setPlainText("Reading sheets…")

        loader = _SheetLoaderWorker(filepath)
        loader.sheets_ready.connect(self._on_sheets_ready)
        loader.load_error.connect(self._on_sheets_error)
        self._sheet_loader = loader
        loader.start()

    def _on_sheets_ready(self, sheet_names: list):
        self._result_text.clear()
        cols = 3
        for i, name in enumerate(sheet_names):
            row, col = divmod(i, cols)
            cb = QCheckBox(name)
            cb.setChecked(i == 0)
            cb.stateChanged.connect(self._on_sheet_checked)
            self._sheet_check_layout.addWidget(cb, row, col)
            self._sheet_checks.append(cb)
        self._update_generate_btn()
        if len(sheet_names) == 1:
            QTimer.singleShot(0, self._run_generation)

    def _on_sheet_checked(self, state):
        """Radio-button behaviour: only one sheet selected at a time."""
        if state == Qt.Checked:
            sender = self.sender()
            for cb in self._sheet_checks:
                if cb is not sender:
                    cb.setChecked(False)
        self._update_generate_btn()

    def _on_sheets_error(self, message: str):
        self._result_text.clear()
        QMessageBox.warning(self, "File Error", f"Could not open file:\n{message}")

    def _update_generate_btn(self):
        has_file = bool(self._file_edit.text().strip())
        has_sel  = any(cb.isChecked() for cb in self._sheet_checks)
        self._generate_btn.setEnabled(has_file and has_sel)

    def _get_selected_sheet(self):
        for cb in self._sheet_checks:
            if cb.isChecked():
                return cb.text()
        return None

    # ------------------------------------------------------------------
    def _run_generation(self):
        filepath = self._file_edit.text().strip()
        sheet    = self._get_selected_sheet()
        if not filepath or not sheet:
            return

        self._generate_btn.setEnabled(False)
        self._generate_btn.setText("Generating...")
        self._save_btn.setEnabled(False)
        self._wb_out = None
        self._result_text.setPlainText("Processing...")

        self._worker = _ReportWorker(filepath, sheet)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _on_worker_finished(self):
        worker       = self._worker
        self._worker = None

        self._generate_btn.setEnabled(True)
        self._generate_btn.setText("  Generate  ")

        if worker is None:
            return

        if worker.success:
            self._wb_out = worker.wb_out
            self._result_text.setPlainText(worker.message)
            self._save_btn.setEnabled(True)
            self._result_text.append('\nClick "Save Financial Statements..." to save the output.')
        else:
            self._result_text.setPlainText(f"ERROR:\n{worker.message}")
            QMessageBox.critical(self, "Generation Failed", worker.message[:600])

        try:
            worker.deleteLater()
        except Exception:
            pass

    # ------------------------------------------------------------------
    def _save_output(self):
        if self._wb_out is None:
            return
        source  = self._file_edit.text().strip()
        base    = os.path.splitext(source)[0] if source else ""
        suggest = (base + "_FinancialStatements.xlsx") if base else ""

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Financial Statements As",
            suggest or _default_dir(),
            "Excel Workbook (*.xlsx);;All Files (*)",
        )
        if not path:
            return
        try:
            self._wb_out.save(path)
            self._result_text.append(f"\nSaved to: {os.path.basename(path)}")
            QMessageBox.information(
                self, "Saved",
                f"Financial statements saved to:\n{path}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Save Failed", str(exc))

    # ------------------------------------------------------------------
    def closeEvent(self, event):
        if self._worker is not None and self._worker.isRunning():
            QMessageBox.warning(
                self,
                "Generation In Progress",
                "Please wait for the generation to finish before closing.",
            )
            event.ignore()
            return
        super().closeEvent(event)

    def _fit_to_screen(self):
        from PySide6.QtGui import QGuiApplication
        screen = self.screen() or QGuiApplication.primaryScreen()
        if screen:
            ag = screen.availableGeometry()
            self.resize(
                min(self.sizeHint().width() + 40, ag.width() - 80),
                min(self.sizeHint().height() + 40, ag.height() - 80),
            )
