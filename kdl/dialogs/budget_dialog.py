"""
GOK IFMIS Statement of Budget Execution - Processing Dialog.

Imports an IFMIS budget Excel file (1-3 sheets), runs the budget
processor engine on the selected sheets, and saves the formatted
output workbook as budget.xlsx.
"""

import os
import xml.etree.ElementTree as ET
import zipfile

from PySide6.QtCore import Qt, QThread, QTimer, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from kdl.styles import accent_button_qss, dialog_qss


def _default_dir() -> str:
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    return downloads if os.path.isdir(downloads) else os.path.expanduser("~")


def _fast_xlsx_sheet_names(filepath: str):
    """Read sheet names from an xlsx/xlsm without loading the workbook."""
    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            with zf.open("xl/workbook.xml") as f:
                tree = ET.parse(f)
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        sheets = tree.findall(".//ns:sheet", ns)
        names = [s.get("name") for s in sheets if s.get("name")]
        return names if names else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Background workers
# ---------------------------------------------------------------------------

class _SheetLoaderWorker(QThread):
    """Reads sheet names from an Excel file without loading cell data."""

    sheets_ready = Signal(list)
    load_error   = Signal(str)

    def __init__(self, filepath: str):
        super().__init__()
        self.filepath = filepath

    def run(self):
        try:
            ext = os.path.splitext(self.filepath)[1].lower()
            if ext in (".xlsx", ".xlsm"):
                names = _fast_xlsx_sheet_names(self.filepath)
                if names:
                    self.sheets_ready.emit(names)
                    return
            import openpyxl
            wb = openpyxl.load_workbook(
                self.filepath, read_only=True, data_only=True, keep_links=False
            )
            try:
                self.sheets_ready.emit(list(wb.sheetnames))
            finally:
                wb.close()
        except Exception as exc:
            self.load_error.emit(str(exc))


class _BudgetWorker(QThread):
    """Loads selected sheets and runs the budget processor engine."""

    def __init__(self, filepath: str, sheet_names: list):
        super().__init__()
        self.filepath    = filepath
        self.sheet_names = sheet_names
        self.wb_out      = None   # openpyxl Workbook on success
        self.message     = ""
        self.success     = False

    def run(self):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(
                self.filepath, data_only=True, keep_links=False
            )
            from kdl.engine.budget_processor import process_budget_sheets
            result = process_budget_sheets(wb, self.sheet_names)
            wb.close()
            self.success = result.success
            self.message = result.message
            self.wb_out  = result.workbook
        except Exception as exc:
            import traceback
            self.success = False
            self.message = f"{exc}\n\n{traceback.format_exc()}"


# ---------------------------------------------------------------------------
# Dialog
# ---------------------------------------------------------------------------

class BudgetDialog(QDialog):
    """Process IFMIS Statement of Budget Execution sheets into a formatted workbook."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GOK IFMIS Budget Processor")
        self.setMinimumWidth(560)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)
        self._worker       = None
        self._sheet_loader = None
        self._sheet_checks: list[QCheckBox] = []
        self._wb_out       = None

        from kdl.config_store import get_dark_mode
        self.setStyleSheet(dialog_qss(get_dark_mode()))
        self._build_ui()
        self._fit_to_screen()

    # ------------------------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── Budget Excel file ──
        file_group  = QGroupBox("Budget Excel File")
        file_layout = QHBoxLayout(file_group)
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText(
            "Select the IFMIS Statement of Budget Execution file..."
        )
        self._file_edit.setReadOnly(True)
        browse_btn = QPushButton("Browse...")
        browse_btn.setFixedWidth(90)
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(self._file_edit)
        file_layout.addWidget(browse_btn)
        layout.addWidget(file_group)

        # ── Sheet selector (multi-select, up to 3 sheets) ──
        sheet_group = QGroupBox("Sheets to Process")
        sheet_outer = QVBoxLayout(sheet_group)
        sheet_outer.setSpacing(6)

        sel_row = QHBoxLayout()
        sel_all_btn = QPushButton("Select All")
        sel_all_btn.setFixedWidth(90)
        sel_all_btn.clicked.connect(self._select_all_sheets)
        sel_none_btn = QPushButton("Select None")
        sel_none_btn.setFixedWidth(90)
        sel_none_btn.clicked.connect(self._select_no_sheets)
        sel_row.addWidget(sel_all_btn)
        sel_row.addWidget(sel_none_btn)
        sel_row.addStretch()
        sheet_outer.addLayout(sel_row)

        self._sheet_check_container = QWidget()
        self._sheet_check_layout    = QGridLayout(self._sheet_check_container)
        self._sheet_check_layout.setSpacing(4)
        self._sheet_check_layout.setContentsMargins(4, 2, 4, 2)
        sheet_outer.addWidget(self._sheet_check_container)
        layout.addWidget(sheet_group)

        # ── Process button ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._process_btn = QPushButton("  Process Budget  ")
        self._process_btn.setEnabled(False)
        from kdl.config_store import get_dark_mode
        self._process_btn.setStyleSheet(accent_button_qss(get_dark_mode()))
        self._process_btn.clicked.connect(self._run_processing)
        btn_row.addWidget(self._process_btn)
        layout.addLayout(btn_row)

        # ── Result ──
        result_group  = QGroupBox("Result")
        result_layout = QVBoxLayout(result_group)
        self._result_text = QTextEdit()
        self._result_text.setReadOnly(True)
        self._result_text.setFixedHeight(120)
        self._result_text.setPlaceholderText("Processing result will appear here...")
        result_layout.addWidget(self._result_text)
        layout.addWidget(result_group)

        # ── Save + Close ──
        action_row = QHBoxLayout()
        self._save_btn = QPushButton("Save Budget...")
        self._save_btn.setEnabled(False)
        self._save_btn.setToolTip(
            "Save the processed budget workbook as an Excel file"
        )
        self._save_btn.clicked.connect(self._save_output)
        self._close_btn = QPushButton("Close")
        self._close_btn.clicked.connect(self.accept)
        action_row.addWidget(self._save_btn)
        action_row.addStretch()
        action_row.addWidget(self._close_btn)
        layout.addLayout(action_row)

    # ------------------------------------------------------------------
    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select IFMIS Budget Excel File",
            _default_dir(),
            "Excel Files (*.xlsx *.xlsm *.xls);;All Files (*)",
        )
        if not path:
            return
        self._file_edit.setText(path)
        self._start_sheet_loader(path)

    def _start_sheet_loader(self, filepath: str):
        if self._sheet_loader is not None and self._sheet_loader.isRunning():
            self._sheet_loader.quit()
            self._sheet_loader.wait()
        self._sheet_loader = None

        for cb in self._sheet_checks:
            self._sheet_check_layout.removeWidget(cb)
            cb.deleteLater()
        self._sheet_checks.clear()
        self._process_btn.setEnabled(False)
        self._save_btn.setEnabled(False)
        self._wb_out = None
        self._result_text.setPlainText("Reading sheets\u2026")

        loader = _SheetLoaderWorker(filepath)
        loader.sheets_ready.connect(self._on_sheets_ready)
        loader.load_error.connect(self._on_sheets_error)
        loader.finished.connect(loader.deleteLater)
        self._sheet_loader = loader
        loader.start()

    def _on_sheets_ready(self, sheet_names: list):
        self._result_text.clear()
        cols = 3
        for i, name in enumerate(sheet_names):
            row, col = divmod(i, cols)
            cb = QCheckBox(name)
            cb.setChecked(True)
            cb.stateChanged.connect(self._update_process_btn)
            self._sheet_check_layout.addWidget(cb, row, col)
            self._sheet_checks.append(cb)
        self._update_process_btn()
        # Auto-run if only one sheet
        if len(sheet_names) == 1:
            QTimer.singleShot(0, self._run_processing)

    def _on_sheets_error(self, message: str):
        self._result_text.clear()
        QMessageBox.warning(self, "File Error", f"Could not open file:\n{message}")

    def _select_all_sheets(self):
        for cb in self._sheet_checks:
            cb.setChecked(True)

    def _select_no_sheets(self):
        for cb in self._sheet_checks:
            cb.setChecked(False)

    def _update_process_btn(self):
        has_file = bool(self._file_edit.text().strip())
        has_sel  = any(cb.isChecked() for cb in self._sheet_checks)
        self._process_btn.setEnabled(has_file and has_sel)

    # ------------------------------------------------------------------
    def _run_processing(self):
        filepath = self._file_edit.text().strip()
        selected = [cb.text() for cb in self._sheet_checks if cb.isChecked()]
        if not filepath or not selected:
            return

        self._process_btn.setEnabled(False)
        self._process_btn.setText("Processing\u2026")
        self._save_btn.setEnabled(False)
        self._wb_out = None
        self._result_text.setPlainText("Processing\u2026")

        self._worker = _BudgetWorker(filepath, selected)
        self._worker.finished.connect(self._on_worker_finished)
        self._worker.start()

    def _on_worker_finished(self):
        worker       = self._worker
        self._worker = None

        self._process_btn.setEnabled(True)
        self._process_btn.setText("  Process Budget  ")

        if worker is None:
            return

        if worker.success:
            self._wb_out = worker.wb_out
            self._result_text.setPlainText(worker.message)
            self._save_btn.setEnabled(True)
            self._result_text.append('\nClick "Save Budget..." to save the output.')
        else:
            self._result_text.setPlainText(f"ERROR:\n{worker.message}")
            QMessageBox.critical(self, "Processing Failed", worker.message[:600])

        try:
            worker.deleteLater()
        except Exception:
            pass

    # ------------------------------------------------------------------
    def _save_output(self):
        if self._wb_out is None:
            return
        source  = self._file_edit.text().strip()
        base    = os.path.splitext(source)[0] if source else ""
        suggest = (base + "_budget.xlsx") if base else "budget.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Budget As",
            suggest,
            "Excel Workbook (*.xlsx);;All Files (*)",
        )
        if not path:
            return
        try:
            self._wb_out.save(path)
            self._result_text.append(f"\nSaved to: {os.path.basename(path)}")
            QMessageBox.information(
                self, "Saved",
                f"Budget saved to:\n{path}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Save Failed", str(exc))

    # ------------------------------------------------------------------
    def closeEvent(self, event):
        if self._worker is not None and self._worker.isRunning():
            QMessageBox.warning(
                self,
                "Processing In Progress",
                "Please wait for processing to finish before closing.",
            )
            event.ignore()
            return
        super().closeEvent(event)

    def _fit_to_screen(self):
        from PySide6.QtGui import QGuiApplication
        screen = self.screen() or QGuiApplication.primaryScreen()
        if screen:
            ag = screen.availableGeometry()
            self.resize(
                min(self.sizeHint().width() + 40, ag.width() - 80),
                min(self.sizeHint().height() + 40, ag.height() - 80),
            )
